import datetime as dt
from datetime import timedelta
from io import BytesIO
from pathlib import Path
import base64
import difflib
import html
import os
import re

import openpyxl
import folium
from folium.plugins import HeatMap, MarkerCluster
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components
from auth_utils import (
    disable_user,
    load_auth_users_from_secrets,
    load_permissions_from_secrets,
    read_auth_store_summary,
    set_user_password,
    set_user_permissions,
    verify_password,
)
from audit_utils import append_audit_event, read_audit_events
from style_utils import apply_global_styles


USUARIOS_APP = load_auth_users_from_secrets()


def get_usuarios_app():
    return load_auth_users_from_secrets()


def get_permissoes():
    permissoes = load_permissions_from_secrets(PERMISSOES_PADRAO)
    usuarios_app = get_usuarios_app()
    for username in set(permissoes.keys()) | set(usuarios_app.keys()):
        if username not in permissoes:
            permissoes[username] = []
        if "*" not in permissoes[username] and PAGINA_PRODUTIVIDADE not in permissoes[username]:
            permissoes[username].append(PAGINA_PRODUTIVIDADE)
        if "*" not in permissoes[username] and "Produtividade UPAs" not in permissoes[username]:
            permissoes[username].append("Produtividade UPAs")
        if "*" not in permissoes[username] and "SAMU" not in permissoes[username]:
            permissoes[username].append("SAMU")
    return permissoes

TEMPO_SESSAO_HORAS = 8
BUILD_TAG = "PM-2026-04-27-08"
PAGINA_PRODUTIVIDADE = "Produtividade UPAs"
ROTULO_PRODUTIVIDADE = "Produtividade Médica UPAs"
PAGINA_ADMIN_ACESSOS = "Administracao de Acessos"
PAGINA_HEATMAP = "Mapa de Calor"
PAGINA_MAPA_TERRITORIAL = "Mapa Territorial"
PAGINAS_LIBERADAS_GLOBAL = {
    "SAMU",
    PAGINA_PRODUTIVIDADE,
    ROTULO_PRODUTIVIDADE,
    "Produtividade Upas",
    PAGINA_HEATMAP,
    PAGINA_MAPA_TERRITORIAL,
}


def get_local_build_stamp():
    try:
        mtime = Path(__file__).stat().st_mtime
        return dt.datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        return "indisponivel"


LOCAL_BUILD_STAMP = get_local_build_stamp()

PERMISSOES_PADRAO = {
    "admin": ["*"],
    "vittor": ["*"],
    "wendel": ["*"],
    "guilherme": ["*"],
    "denis": ["*"],
    "prefeitura": [
        "UPA Luziânia",
        "UPA Jardim Ingá",
        "SAMU",
        "HMJI",
        PAGINA_MAPA_TERRITORIAL,
        "Atenção Secundária",
        "Saúde Mental",
        "Atenção Primária",
        "Gestão de Pessoas",
        "Metas do Plano",
        PAGINA_PRODUTIVIDADE,
    ],
}

PERMISSOES = get_permissoes()


def render_login():
    base_dir = Path(__file__).resolve().parent
    logo_patris = base_dir / "assets" / "patris.png"
    logo_prefeitura = base_dir / "assets" / "prefeitura.png"

    def _logo_b64(path):
        if not path.exists():
            return ""
        return base64.b64encode(path.read_bytes()).decode("utf-8")

    logo_patris_b64 = _logo_b64(logo_patris)
    logo_prefeitura_b64 = _logo_b64(logo_prefeitura)

    st.markdown(
        """
    <style>
    .login-header-row {
        max-width: 1580px;
        margin: 14px auto 18px auto;
        display: grid;
        grid-template-columns: 230px 1fr 230px;
        gap: 20px;
        align-items: center;
    }
    .login-side-logo {
        display: flex;
        align-items: center;
        justify-content: center;
        min-height: 120px;
        overflow: hidden;
    }
    .login-side-logo img {
        max-width: none;
        max-height: none;
        width: auto;
        object-fit: contain;
        display: block;
    }
    .login-side-logo-patris img {
        width: min(100%, 260px) !important;
        max-width: 260px !important;
        transform: none;
        object-fit: contain;
        display: block;
    }
    .login-side-logo-patris {
        min-height: 140px;
    }
    .login-side-logo-prefeitura img {
        object-fit: contain;
        display: block;
    }
    .login-side-logo-prefeitura {
        min-height: 130px;
    }
    .login-hero {
        border-radius: 22px;
        background: linear-gradient(135deg, #0d7d57 0%, #16a56f 45%, #1fb77f 100%);
        border: 1px solid rgba(255,255,255,0.22);
        box-shadow: 0 16px 34px rgba(2, 6, 23, 0.20);
        padding: 16px 20px 18px 20px;
        text-align: center;
    }
    .login-kicker {
        display: inline-block;
        font-size: 11px;
        text-transform: uppercase;
        letter-spacing: 1.4px;
        font-weight: 800;
        color: #d1fae5;
        margin-bottom: 6px;
    }
    .login-heading {
        color: #f8fafc;
        font-size: 44px;
        font-weight: 900;
        line-height: 1.05;
        margin-bottom: 8px;
    }
    .login-sub {
        color: rgba(241, 245, 249, 0.96);
        font-size: 29px;
        margin: 0 0 12px 0;
    }
    .login-pills {
        display: flex;
        justify-content: center;
        gap: 10px;
        flex-wrap: wrap;
    }
    .login-pill {
        font-size: 12px;
        color: #f8fafc;
        padding: 6px 12px;
        border-radius: 999px;
        background: rgba(255, 255, 255, 0.16);
        border: 1px solid rgba(255, 255, 255, 0.20);
    }
    .login-title {
        text-align: center;
        margin: 8px 0 4px 0;
        font-size: 44px;
        font-weight: 900;
        color: #e2e8f0;
    }
    .login-subtitle {
        text-align: center;
        font-size: 22px;
        color: #94a3b8;
        margin-bottom: 14px;
    }
    @media (max-width: 900px) {
        .login-header-row {
            grid-template-columns: 1fr;
            gap: 10px;
        }
        .login-side-logo-patris img {
            transform: scale(1.15);
        }
        .login-side-logo-prefeitura img {
            transform: scale(1.4);
        }
        .login-side-logo-patris,
        .login-side-logo-prefeitura {
            min-height: 120px;
        }
        .login-heading {
            font-size: 30px;
        }
        .login-sub {
            font-size: 18px;
        }
        .login-title {
            font-size: 32px;
        }
        .login-subtitle {
            font-size: 16px;
        }
    }
    [data-testid="stTextInput"] input {
        height: 46px;
        border-radius: 12px;
        border: 1px solid #334155;
        background: rgba(15, 23, 42, 0.72);
        color: #f8fafc;
    }
    [data-testid="stTextInput"] label {
        font-weight: 700;
        color: #e2e8f0;
    }
    [data-testid="stButton"] button {
        height: 46px;
        border-radius: 12px;
        border: none;
        background: linear-gradient(130deg, #0f766e 0%, #0ea5e9 100%);
        color: #f8fafc;
        font-weight: 800;
        letter-spacing: 0.2px;
        box-shadow: 0 10px 24px rgba(14, 116, 144, 0.35);
    }
    </style>
    """,
        unsafe_allow_html=True,
    )

    patris_html = (
        f'<img src="data:image/png;base64,{logo_patris_b64}" alt="Instituto Patris" style="width:240px; max-width:100%; height:auto;" />'
        if logo_patris_b64
        else '<strong style="color:#e2e8f0; font-size:18px;">Patris</strong>'
    )
    prefeitura_html = (
        f'<img src="data:image/png;base64,{logo_prefeitura_b64}" alt="Prefeitura de Luziania" style="width:198px; max-width:198px; height:auto;" />'
        if logo_prefeitura_b64
        else '<strong style="color:#e2e8f0; font-size:18px;">Prefeitura</strong>'
    )

    st.markdown(
        f"""
    <div class="login-header-row">
        <div class="login-side-logo login-side-logo-patris">{patris_html}</div>
        <div class="login-hero">
            <div class="login-kicker">PATRIS • GESTAO MUNICIPAL</div>
            <div class="login-heading">Painel de Gestao Patris</div>
            <p class="login-sub">Gestao estrategica da producao assistencial e desempenho operacional</p>
            <div class="login-pills">
                <span class="login-pill">Pagina: UPA Luziania</span>
                <span class="login-pill">Periodo: Mar/26</span>
                <span class="login-pill">Atualizado em: {globals().get("LOCAL_BUILD_STAMP", "indisponivel")}</span>
            </div>
        </div>
        <div class="login-side-logo login-side-logo-prefeitura">{prefeitura_html}</div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="login-title">🔐 Acesso ao Painel</div>', unsafe_allow_html=True)
    st.markdown('<div class="login-subtitle">Informe usuario e senha para continuar</div>', unsafe_allow_html=True)
    st.caption("Login build: LG-2026-04-27-12")

    usuarios_app = get_usuarios_app()
    if not usuarios_app:
        st.error("Autenticação não configurada. Defina auth.users no secrets.toml ou configure auth_store persistente.")
        st.stop()

    col1, col2, col3 = st.columns([1, 1.8, 1])
    with col2:
        usuario = st.text_input("Usuário")
        senha = st.text_input("Senha", type="password")
        entrar = st.button("Entrar", width="stretch")

    if entrar:
        usuarios_app = get_usuarios_app()
        usuario_ok = usuario in usuarios_app
        senha_ok = usuario_ok and verify_password(senha, usuarios_app[usuario])

        if usuario_ok and senha_ok:
            st.session_state["autenticado"] = True
            st.session_state["usuario_logado"] = usuario
            st.session_state["login_em"] = dt.datetime.now()
            st.session_state["expira_em"] = dt.datetime.now() + timedelta(hours=TEMPO_SESSAO_HORAS)
            append_audit_event(
                event="login_success",
                user=usuario,
                session_id=st.session_state.get("session_id", ""),
                details="Login validado",
            )
            st.rerun()
        else:
            st.error("Usuário ou senha inválidos.")


def check_login():
    if "autenticado" not in st.session_state:
        st.session_state["autenticado"] = False

    if "session_id" not in st.session_state:
        st.session_state["session_id"] = f"sess-{dt.datetime.now().strftime('%Y%m%d%H%M%S%f')}"

    if st.session_state["autenticado"]:
        expira_em = st.session_state.get("expira_em")

        if expira_em and dt.datetime.now() > expira_em:
            append_audit_event(
                event="session_expired",
                user=st.session_state.get("usuario_logado", ""),
                page=st.session_state.get("pagina_selecionada", ""),
                session_id=st.session_state.get("session_id", ""),
                details="Sessao expirada por tempo limite",
            )
            st.session_state["autenticado"] = False
            st.session_state["usuario_logado"] = None
            st.session_state["login_em"] = None
            st.session_state["expira_em"] = None

    if not st.session_state["autenticado"]:
        render_login()
        st.stop()


st.set_page_config(page_title="Painel de Gestão Patris", page_icon="📊", layout="wide")
check_login()

BASE_DIR = Path(__file__).resolve().parent
ASSETS_DIR = BASE_DIR / "assets"

LOGO_PATRIS = ASSETS_DIR / "patris.png"
LOGO_SIDEBAR = ASSETS_DIR / "logosemfundo.png"
LOGO_PREFEITURA = ASSETS_DIR / "prefeitura.png"
BACKGROUND_IMG = ASSETS_DIR / "background.png"


def usuario_pode_ver_pagina(usuario, pagina):
    if pagina == PAGINA_ADMIN_ACESSOS:
        return usuario == "admin"

    if pagina == "Auditoria de Acesso":
        return usuario == "admin"

    pagina_norm = normalize_text(pagina)
    liberadas_norm = {normalize_text(p) for p in PAGINAS_LIBERADAS_GLOBAL}
    if pagina_norm in liberadas_norm:
        return True

    permissoes = get_permissoes().get(usuario, [])
    permissoes_norm = {normalize_text(p) for p in permissoes}
    if pagina_norm in permissoes_norm:
        return True

    return "*" in permissoes or pagina in permissoes


def image_to_base64(path):
    if not path.exists():
        return ""
    return base64.b64encode(path.read_bytes()).decode("utf-8")


BACKGROUND_BASE64 = image_to_base64(BACKGROUND_IMG)
LOGO_PATRIS_BASE64 = image_to_base64(LOGO_PATRIS)
LOGO_SIDEBAR_BASE64 = image_to_base64(LOGO_SIDEBAR) or LOGO_PATRIS_BASE64
apply_global_styles(st, BACKGROUND_BASE64)


MESES = [
    "MARCO.26", "ABRIL.26", "MAIO.26", "JUNHO.26",
    "JULHO.26", "AGOSTO.26", "SETEMBRO.26", "OUTUBRO.26",
    "NOVEMBRO.26", "DEZEMBRO.26", "JANEIRO.27", "FEVEREIRO.27"
]

MESES_LABEL = {
    "MARCO.26": "Mar/26",
    "ABRIL.26": "Abr/26",
    "MAIO.26": "Mai/26",
    "JUNHO.26": "Jun/26",
    "JULHO.26": "Jul/26",
    "AGOSTO.26": "Ago/26",
    "SETEMBRO.26": "Set/26",
    "OUTUBRO.26": "Out/26",
    "NOVEMBRO.26": "Nov/26",
    "DEZEMBRO.26": "Dez/26",
    "JANEIRO.27": "Jan/27",
    "FEVEREIRO.27": "Fev/27"
}


def default_previous_month_selection():
    month_name_to_number = {
        "JANEIRO": 1,
        "FEVEREIRO": 2,
        "MARCO": 3,
        "ABRIL": 4,
        "MAIO": 5,
        "JUNHO": 6,
        "JULHO": 7,
        "AGOSTO": 8,
        "SETEMBRO": 9,
        "OUTUBRO": 10,
        "NOVEMBRO": 11,
        "DEZEMBRO": 12,
    }

    month_abbr = {
        1: "Jan",
        2: "Fev",
        3: "Mar",
        4: "Abr",
        5: "Mai",
        6: "Jun",
        7: "Jul",
        8: "Ago",
        9: "Set",
        10: "Out",
        11: "Nov",
        12: "Dez",
    }

    today = dt.datetime.now().date()
    first_day_current_month = today.replace(day=1)
    previous_month_date = first_day_current_month - timedelta(days=1)
    previous_month_label = f"{month_abbr[previous_month_date.month]}/{str(previous_month_date.year)[-2:]}"

    options = [MESES_LABEL[m] for m in MESES]
    if previous_month_label in options:
        return [previous_month_label]

    available_dates = []
    for month_key in MESES:
        month_name, year_suffix = month_key.split(".")
        month_number = month_name_to_number.get(normalize_text(month_name))
        if month_number is None:
            continue
        year = 2000 + int(year_suffix)
        available_dates.append((dt.date(year, month_number, 1), MESES_LABEL.get(month_key, month_key)))

    if not available_dates:
        return options

    available_dates.sort(key=lambda x: x[0])
    target_date = previous_month_date.replace(day=1)
    candidates = [label for date_value, label in available_dates if date_value <= target_date]

    if candidates:
        return [candidates[-1]]

    return [available_dates[0][1]]


RISK_COLORS = {
    "NÃO URGENTE (AZUL)": "#1E3A8A",
    "POUCO URGENTE (VERDE)": "#16A34A",
    "URGENTE (AMARELO)": "#EAB308",
    "MUITO URGENTE (LARANJA)": "#F97316",
    "EMERGÊNCIA (VERMELHO)": "#DC2626",
    "NÃO INFORMADO": "#6B7280",
}

_plot_counter = 0


def _strip_html_text(value):
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _get_plot_title_subtitle(fig):
    title_obj = getattr(fig.layout, "title", None)
    raw_title = getattr(title_obj, "text", None) if title_obj is not None else None
    if not raw_title:
        return "", ""

    parts = str(raw_title).split("<br>", 1)
    title = _strip_html_text(parts[0]) if parts else ""
    subtitle = _strip_html_text(parts[1]) if len(parts) > 1 else ""
    return title, subtitle


def _to_number(value):
    try:
        if value is None or pd.isna(value):
            return None
        return float(value)
    except Exception:
        return None


def _is_inverse_indicator(indicator_hint):
    text = normalize_text(indicator_hint) or ""
    inverse_tokens = [
        "TEMPO DE ESPERA",
        "TEMPO MEDIO",
        "TEMPO MÉDIO",
        "TEMPO DE PERMANENCIA",
        "TEMPO DE PERMANÊNCIA",
        "OBITO",
        "ÓBITO",
        "ACIDENTE DE TRABALHO",
        "ABSENTEISMO",
        "ABSENTEÍSMO",
        "TURNOVER",
    ]
    return any(token in text for token in inverse_tokens)


def _status_threshold(indicator_hint, inverse_logic=False):
    """Define limiar percentual de alerta por contexto do indicador."""
    text = normalize_text(indicator_hint) or ""

    if inverse_logic:
        return 0.02

    strict_tokens = [
        "GASTO",
        "FINANCEIRO",
        "CUSTO",
        "VALOR",
        "DESPESA",
    ]
    if any(token in text for token in strict_tokens):
        return 0.05

    return 0.03


def _chart_exec_status(fig, indicator_hint=""):
    """Calcula um status executivo simples com base na tendencia dos dois ultimos pontos."""
    inverse_logic = _is_inverse_indicator(indicator_hint)
    threshold = _status_threshold(indicator_hint, inverse_logic=inverse_logic)

    for trace in fig.data:
        trace_name_obj = getattr(trace, "name", None)
        trace_name = str(trace_name_obj).upper() if trace_name_obj is not None else ""
        if "META" in trace_name:
            continue

        xs_raw = getattr(trace, "x", None)
        ys_values = getattr(trace, "y", None)

        if xs_raw is None or ys_values is None:
            continue

        try:
            xs = list(xs_raw)
            ys_raw = list(ys_values)
        except Exception:
            continue

        if not xs or not ys_raw:
            continue

        ys = [_to_number(v) for v in ys_raw]
        points = [(x, y) for x, y in zip(xs, ys) if y is not None]
        if len(points) < 2:
            continue

        labels = [str(x) for x, _ in points]
        values = [y for _, y in points]

        is_time_like = all(lbl in MESES_LABEL.values() for lbl in labels)
        if not is_time_like:
            continue

        atual = values[-1]
        anterior = values[-2]

        if anterior == 0:
            if atual == 0:
                return {
                    "label": "Sem movimentacao",
                    "tone": "neutral",
                    "detail": None,
                }
            return {
                "label": "Entrada de valor",
                "tone": "info",
                "detail": None,
            }

        delta = (atual - anterior) / abs(anterior)
        delta_txt = f"{delta * 100:+.1f}%".replace(".", ",")

        if inverse_logic:
            if delta <= -threshold:
                return {
                    "label": "Em melhora",
                    "tone": "success",
                    "detail": delta_txt,
                }
            if delta >= threshold:
                return {
                    "label": "Em piora",
                    "tone": "danger",
                    "detail": delta_txt,
                }
            return {
                "label": "Estavel",
                "tone": "warning",
                "detail": delta_txt,
            }

        if delta >= threshold:
            return {
                "label": "Em alta",
                "tone": "success",
                "detail": delta_txt,
            }
        if delta <= -threshold:
            return {
                "label": "Em queda",
                "tone": "danger",
                "detail": delta_txt,
            }
        return {
            "label": "Estavel",
            "tone": "warning",
            "detail": delta_txt,
        }

    return {
        "label": "Consolidado",
        "tone": "info",
        "detail": None,
    }


def plot(fig, prefix="grafico", show_status_chip=True):
    global _plot_counter
    _plot_counter += 1

    title, subtitle = _get_plot_title_subtitle(fig)
    if title:
        indicator_hint = f"{title} {subtitle}".strip()
        status = _chart_exec_status(fig, indicator_hint=indicator_hint)
        no_status_titles = {
            "Pacientes recepcionados por mês",
            "Atendimentos médicos vs meta",
            "Produção diária do SAMU",
        }
        show_status_chip = show_status_chip and (title not in no_status_titles)
        status_label = html.escape(status["label"])
        status_detail = f" {html.escape(status['detail'])}" if status.get("detail") else ""
        subtitle_text = subtitle if subtitle else ""
        subtitle_html = ""
        if subtitle_text:
            subtitle_safe = html.escape(subtitle_text)
            subtitle_html = f'<div class="chart-exec-subtitle">{subtitle_safe}</div>'
        title_safe = html.escape(title)
        chip_html = ""
        if show_status_chip:
            chip_palette = {
                "success": ("#16A34A", "#FFFFFF"),
                "warning": ("#D97706", "#FFFFFF"),
                "danger": ("#DC2626", "#FFFFFF"),
                "neutral": ("#64748B", "#FFFFFF"),
                "info": ("#0F6CBD", "#FFFFFF"),
            }
            chip_bg, chip_fg = chip_palette.get(status.get("tone"), ("#0F6CBD", "#FFFFFF"))
            if title in {"Consultas Médicas", "Nível Superior (Exceto Médico)"}:
                detail_raw = str(status.get("detail") or "").strip()
                is_negative = detail_raw.startswith("-") or detail_raw.startswith("−")
                chip_bg, chip_fg = ("#DC2626", "#FFFFFF") if is_negative else ("#16A34A", "#FFFFFF")
            chip_html = (
                f'<div class="chart-exec-chip chart-exec-chip-{status["tone"]}" '
                f'style="display:inline-flex;align-items:center;justify-content:center;white-space:nowrap;'
                f'font-size:11px;font-weight:800;color:{chip_fg};background:{chip_bg};'
                f'border:1px solid {chip_bg};border-radius:999px;padding:6px 11px;line-height:1;">'
                f'{status_label}{status_detail}</div>'
            )

        header_html = (
            '<div class="chart-exec-header">'
            '<div class="chart-exec-row">'
            '<div>'
            f'<div class="chart-exec-title" style="font-size:18px;font-weight:800;color:#0B1220;letter-spacing:-0.3px;line-height:1.22;"><strong style="font-weight:800">{title_safe}</strong></div>'
            f'{subtitle_html}'
            '</div>'
            f'{chip_html}'
            '</div>'
            '</div>'
        )

        st.markdown(header_html, unsafe_allow_html=True)

        current_margin = getattr(getattr(fig.layout, "margin", None), "t", None)
        new_margin_top = max(28, int(current_margin) - 44) if current_margin is not None else 34
        fig.update_layout(title_text="", margin=dict(t=new_margin_top))

    st.plotly_chart(fig, width="stretch", key=f"{prefix}_{_plot_counter}")

def local_excel_path():
    base = Path(__file__).parent
    names = [
        "dashboard_municipio.xlsx",
        "DASH BORD NOVO MUNICIPIO ATUALIZADO.xlsx",
        "DASH BORD NOVO MUNICIPIO ATUALIZADO.xlsm",
    ]
    for name in names:
        p = base / name
        if p.exists():
            return p
    found = sorted(base.glob("*.xlsx")) + sorted(base.glob("*.xlsm"))
    return found[0] if found else None

def normalize_value(v):
    if v in (None, ""):
        return None

    # Excel pode entregar duração como timedelta
    if isinstance(v, dt.timedelta):
        return v.total_seconds() / 3600  # horas, sem arredondar

    # Excel pode entregar horário como dt.time
    if isinstance(v, dt.time):
        return v.hour + (v.minute / 60) + (v.second / 3600)  # horas, sem arredondar

    if isinstance(v, str):
        if v.startswith("#DIV/0"):
            return None

        vv = v.strip()

        # interpreta textos tipo 01:30 ou 01:30:00 como horas
        if ":" in vv:
            try:
                partes = vv.split(":")
                if len(partes) == 2:
                    h, m = partes
                    s = 0
                elif len(partes) == 3:
                    h, m, s = partes
                else:
                    h = m = s = None

                if h is not None:
                    return float(h) + float(m) / 60 + float(s) / 3600
            except Exception:
                pass

        # Normaliza valores numéricos textuais no padrão BR/EN (milhar e decimal)
        vv = vv.replace("R$", "").replace(" ", "")
        if "." in vv and "," in vv:
            # Ex.: 1.234,56 -> 1234.56
            vv = vv.replace(".", "").replace(",", ".")
        elif "," in vv:
            # Ex.: 1234,56 -> 1234.56
            vv = vv.replace(",", ".")
        else:
            # Ex.: 1.234.567 (milhar) -> 1234567
            if vv.count(".") > 1:
                vv = vv.replace(".", "")

        try:
            return float(vv)
        except Exception:
            return v.strip()

    if isinstance(v, (int, float)):
        return float(v)

    return v


def normalize_text(value):
    if value is None:
        return None

    import unicodedata
    import re

    text = str(value).strip().upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text).strip()

    return text if text else None


def row_values(ws, r, n=14):
    return [ws.cell(r, c).value for c in range(1, n + 1)]


def is_month_row(vals):
    months = [normalize_text(v) for v in vals[2:14] if v is not None]
    return len(months) >= 3 and all(m in MESES for m in months)


def parse_sheet(ws, sheet_name):
    rows = []
    unidade = str(ws["A2"].value).strip() if ws["A2"].value else sheet_name
    unidade_norm = normalize_text(unidade)

    painel = None
    painel_norm = None
    meses = None

    labels_especiais = {
        "META",
        "MEDIA DIARIA",
        "MÉDIA DIÁRIA",
        "MEDIA DIARIA",
    }

    for r in range(1, ws.max_row + 1):
        vals = row_values(ws, r)
        a, b = vals[0], vals[1]

        if is_month_row(vals):
            meses = [
                normalize_text(ws.cell(r, c).value) if ws.cell(r, c).value is not None else None
                for c in range(3, 15)
            ]
            continue

        if not any(v is not None for v in vals[2:14]):
            continue

        a_str = a.strip() if isinstance(a, str) else None
        b_str = b.strip() if isinstance(b, str) else None

        a_norm = normalize_text(a_str)
        b_norm = normalize_text(b_str)

        if a_norm == "INDICADOR":
            continue

        if a_norm and a_norm not in labels_especiais and b_norm:
            painel = a_str
            painel_norm = a_norm
            serie = b_str
        elif a_norm and a_norm not in labels_especiais and not b_norm:
            painel = a_str
            painel_norm = a_norm
            serie = a_str
        elif a_norm in labels_especiais and painel:
            serie = a_str
        elif b_norm and painel:
            serie = b_str
        else:
            continue

        serie_norm = normalize_text(serie)

        for i, c in enumerate(range(3, 15)):
            mes = meses[i] if meses and i < len(meses) else None
            rows.append({
                "aba": sheet_name,
                "unidade": unidade,
                "unidade_norm": unidade_norm,
                "painel": painel,
                "painel_norm": painel_norm,
                "serie": serie,
                "serie_norm": serie_norm,
                "mes": mes,
                "mes_label": MESES_LABEL.get(mes, mes),
                "valor": normalize_value(ws.cell(r, c).value),
            })

    df = pd.DataFrame(rows)

    if not df.empty:
        df["valor_num"] = pd.to_numeric(df["valor"], errors="coerce")
        df["mes"] = pd.Categorical(df["mes"], categories=MESES, ordered=True)
        df = df.sort_values(["unidade_norm", "painel_norm", "serie_norm", "mes"])
    return df


def _local_file_mtime():
    """Retorna o timestamp de modificação do Excel local (para invalidar cache automaticamente)."""
    p = local_excel_path()
    return p.stat().st_mtime if p else 0


@st.cache_data(show_spinner=False)
def load_workbook_data(file_bytes=None, _mtime=None):
    if file_bytes is None:
        path = local_excel_path()
        if not path:
            return pd.DataFrame(), None
        wb = openpyxl.load_workbook(path, data_only=True)
        source_name = path.name
    else:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        source_name = "upload.xlsx"

    sheet_order = [
        "INDICADORES UPA LUZIÂNIA",
        "INDICADORES UPA JARDIM INGÁ",
        "INDICADORES HMJI",
        "INDICADORES ATENÇÃO SECUNDÁRIA",
        "INDICADORES SAÚDE MENTAL",
        "INDICADORES ATENÇÃO PRIMÁRIA",
        "INDICADORES RH"
    ]

    frames = []
    for s in sheet_order:
        if s in wb.sheetnames:
            part = parse_sheet(wb[s], s)
            if not part.empty:
                frames.append(part)

    if not frames:
        return pd.DataFrame(), source_name

    data = pd.concat(frames, ignore_index=True)
    return data, source_name


@st.cache_data(show_spinner=False)
def load_metas_data(file_bytes=None, _mtime=None):
    colunas_padrao = [
        "categoria",
        "categoria_norm",
        "mes",
        "mes_label",
        "executado",
        "meta",
        "executado_total",
        "meta_total",
        "executado_total_geral",
    ]

    if file_bytes is None:
        path = local_excel_path()
        if not path:
            return pd.DataFrame(columns=colunas_padrao)
        wb = openpyxl.load_workbook(path, data_only=True)
    else:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    nome_aba = "METAS DO PLANO DE TRABALHO"
    if nome_aba not in wb.sheetnames:
        return pd.DataFrame(columns=colunas_padrao)

    ws = wb[nome_aba]

    rows = []
    meses = None
    categoria_atual = None
    total_geral_por_mes = {}

    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 16)]

        linha_meses = [normalize_text(v) for v in vals[2:14] if v is not None]
        if len(linha_meses) >= 3 and all(m in MESES for m in linha_meses):
            meses = [normalize_text(v) if v is not None else None for v in vals[2:14]]
            continue

        col_b = vals[1]
        col_b_norm = normalize_text(col_b)

        if not col_b_norm:
            continue

        if col_b_norm == "TOTAL GERAL":
            for i, c in enumerate(range(3, 15)):  # C:N
                mes = meses[i] if meses and i < len(meses) else None
                if mes is None:
                    continue

                valor = normalize_value(ws.cell(r, c).value)
                valor_num = pd.to_numeric(pd.Series([valor]), errors="coerce").iloc[0]
                if pd.notna(valor_num):
                    total_geral_por_mes[mes] = float(valor_num)
            continue

        # linha da categoria = executado
        if col_b_norm != "META":
            categoria_atual = str(col_b).strip()

            for i, c in enumerate(range(3, 15)):  # C:N
                mes = meses[i] if meses and i < len(meses) else None
                valor = normalize_value(ws.cell(r, c).value)
                valor_num = pd.to_numeric(pd.Series([valor]), errors="coerce").iloc[0]

                rows.append({
                    "categoria": categoria_atual,
                    "categoria_norm": normalize_text(categoria_atual),
                    "mes": mes,
                    "mes_label": MESES_LABEL.get(mes, mes),
                    "executado": float(valor_num) if pd.notna(valor_num) else 0.0,
                    "meta": None,
                    "executado_total": None,
                    "meta_total": None,
                })
            continue

        # linha META = meta
        if col_b_norm == "META" and categoria_atual and meses:
            for i, c in enumerate(range(3, 15)):  # C:N
                mes = meses[i] if i < len(meses) else None
                valor = normalize_value(ws.cell(r, c).value)
                valor_num = pd.to_numeric(pd.Series([valor]), errors="coerce").iloc[0]

                rows.append({
                    "categoria": categoria_atual,
                    "categoria_norm": normalize_text(categoria_atual),
                    "mes": mes,
                    "mes_label": MESES_LABEL.get(mes, mes),
                    "executado": None,
                    "meta": float(valor_num) if pd.notna(valor_num) else 0.0,
                    "executado_total": None,
                    "meta_total": None,
                })

    # Leitura deterministica do TOTAL GERAL na linha 18 (colunas C:N),
    # conforme layout da planilha de Metas informado pelo usuario.
    linha_total_geral = 18
    if ws.max_row >= linha_total_geral:
        for i, c in enumerate(range(3, 15)):  # C:N
            mes = None
            if meses and i < len(meses):
                mes = meses[i]
            elif i < len(MESES):
                mes = MESES[i]

            if mes is None:
                continue

            valor_linha_18 = normalize_value(ws.cell(linha_total_geral, c).value)
            valor_linha_18_num = pd.to_numeric(pd.Series([valor_linha_18]), errors="coerce").iloc[0]
            if pd.notna(valor_linha_18_num):
                total_geral_por_mes[mes] = float(valor_linha_18_num)

    df = pd.DataFrame(rows, columns=colunas_padrao)

    if df.empty:
        return pd.DataFrame(columns=colunas_padrao)

    df["executado"] = pd.to_numeric(df["executado"], errors="coerce")
    df["meta"] = pd.to_numeric(df["meta"], errors="coerce")

    # evita bug do groupby com categorical
    df["mes"] = df["mes"].astype(str)
    df.loc[df["mes"].isin(["None", "nan"]), "mes"] = None

    resumo = (
        df.pivot_table(
            index=["categoria", "categoria_norm", "mes", "mes_label"],
            values=["executado", "meta"],
            aggfunc={"executado": "sum", "meta": "max"},
            dropna=False,
        )
        .reset_index()
    )

    resumo["mes_ord"] = resumo["mes"].apply(lambda x: MESES.index(x) if x in MESES else 999)
    resumo = resumo.sort_values(["categoria_norm", "mes_ord"]).drop(columns=["mes_ord"])

    totais = (
        resumo.groupby(["categoria", "categoria_norm"], dropna=False)[["executado", "meta"]]
        .sum(min_count=1)
        .reset_index()
        .rename(columns={
            "executado": "executado_total",
            "meta": "meta_total",
        })
    )

    resumo = resumo.merge(
        totais,
        on=["categoria", "categoria_norm"],
        how="left"
    )

    resumo["executado_total_geral"] = resumo["mes"].map(total_geral_por_mes)

    return resumo[colunas_padrao].reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_metas_total_geral_map(file_bytes=None, _mtime=None):
    """Retorna TOTAL GERAL por mês (Mar/26..Fev/27) lendo diretamente a linha 18, colunas C:N."""
    if file_bytes is None:
        path = local_excel_path()
        if not path:
            return {}
        wb = openpyxl.load_workbook(path, data_only=True)
    else:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    nome_aba = "METAS DO PLANO DE TRABALHO"
    if nome_aba not in wb.sheetnames:
        return {}

    ws = wb[nome_aba]
    linha_total_geral = 18
    if ws.max_row < linha_total_geral:
        return {}

    total_geral_map = {}
    for i, c in enumerate(range(3, 15)):  # C:N
        mes_key = MESES[i] if i < len(MESES) else None
        if mes_key is None:
            continue

        valor = normalize_value(ws.cell(linha_total_geral, c).value)
        valor_num = pd.to_numeric(pd.Series([valor]), errors="coerce").iloc[0]
        total_geral_map[MESES_LABEL.get(mes_key, mes_key)] = float(valor_num) if pd.notna(valor_num) else 0.0

    return total_geral_map

@st.cache_data(show_spinner=False)
def load_financeiro_data(file_bytes=None, _mtime=None):
    colunas = [
        "grupo",
        "grupo_norm",
        "fornecedor",
        "fornecedor_norm",
        "mes",
        "mes_label",
        "valor",
        "valor_num",
    ]

    if file_bytes is None:
        path = local_excel_path()
        if not path:
            return pd.DataFrame(columns=colunas)
        wb = openpyxl.load_workbook(path, data_only=True)
    else:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    nome_aba = "Financeiro"
    if nome_aba not in wb.sheetnames:
        return pd.DataFrame(columns=colunas)

    ws = wb[nome_aba]

    rows = []
    meses = None
    grupo_atual = None

    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 16)]  # A:O

        linha_meses = [normalize_text(v) for v in vals[2:14] if v is not None]
        if len(linha_meses) >= 3 and all(m in MESES for m in linha_meses):
            meses = [normalize_text(v) if v is not None else None for v in vals[2:14]]
            continue

        col_a = vals[0]
        col_b = vals[1]

        col_a_norm = normalize_text(col_a)
        col_b_norm = normalize_text(col_b)

        if not any(v not in (None, "") for v in vals[2:14]):
            continue

        # linha de grupo / seção
        if col_a_norm and col_a_norm != "TOTAL" and not col_b_norm:
            grupo_atual = str(col_a).strip()
            continue

        # ignora linha TOTAL geral do bloco
        if col_b_norm == "TOTAL":
            continue

        fornecedor = str(col_b).strip() if col_b else None
        if not fornecedor or not meses:
            continue

        for i, c in enumerate(range(3, 15)):  # C:N
            mes = meses[i] if i < len(meses) else None
            valor = normalize_value(ws.cell(r, c).value)
            valor_num = pd.to_numeric(pd.Series([valor]), errors="coerce").iloc[0]

            rows.append({
                "grupo": grupo_atual,
                "grupo_norm": normalize_text(grupo_atual),
                "fornecedor": fornecedor,
                "fornecedor_norm": normalize_text(fornecedor),
                "mes": mes,
                "mes_label": MESES_LABEL.get(mes, mes),
                "valor": valor,
                "valor_num": float(valor_num) if pd.notna(valor_num) else 0.0,
            })

    df = pd.DataFrame(rows, columns=colunas)

    if df.empty:
        return pd.DataFrame(columns=colunas)

    df["mes"] = pd.Categorical(df["mes"], categories=MESES, ordered=True)
    df = df.sort_values(["grupo_norm", "fornecedor_norm", "mes"]).reset_index(drop=True)
    return df


@st.cache_data(show_spinner=False)
def load_mapa_territorial_data(file_bytes=None, _mtime=None):
    """Carrega GEO_UNIDADES e COLABORADORES para o mapa territorial."""
    if file_bytes is None:
        path = local_excel_path()
        if not path:
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), ["Arquivo Excel local não encontrado."]
        xls = pd.ExcelFile(path)
    else:
        xls = pd.ExcelFile(BytesIO(file_bytes))

    erros = []
    if "GEO_UNIDADES" not in xls.sheet_names:
        erros.append("A aba GEO_UNIDADES não foi encontrada na planilha.")
    if "COLABORADORES" not in xls.sheet_names:
        erros.append("A aba COLABORADORES não foi encontrada na planilha.")
    if erros:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), erros

    geo_raw = pd.read_excel(xls, sheet_name="GEO_UNIDADES")
    colab_raw = pd.read_excel(xls, sheet_name="COLABORADORES")

    def _pick_col(df_cols, aliases):
        cols_norm = {normalize_text(c): c for c in df_cols}
        for alias in aliases:
            if alias in cols_norm:
                return cols_norm[alias]
        for col in df_cols:
            col_norm = normalize_text(col) or ""
            if any(alias in col_norm for alias in aliases):
                return col
        return None

    geo_unid_col = _pick_col(geo_raw.columns, {"UNIDADE"})
    geo_lat_col = _pick_col(geo_raw.columns, {"LATITUDE", "LAT"})
    geo_lon_col = _pick_col(geo_raw.columns, {"LONGITUDE", "LONG", "LON"})
    geo_tipo_col = _pick_col(geo_raw.columns, {"TIPO"})

    if not geo_unid_col or not geo_lat_col or not geo_lon_col:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), [
            "A aba GEO_UNIDADES precisa das colunas: unidade, latitude e longitude.",
        ]

    geo = geo_raw.copy()
    geo = geo.rename(columns={
        geo_unid_col: "unidade",
        geo_lat_col: "latitude",
        geo_lon_col: "longitude",
        geo_tipo_col: "tipo" if geo_tipo_col else "tipo",
    })
    if "tipo" not in geo.columns:
        geo["tipo"] = "Não informado"

    geo["unidade"] = geo["unidade"].astype(str).str.strip()
    geo["unidade_norm"] = geo["unidade"].map(normalize_text)
    geo["tipo"] = geo["tipo"].fillna("Não informado").astype(str).str.strip()

    geo["latitude"] = (
        geo["latitude"].astype(str).str.replace(",", ".", regex=False)
    )
    geo["longitude"] = (
        geo["longitude"].astype(str).str.replace(",", ".", regex=False)
    )
    geo["latitude"] = pd.to_numeric(geo["latitude"], errors="coerce")
    geo["longitude"] = pd.to_numeric(geo["longitude"], errors="coerce")
    geo = geo.dropna(subset=["unidade_norm", "latitude", "longitude"]).copy()

    # Corrige casos evidentes de lat/lon invertidos em bases externas.
    swapped_mask = geo["latitude"].abs().gt(33) & geo["longitude"].abs().lt(33)
    if swapped_mask.any():
        geo.loc[swapped_mask, ["latitude", "longitude"]] = geo.loc[
            swapped_mask, ["longitude", "latitude"]
        ].to_numpy()

    # Correcoes pontuais validadas no Google Maps para unidades com coordenada digitada incorretamente.
    geo_overrides = {
        "UBS JARDIM INGA - UNIDADE BASICA DE SAUD": (-16.1428497, -47.9505777),
        "UBSF - TRES VENDAS": (-16.3136192, -48.0148513),
    }
    override_mask = geo["unidade_norm"].isin(geo_overrides)
    if override_mask.any():
        geo.loc[override_mask, "latitude"] = geo.loc[override_mask, "unidade_norm"].map(
            lambda key: geo_overrides.get(key, (None, None))[0]
        )
        geo.loc[override_mask, "longitude"] = geo.loc[override_mask, "unidade_norm"].map(
            lambda key: geo_overrides.get(key, (None, None))[1]
        )

    geo = geo.drop_duplicates(subset=["unidade_norm"], keep="first")

    colab_nome_col = _pick_col(colab_raw.columns, {"COLABORADOR", "NOME"})
    colab_cpf_col = _pick_col(colab_raw.columns, {"CPF"})
    colab_cargo_col = _pick_col(colab_raw.columns, {"CARGO"})
    colab_regime_col = _pick_col(colab_raw.columns, {"REGIME DE TRABALHO", "REGIME"})
    colab_crm_col = _pick_col(colab_raw.columns, {"CRM"})
    colab_adm_col = _pick_col(colab_raw.columns, {"ADM", "ADMISSAO", "ADMISSÃO", "DATA ADMISSAO", "DATA ADMISSAO"})
    colab_situacao_col = _pick_col(colab_raw.columns, {"SITUACAO", "SITUAÇÃO", "STATUS"})

    def _slug_token(value):
        txt = normalize_text(value) or ""
        txt = re.sub(r"[^A-Z0-9]+", "_", txt).strip("_").lower()
        return txt or "mes"

    def _fmt_adm(value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return ""
        try:
            dt_val = pd.to_datetime(value, errors="coerce")
            if pd.notna(dt_val):
                return dt_val.strftime("%d/%m/%Y")
        except Exception:
            pass
        return str(value).strip()

    def _to_money(value):
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return pd.NA
        parsed = normalize_value(value)
        num = pd.to_numeric(pd.Series([parsed]), errors="coerce").iloc[0]
        return float(num) if pd.notna(num) else pd.NA

    proventos_meta = []
    used_tokens = set()
    for col in colab_raw.columns:
        col_norm = normalize_text(col) or ""
        if not col_norm.startswith("PROVENTOS"):
            continue
        mes_label = re.sub(r"(?i)^\s*proventos\s*", "", str(col)).strip()
        if not mes_label:
            parts = col_norm.split(" ", 1)
            mes_label = parts[1] if len(parts) > 1 else "Mes"
        token = _slug_token(mes_label)
        if token in used_tokens:
            token = f"{token}_{len(used_tokens) + 1}"
        used_tokens.add(token)
        proventos_meta.append({
            "col": col,
            "mes_label": mes_label,
            "field": f"proventos_{token}",
        })
    colab_unid_cols = [
        c for c in [
            _pick_col(colab_raw.columns, {"UNIDADE"}),
            _pick_col(colab_raw.columns, {"UNIDADE 2", "UNIDADE2"}),
            _pick_col(colab_raw.columns, {"UNIDADE 3", "UNIDADE3"}),
        ]
        if c is not None
    ]

    if not colab_unid_cols:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), [
            "A aba COLABORADORES precisa ter ao menos a coluna unidade.",
        ]

    rows = []
    for idx, row in colab_raw.iterrows():
        nome = str(row[colab_nome_col]).strip() if colab_nome_col and pd.notna(row[colab_nome_col]) else f"COLAB_{idx + 1}"
        cpf_raw = str(row[colab_cpf_col]).strip() if colab_cpf_col and pd.notna(row[colab_cpf_col]) else ""
        cpf_digits = re.sub(r"\D", "", cpf_raw)
        colab_id = cpf_digits if cpf_digits else f"{normalize_text(nome) or 'SEM_NOME'}_{idx}"
        adm_txt = _fmt_adm(row[colab_adm_col]) if colab_adm_col and colab_adm_col in row.index else ""
        situacao_txt = str(row[colab_situacao_col]).strip() if colab_situacao_col and pd.notna(row[colab_situacao_col]) else ""

        proventos_row = {}
        for meta in proventos_meta:
            raw_val = row[meta["col"]] if meta["col"] in row.index else None
            proventos_row[meta["field"]] = _to_money(raw_val)

        unidades = []
        for uc in colab_unid_cols:
            val = row[uc]
            if pd.isna(val):
                continue
            u_txt = str(val).strip()
            u_norm = normalize_text(u_txt)
            if u_norm:
                unidades.append((u_txt, u_norm))

        unidades = list(dict.fromkeys(unidades))
        for u_txt, u_norm in unidades:
            item = {
                "colaborador_id": colab_id,
                "colaborador": nome,
                "cpf": cpf_raw,
                "cargo": str(row[colab_cargo_col]).strip() if colab_cargo_col and pd.notna(row[colab_cargo_col]) else "",
                "regime_trabalho": str(row[colab_regime_col]).strip() if colab_regime_col and pd.notna(row[colab_regime_col]) else "",
                "crm": str(row[colab_crm_col]).strip() if colab_crm_col and pd.notna(row[colab_crm_col]) else "",
                "adm": adm_txt,
                "situacao": situacao_txt,
                "unidade": u_txt,
                "unidade_norm": u_norm,
            }
            item.update(proventos_row)
            rows.append(item)

    aloc = pd.DataFrame(rows)
    if aloc.empty:
        geo["qtd_colaboradores"] = 0
        geo["qtd_medicos"] = 0
        return geo, pd.DataFrame(columns=["unidade", "qtd_colaboradores", "qtd_medicos"]), pd.DataFrame(), []

    aloc = aloc.drop_duplicates(subset=["colaborador_id", "unidade_norm"]).copy()
    aloc["is_medico"] = aloc["crm"].astype(str).str.strip().ne("")

    resumo_unidade = (
        aloc.groupby("unidade_norm", as_index=False)
        .agg(
            qtd_colaboradores=("colaborador_id", "nunique"),
            qtd_medicos=("is_medico", "sum"),
        )
    )

    geo = geo.merge(resumo_unidade, on="unidade_norm", how="left")
    geo["qtd_colaboradores"] = geo["qtd_colaboradores"].fillna(0).astype(int)
    geo["qtd_medicos"] = geo["qtd_medicos"].fillna(0).astype(int)

    ranking = geo[["unidade", "tipo", "qtd_colaboradores", "qtd_medicos"]].copy()
    ranking = ranking.sort_values(["qtd_colaboradores", "qtd_medicos"], ascending=False).reset_index(drop=True)

    proventos_fields = [meta["field"] for meta in proventos_meta]
    detail_cols = [
        "unidade",
        "unidade_norm",
        "colaborador",
        "cpf",
        "cargo",
        "regime_trabalho",
        "crm",
        "adm",
        "situacao",
        "colaborador_id",
    ] + proventos_fields

    aloc_detail = (
        aloc[detail_cols]
        .sort_values(["unidade", "colaborador"], ascending=[True, True])
        .reset_index(drop=True)
    )

    return geo, ranking, aloc_detail, []


@st.cache_data(show_spinner=False)
def load_colaboradores_sheet(file_bytes=None, _mtime=None):
    """Carrega a aba COLABORADORES completa do arquivo fonte ativo (local/upload)."""
    try:
        if file_bytes is None:
            path = local_excel_path()
            if not path:
                return pd.DataFrame()
            xls = pd.ExcelFile(path)
            if "COLABORADORES" not in xls.sheet_names:
                return pd.DataFrame()
            df = pd.read_excel(xls, sheet_name="COLABORADORES")
        else:
            xls = pd.ExcelFile(BytesIO(file_bytes))
            if "COLABORADORES" not in xls.sheet_names:
                return pd.DataFrame()
            df = pd.read_excel(xls, sheet_name="COLABORADORES")
    except Exception:
        return pd.DataFrame()

    if df is None or df.empty:
        return pd.DataFrame()

    # Mantem nomes originais para exibicao, removendo apenas espacos excedentes.
    df = df.rename(columns={c: str(c).strip() for c in df.columns})

    for col in df.columns:
        col_norm = normalize_text(col) or ""
        if any(token in col_norm for token in ["PROVENTOS", "DESCONTOS", "LIQUIDO"]):
            df[col] = pd.to_numeric(df[col].map(normalize_value), errors="coerce")
        if "ADM" in col_norm:
            parsed = pd.to_datetime(df[col], errors="coerce")
            if parsed.notna().any():
                df[col] = parsed.dt.strftime("%d/%m/%Y")

    return df



# ---------------------------------------------------------------------------
# LOADER CELK — dados transacionais brutos (linha por atendimento)
# ---------------------------------------------------------------------------

def _celk_paths():
    """Retorna todos os arquivos CELK presentes em data_raw/, ordenados por mtime."""
    import glob as _glob
    base_dir = os.path.join(os.path.dirname(__file__), "data_raw")
    if not os.path.isdir(base_dir):
        return []
    pattern = os.path.join(base_dir, "producao_consolidada_*.xlsx")
    files = sorted(_glob.glob(pattern), key=os.path.getmtime)
    return files


def _celk_mtime():
    paths = _celk_paths()
    return sum(int(os.path.getmtime(p) * 1000) for p in paths) if paths else 0


def _celk_alias_files():
    base_dir = Path(__file__).resolve().parent
    data_raw_dir = base_dir / "data_raw"
    return [
        data_raw_dir / "mapeamento_unidades_celk.xlsx",
        data_raw_dir / "mapeamento_unidades_celk.csv",
        data_raw_dir / "relacao_unidades_celk.xlsx",
        data_raw_dir / "relacao_unidades_celk.csv",
        base_dir / "mapeamento_unidades_celk.xlsx",
        base_dir / "mapeamento_unidades_celk.csv",
        base_dir / "relacao_unidades_celk.xlsx",
        base_dir / "relacao_unidades_celk.csv",
    ]


@st.cache_data(show_spinner=False)
def load_celk_unit_aliases(_mtime=None):
    """
    Carrega relação A->B para unidades CELK (de-para).
    A = nome da unidade na planilha CELK, B = nome canônico que deve prevalecer no app.
    """
    alias_map = {}

    def _consume_df(df_alias):
        if df_alias is None or df_alias.empty or df_alias.shape[1] < 2:
            return
        for _, row in df_alias.iloc[:, :2].iterrows():
            unidade_a = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            unidade_b = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            if not unidade_a or not unidade_b:
                continue
            key = normalize_text(unidade_a)
            if not key or key in {"UNIDADE", "COLUNA A", "A"}:
                continue
            alias_map[key] = unidade_b

    for alias_file in _celk_alias_files():
        if not alias_file.exists():
            continue
        try:
            if alias_file.suffix.lower() == ".csv":
                _consume_df(pd.read_csv(alias_file, dtype=str, header=None))
            else:
                excel_file = pd.ExcelFile(alias_file)
                for sheet_name in excel_file.sheet_names:
                    _consume_df(pd.read_excel(alias_file, sheet_name=sheet_name, dtype=str, header=None))
        except Exception:
            pass

    mapping_sheet_tokens = ("MAPE", "RELAC", "DE PARA", "DE-PARA", "CORRESP")
    for celk_file in _celk_paths():
        try:
            excel_file = pd.ExcelFile(celk_file)
            for sheet_name in excel_file.sheet_names:
                sheet_norm = normalize_text(sheet_name) or ""
                if not any(token in sheet_norm for token in mapping_sheet_tokens):
                    continue
                _consume_df(pd.read_excel(celk_file, sheet_name=sheet_name, dtype=str, header=None))
        except Exception:
            pass

    return alias_map


def _infer_celk_group(unidade):
    txt = normalize_text(unidade) or ""
    if "UPA" in txt:
        return "UPA"
    if any(token in txt for token in ["PSF", "UBS", "UBSF", "SAD"]):
        return "Atenção Básica"
    if any(token in txt for token in ["CEO", "ODONTO"]):
        return "Odontologia"
    if any(token in txt for token in ["CAIS", "MATERNO", "CAPS"]):
        return "Atenção Secundária"
    return None


def _is_upa_unit(unidade):
    txt = normalize_text(unidade) or ""
    return "UPA" in txt


@st.cache_data(show_spinner=False)
def load_celk_data(_mtime=None):
    """
    Lê os arquivos CELK da pasta data_raw e retorna DataFrame normalizado.
    Colunas resultantes: DATA (datetime), PACIENTE, CARTAO_SUS, IDADE,
    TIPO_ATENDIMENTO, CODIGO_SIGTAP, DESCRICAO_PROCEDIMENTO, PROFISSIONAL, UNIDADE,
    + derivadas: HORA (int), DIA_SEMANA (str PT-BR), DIA_MES (int),
    MES_LABEL (str), SEMANA_MES (str S1-S6).
    """
    paths = _celk_paths()
    if not paths:
        return pd.DataFrame()

    frames = []
    for path in paths:
        try:
            df = pd.read_excel(
                path,
                sheet_name=0,
                usecols="A:I",
                dtype=str,
            )
            frames.append(df)
        except Exception:
            pass

    if not frames:
        return pd.DataFrame()

    df = pd.concat(frames, ignore_index=True)

    # Normaliza nomes de coluna
    df.columns = [c.strip().upper() for c in df.columns]
    rename_map = {
        "DESCRIÇÃO DO PROCEDIMENTO": "DESCRICAO_PROCEDIMENTO",
        "DESCRICAO DO PROCEDIMENTO": "DESCRICAO_PROCEDIMENTO",
        "DESCRI\u00c7\u00c3O DO PROCEDIMENTO": "DESCRICAO_PROCEDIMENTO",
    }
    df.rename(columns=rename_map, inplace=True)

    # Parse DATA → datetime
    # ISO format YYYY-MM-DD HH:MM:SS vindo do Excel → dayfirst=False
    df["DATA"] = pd.to_datetime(df["DATA"], errors="coerce", dayfirst=False)
    df = df.dropna(subset=["DATA"])

    # Campos numéricos
    if "IDADE" in df.columns:
        df["IDADE"] = pd.to_numeric(df["IDADE"], errors="coerce")

    # Derivadas temporais
    _dow_map = {0: "Segunda", 1: "Terça", 2: "Quarta", 3: "Quinta", 4: "Sexta", 5: "Sábado", 6: "Domingo"}
    df["HORA"] = df["DATA"].dt.hour.astype(int)
    df["DIA_MES"] = df["DATA"].dt.day.astype(int)
    df["DIA_SEMANA"] = df["DATA"].dt.dayofweek.map(_dow_map)
    df["MES_LABEL"] = df["DATA"].dt.to_period("M").dt.strftime("%b/%y").str.capitalize()
    df["SEMANA_MES"] = df["DATA"].dt.day.apply(lambda d: f"S{(d - 1) // 7 + 1}")

    # Normaliza UNIDADE
    if "UNIDADE" in df.columns:
        df["UNIDADE"] = df["UNIDADE"].str.strip()
        df["UNIDADE_CELK_ORIGINAL"] = df["UNIDADE"]
        unidade_celk_norm = df["UNIDADE"].map(normalize_text)

        _CELK_UNIT_MAP = {
            # ── UPAs ───────────────────────────────────────────────────
            "UNIDADE DE PRONTO ATENDIMENTO DE LUZIANIA UPA":    ("UPA", "UPA II Luziânia"),
            "UNIDADE DE PRONTO ATENDIMENTO DO JARDIM INGA UPA": ("UPA", "UPA I Jardim Ingá"),
            # ── Atenção Básica ─────────────────────────────────────────
            "PSF ALTO DAS CARAIBAS":                            ("Atenção Básica", "PSF Alto das Caraíbas"),
            "PSF AMERICANOS":                                   ("Atenção Básica", "PSF Americanos"),
            "PSF CRUZEIRO":                                     ("Atenção Básica", "PSF Cruzeiro"),
            "PSF JARDIM DO INGA":                               ("Atenção Básica", "PSF Jardim do Ingá"),
            "PSF JARDIM IPE":                                   ("Atenção Básica", "PSF Jardim Ipê"),
            "PSF JARDIM MARILIA":                               ("Atenção Básica", "PSF Jardim Marília"),
            "PSF JARDIM PLANALTO":                              ("Atenção Básica", "PSF Jardim Planalto"),
            "PSF MANDU":                                        ("Atenção Básica", "PSF Mandu"),
            "PSF MANIRATUBA":                                   ("Atenção Básica", "PSF Maniratuba"),
            "PSF MINGONE I":                                    ("Atenção Básica", "PSF Mingone I"),
            "PSF MINGONE II":                                   ("Atenção Básica", "PSF Mingone II"),
            "PSF NORTE MARAVILHA":                              ("Atenção Básica", "PSF Norte Maravilha"),
            "PSF PARQUE ALVORADA":                              ("Atenção Básica", "PSF Parque Alvorada"),
            "PSF PARQUE ESTRELA DALVA IX":                      ("Atenção Básica", "PSF Parque Estrela D'Alva IX"),
            "PSF PARQUE ESTRELA DALVA VIII":                    ("Atenção Básica", "PSF Parque Estrela D'Alva VIII"),
            "PSF PARQUE ESTRELA DALVA X":                       ("Atenção Básica", "PSF Parque Estrela D'Alva X"),
            "PSF PARQUE SANTA FE":                              ("Atenção Básica", "PSF Parque Santa Fé"),
            "PSF PARQUE SOL NASCENTE":                          ("Atenção Básica", "PSF Parque Sol Nascente"),
            "PSF RESIDENCIAL COPAIBAS":                         ("Atenção Básica", "PSF Residencial Copaíbas"),
            "PSF SETOR LESTE":                                  ("Atenção Básica", "PSF Setor Leste"),
            "PSF SETOR LESTE / SAO CAETANO":                    ("Atenção Básica", "PSF Setor Leste / São Caetano"),
            "PSF TRES VENDAS":                                  ("Atenção Básica", "PSF Três Vendas"),
            "PSF VILA ESPERANCA":                               ("Atenção Básica", "PSF Vila Esperança"),
            "PSF VILA JURACY":                                  ("Atenção Básica", "PSF Vila Juracy"),
            "SERVICO DE ATENDIMENTO DOMICILIAR - SAD":          ("Atenção Básica", "SAD"),
            "UNIDADE BASICA DE SAUDE CIDADE OSFAYA":            ("Atenção Básica", "UBS Cidade Osfaya"),
            "UNIDADE BASICA DE SAUDE DO JARDIM SAO PAULO":      ("Atenção Básica", "UBS Jardim São Paulo"),
            "UNIDADE BASICA DE SAUDE JARDIM LUZILIA":           ("Atenção Básica", "UBS Jardim Luzília"),
            "UNIDADE BASICA DE SAUDE PED IX I":                 ("Atenção Básica", "UBS PED IX I"),
            "UNIDADE BASICA DE SAUDE SETOR AEROPORTO":          ("Atenção Básica", "UBS Setor Aeroporto"),
            "UNIDADE BASICA DE SAUDE SETOR INDUSTRIAL":         ("Atenção Básica", "UBS Setor Industrial"),
            # ── Atenção Secundária ─────────────────────────────────────
            "CAIS I":                                           ("Atenção Secundária", "CAIS I"),
            "MATERNO INFANTIL":                                 ("Atenção Secundária", "Materno Infantil"),
            # ── Odontologia ────────────────────────────────────────────
            "CENTRO ESPECIALIZADO DE ODONTOLOGIA CEO":          ("Odontologia", "CEO"),
        }
        unit_map_norm = {normalize_text(k): v for k, v in _CELK_UNIT_MAP.items()}
        fallback_group_map = {k: v[0] for k, v in unit_map_norm.items()}
        fallback_name_map = {k: v[1] for k, v in unit_map_norm.items()}

        alias_map = load_celk_unit_aliases(_mtime=_mtime)
        unidade_canonica = unidade_celk_norm.map(alias_map)
        unidade_canonica = unidade_canonica.fillna(unidade_celk_norm.map(fallback_name_map)).fillna(df["UNIDADE"])
        unidade_canonica = unidade_canonica.astype(str).str.strip()

        grupo_painel = unidade_celk_norm.map(fallback_group_map)
        grupo_painel = grupo_painel.fillna(unidade_canonica.map(_infer_celk_group))

        df["UNIDADE"] = unidade_canonica
        df["UNIDADE_NORM"] = df["UNIDADE"].map(normalize_text)
        df["GRUPO_PAINEL"] = grupo_painel
        df["UNIDADE_PAINEL"] = unidade_canonica

    return df


@st.cache_data(show_spinner=False)
def load_produtividade_data(_mtime=None):
    paths = _urgencia_paths()

    empty = {
        "kpi_diario": pd.DataFrame(),
        "kpi_diario_unidade": pd.DataFrame(),
        "kpi_semanal": pd.DataFrame(),
        "ranking": pd.DataFrame(),
        "top5_geral": pd.DataFrame(),
        "top5_upa2": pd.DataFrame(),
        "top5_upa1": pd.DataFrame(),
    }
    if not paths:
        return empty

    def _read_sheet(path, name):
        xls = pd.ExcelFile(path)
        if name not in xls.sheet_names:
            return pd.DataFrame()
        return pd.read_excel(path, sheet_name=name)

    def _concat_sheet(name):
        frames = [_read_sheet(p, name) for p in paths]
        frames = [f for f in frames if not f.empty]
        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    # Mantém compatibilidade com código que usa xls/_sheet internamente
    path = paths[-1]
    xls = pd.ExcelFile(path)

    def _sheet(name):
        return _concat_sheet(name)

    kpi_diario = _sheet("KPI_DIARIO_GERAL")
    kpi_diario_unidade = _sheet("KPI_DIARIO_UNIDADE")
    kpi_semanal = _sheet("KPI_SEMANAL_GERAL")
    ranking = _sheet("RANKING_MEDICOS")
    top5_geral = _sheet("TOP5_GERAL")
    top5_upa2 = _sheet("TOP5_UPA_II")
    top5_upa1 = _sheet("TOP5_UPA_I")

    if "Data" in kpi_diario.columns:
        kpi_diario["Data"] = pd.to_datetime(kpi_diario["Data"], errors="coerce", dayfirst=True)
    for col in ["UPA II DE LUZI\u00c2NIA", "UPA I JARDIM ING\u00c1", "SAMU", "Total_Geral_24h"]:
        if col in kpi_diario.columns:
            kpi_diario[col] = pd.to_numeric(kpi_diario[col], errors="coerce")

    if "Data" in kpi_diario_unidade.columns:
        kpi_diario_unidade["Data"] = pd.to_datetime(kpi_diario_unidade["Data"], errors="coerce", dayfirst=True)
    for col in ["Total_24h_Final", "Media_Hora_24h", "Subtotal_Diurno", "Subtotal_Noturno"]:
        if col in kpi_diario_unidade.columns:
            kpi_diario_unidade[col] = pd.to_numeric(kpi_diario_unidade[col], errors="coerce")

    if "Semana_Inicio" in kpi_semanal.columns:
        kpi_semanal["Semana_Inicio"] = pd.to_datetime(kpi_semanal["Semana_Inicio"], errors="coerce", dayfirst=True)
    if "Semana_Fim" in kpi_semanal.columns:
        kpi_semanal["Semana_Fim"] = pd.to_datetime(kpi_semanal["Semana_Fim"], errors="coerce", dayfirst=True)
    for col in ["Total_Semana_Geral", "Media_Diaria_Geral", "Total_Semana_UPA_II", "Total_Semana_UPA_I", "Total_Semana_SAMU"]:
        if col in kpi_semanal.columns:
            kpi_semanal[col] = pd.to_numeric(kpi_semanal[col], errors="coerce")

    for col in ["Total_Atendimentos", "Plantoes", "Media_por_Plantao", "Media_por_Hora"]:
        for df_r in [ranking, top5_geral, top5_upa2, top5_upa1]:
            if col in df_r.columns:
                df_r[col] = pd.to_numeric(df_r[col], errors="coerce")

    return {
        "kpi_diario": kpi_diario,
        "kpi_diario_unidade": kpi_diario_unidade,
        "kpi_semanal": kpi_semanal,
        "ranking": ranking,
        "top5_geral": top5_geral,
        "top5_upa2": top5_upa2,
        "top5_upa1": top5_upa1,
    }


def _urgencia_paths():
    """Retorna todos os arquivos urgencia*validado*.xlsx ordenados por data de modificacao."""
    base = Path(__file__).parent
    candidates = sorted(base.glob("urgencia*validado*.xlsx"), key=lambda p: p.stat().st_mtime)
    if candidates:
        return candidates
    # Fallback para urgencia_tratado_final.xlsx
    fallback = base / "urgencia_tratado_final.xlsx"
    return [fallback] if fallback.exists() else []


def _urgencia_path():
    """Retorna o arquivo urgencia mais recente (compatibilidade)."""
    paths = _urgencia_paths()
    return paths[-1] if paths else None


def _samu_file_mtime():
    paths = _urgencia_paths()
    return sum(p.stat().st_mtime for p in paths) if paths else 0


def _parse_samu_file(path):
    """Lê e parseia a aba SAMU de um arquivo. Retorna (diario_rows, resumo_rows, titulo)."""
    raw = pd.read_excel(path, sheet_name="SAMU", header=None)
    if raw.empty or raw.shape[1] < 2:
        return [], [], "SAMU"

    titulo = str(raw.iloc[0, 0]).strip() if pd.notna(raw.iloc[0, 0]) else "SAMU"

    header_row = None
    for idx in range(len(raw)):
        c0 = normalize_text(raw.iloc[idx, 0])
        c1 = normalize_text(raw.iloc[idx, 1])
        if c0 == "DESCRICAO" and c1 and "SIGTAP" in c1:
            header_row = idx
            break

    if header_row is None:
        return [], [], titulo

    header_vals = raw.iloc[header_row]
    day_cols = []
    total_col = None
    falta_col = None
    eficacia_col = None
    meta_col = None

    for col_idx in range(2, raw.shape[1]):
        hv = header_vals.iloc[col_idx]
        hv_norm = normalize_text(hv)
        hv_num = pd.to_numeric(pd.Series([hv]), errors="coerce").iloc[0]

        if pd.notna(hv_num) and float(hv_num).is_integer() and 1 <= int(hv_num) <= 31:
            day_cols.append((col_idx, int(hv_num)))
            continue

        if hv_norm == "TOTAL":
            total_col = col_idx
        elif hv_norm == "FALTA":
            falta_col = col_idx
        elif hv_norm and "% EFICACIA" in hv_norm:
            eficacia_col = col_idx
        elif hv_norm == "META":
            meta_col = col_idx

    month_map = {
        "JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "ABRIL": 4,
        "MAIO": 5, "JUNHO": 6, "JULHO": 7, "AGOSTO": 8,
        "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
    }
    titulo_norm = normalize_text(titulo) or ""
    month_year = re.search(r"(JANEIRO|FEVEREIRO|MARCO|ABRIL|MAIO|JUNHO|JULHO|AGOSTO|SETEMBRO|OUTUBRO|NOVEMBRO|DEZEMBRO)\s+(\d{4})", titulo_norm)

    base_date = None
    if month_year:
        month_name, year_str = month_year.groups()
        month_num = month_map.get(month_name)
        if month_num:
            base_date = dt.date(int(year_str), month_num, 1)

    diario_rows = []
    resumo_rows = []

    for ridx in range(header_row + 1, len(raw)):
        row = raw.iloc[ridx]
        desc = row.iloc[0] if pd.notna(row.iloc[0]) else None
        if not desc:
            continue

        desc_text = str(desc).strip()
        if normalize_text(desc_text) == "TOTAL:":
            continue

        codigo = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else None

        total_val = pd.to_numeric(pd.Series([row.iloc[total_col]]), errors="coerce").iloc[0] if total_col is not None else pd.NA
        falta_val = pd.to_numeric(pd.Series([row.iloc[falta_col]]), errors="coerce").iloc[0] if falta_col is not None else pd.NA
        efic_val = pd.to_numeric(pd.Series([row.iloc[eficacia_col]]), errors="coerce").iloc[0] if eficacia_col is not None else pd.NA
        meta_val = pd.to_numeric(pd.Series([row.iloc[meta_col]]), errors="coerce").iloc[0] if meta_col is not None else pd.NA

        resumo_rows.append({
            "Descricao": desc_text,
            "Codigo_SIGTAP": codigo,
            "Total": float(total_val) if pd.notna(total_val) else pd.NA,
            "Falta": float(falta_val) if pd.notna(falta_val) else pd.NA,
            "Eficacia": float(efic_val) if pd.notna(efic_val) else pd.NA,
            "Meta": float(meta_val) if pd.notna(meta_val) else pd.NA,
        })

        for col_idx, day_num in day_cols:
            val = pd.to_numeric(pd.Series([row.iloc[col_idx]]), errors="coerce").iloc[0]
            if pd.isna(val):
                continue

            data_ref = None
            if base_date is not None:
                try:
                    data_ref = dt.date(base_date.year, base_date.month, day_num)
                except ValueError:
                    data_ref = None

            diario_rows.append({
                "Data": pd.to_datetime(data_ref) if data_ref is not None else pd.NaT,
                "Dia": day_num,
                "Descricao": desc_text,
                "Codigo_SIGTAP": codigo,
                "Atendimentos": float(val),
            })

    return diario_rows, resumo_rows, titulo


@st.cache_data(show_spinner=False)
def load_samu_data(_mtime=None):
    paths = _urgencia_paths()

    empty_result = {
        "diario": pd.DataFrame(columns=["Data", "Dia", "Descricao", "Codigo_SIGTAP", "Atendimentos"]),
        "resumo": pd.DataFrame(columns=["Descricao", "Codigo_SIGTAP", "Total", "Falta", "Eficacia", "Meta"]),
        "titulo": "SAMU",
    }

    if not paths:
        return empty_result

    all_diario = []
    all_resumo = []
    titulos = []

    for p in paths:
        xls_check = pd.ExcelFile(p)
        if "SAMU" not in xls_check.sheet_names:
            continue
        d_rows, r_rows, titulo = _parse_samu_file(p)
        all_diario.extend(d_rows)
        all_resumo.extend(r_rows)
        titulos.append(titulo)

    if not all_diario and not all_resumo:
        return empty_result

    diario_df = pd.DataFrame(all_diario)
    resumo_df = pd.DataFrame(all_resumo)

    if not diario_df.empty:
        diario_df["Atendimentos"] = pd.to_numeric(diario_df["Atendimentos"], errors="coerce")
        diario_df = diario_df.dropna(subset=["Atendimentos"]).copy()

    titulo_final = " · ".join(titulos) if titulos else "SAMU"

    return {
        "diario": diario_df,
        "resumo": resumo_df,
        "titulo": titulo_final,
    }


def format_currency_br(x):
    if x is None or pd.isna(x):
        return "R$ -"
    return f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def financeiro_kpis(fin_df):
    work = fin_df.dropna(subset=["valor_num"]).copy()

    if work.empty:
        return {
            "total": 0.0,
            "media_mensal": 0.0,
            "maior_mes": 0.0,
            "fornecedores_ativos": 0,
        }

    total = float(work["valor_num"].sum())

    mensal = (
        work.groupby(["mes", "mes_label"], observed=True)["valor_num"]
        .sum()
        .reset_index()
        .sort_values("mes")
    )

    # considera apenas meses com movimentação real (valor > 0)
    mensal_com_dados = mensal[mensal["valor_num"] > 0]

    media_mensal = float(mensal_com_dados["valor_num"].mean()) if not mensal_com_dados.empty else 0.0
    maior_mes = float(mensal_com_dados["valor_num"].max()) if not mensal_com_dados.empty else 0.0

    fornecedores_ativos = int(
        work.groupby("fornecedor")["valor_num"].sum().gt(0).sum()
    )

    return {
        "total": total,
        "media_mensal": media_mensal,
        "maior_mes": maior_mes,
        "fornecedores_ativos": fornecedores_ativos,
    }


def render_financeiro_page(fin_df, meses_filtrados):
    st.subheader("Financeiro")

    if fin_df is None or fin_df.empty:
        st.warning("A aba 'Financeiro' não foi encontrada ou está vazia.")
        return

    work = fin_df.copy()

    # respeita o filtro global de período do app
    if meses_filtrados and "mes_label" in work.columns:
        work = work[work["mes_label"].isin(meses_filtrados)].copy()

    work = work.dropna(subset=["valor_num"])
    if work.empty:
        st.info("Sem dados financeiros para o período selecionado.")
        return

    kpis = financeiro_kpis(work)

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        top_kpi_card("Gasto total", format_currency_br(kpis["total"]), icon="💰", subtitle="▲ total no período", accent_color="#22C55E", subtitle_color="#16A34A")
    with c2:
        top_kpi_card("Média mensal", format_currency_br(kpis["media_mensal"]), icon="📊", subtitle="▲ média dos meses filtrados", accent_color="#3B82F6", subtitle_color="#2563EB")
    with c3:
        top_kpi_card("Maior mês", format_currency_br(kpis["maior_mes"]), icon="📈", subtitle="▲ pico de gasto mensal", accent_color="#F97316", subtitle_color="#EA580C")
    with c4:
        top_kpi_card("Fornecedores ativos", format_int(kpis["fornecedores_ativos"]), icon="🏢", subtitle="▲ com lançamento no período", accent_color="#EF4444", subtitle_color="#DC2626")
    section_end()

    mensal = (
        work.groupby(["mes", "mes_label"], observed=True)["valor_num"]
        .sum()
        .reset_index()
        .sort_values("mes")
    )
    mensal = mensal[mensal["valor_num"] > 0]

    section_start(
        "",
        ""
    )
    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=mensal["mes_label"],
            y=mensal["valor_num"],
            name="Gasto mensal",
            marker_color=SEMANTIC_COLORS["realizado"],
            hovertemplate="<b>Gasto mensal</b><br>Mês: %{x}<br>Valor: R$ %{y:,.2f}<extra></extra>"
        )
    )
    fig.add_trace(
        go.Scatter(
            x=mensal["mes_label"],
            y=mensal["valor_num"],
            mode="lines+markers",
            name="Tendência",
            line=dict(color=SEMANTIC_COLORS["media"], width=3),
            marker=dict(size=7, color=SEMANTIC_COLORS["media"]),
            hovertemplate="<b>Tendência</b><br>Mês: %{x}<br>Valor: R$ %{y:,.2f}<extra></extra>"
        )
    )
    fig = apply_plotly_theme(
        fig,
        title="Gasto total por mês",
        subtitle="",
        yaxis_title="Valor (R$)",
        height=390,
        legend=True,
        legend_orientation="h"
    )
    fig = apply_month_axis_order(fig, mensal)
    plot(fig, "financeiro_mensal")
    section_end()

    fornecedores = (
        work.groupby("fornecedor", as_index=False)["valor_num"]
        .sum()
        .sort_values("valor_num", ascending=False)
    )

    top_fornecedores = fornecedores.head(10).copy()

    section_start(
        "",
        ""
    )
    fig_top = go.Figure()
    fig_top.add_trace(
        go.Bar(
            x=top_fornecedores["valor_num"],
            y=top_fornecedores["fornecedor"],
            orientation="h",
            name="Total",
            marker_color=SEMANTIC_COLORS["primary_soft"],
            hovertemplate="<b>%{y}</b><br>Total: R$ %{x:,.2f}<extra></extra>"
        )
    )
    fig_top = apply_plotly_theme(
        fig_top,
        title="Top 10 fornecedores por gasto",
        subtitle="",
        yaxis_title="",
        height=430,
        legend=False
    )
    fig_top.update_layout(yaxis=dict(autorange="reversed"))
    plot(fig_top, "financeiro_top_fornecedores")
    section_end()

    composicao = top_fornecedores.copy()
    total_comp = composicao["valor_num"].sum()
    composicao["participacao_pct"] = (
        (composicao["valor_num"] / total_comp) * 100 if total_comp else 0
    )

    section_start(
        "<span style='font-size:1.52rem;font-weight:900;line-height:1.15;'>Detalhamento analítico</span>",
        "Tabela executiva com consolidação por fornecedor"
    )
    tabela = fornecedores.copy()
    tabela["Média mensal"] = tabela["valor_num"] / max(len(mensal), 1)
    tabela["Participação %"] = (
        (tabela["valor_num"] / tabela["valor_num"].sum()) * 100
        if tabela["valor_num"].sum() > 0 else 0
    )

    tabela_view = tabela.rename(columns={
        "fornecedor": "Fornecedor",
        "valor_num": "Total no período",
    }).copy()

    tabela_view["Total no período"] = tabela_view["Total no período"].apply(format_currency_br)
    tabela_view["Média mensal"] = tabela_view["Média mensal"].apply(format_currency_br)
    tabela_view["Participação %"] = tabela_view["Participação %"].apply(format_pct_br)

    st.table(
        tabela_view[["Fornecedor", "Total no período", "Média mensal", "Participação %"]]
        .reset_index(drop=True)
    )
    section_end()

def filter_panel(df, unidade, painel):
    unidade_norm = normalize_text(unidade)
    painel_norm = normalize_text(painel)

    df_test = df[df["unidade_norm"] == unidade_norm]

    # tenta match exato
    result = df_test[df_test["painel_norm"] == painel_norm]

    # fallback por contains quando há pequenas variações de rótulo
    if result.empty:
        result = df_test[
            df_test["painel_norm"].str.contains(painel_norm, na=False)
        ]

    # fallback fuzzy para casos de texto corrompido no Excel (ex.: CLASSIFICA��O)
    if result.empty and not df_test.empty and painel_norm:
        import difflib

        candidatos = (
            df_test["painel_norm"]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        melhor = None
        melhor_score = 0.0
        for cand in candidatos:
            score = difflib.SequenceMatcher(None, painel_norm, cand).ratio()
            if score > melhor_score:
                melhor_score = score
                melhor = cand

        if melhor is not None and melhor_score >= 0.72:
            result = df_test[df_test["painel_norm"] == melhor]

    return result.copy()

def format_int(x):
    if pd.isna(x):
        return "-"
    return f"{int(round(x)):,}".replace(",", ".")

def clean_card_value(value):
    if value is None:
        return "-"

    value = str(value)

    replacements = [
        "<div style='",
        '<div style="',
        "</div>",
        "<div>",
        "</span>",
        "<span>",
        "&nbsp;"
    ]
    for item in replacements:
        value = value.replace(item, "")

    import re
    value = re.sub(r"<[^>]+>", "", value)
    value = re.sub(r"\s+", " ", value).strip()

    return value if value else "-"


def metric_sum(df, serie_norm=None, exclude_series_norm=None, month=None):
    work = df.copy()

    if month is not None:
        work = work[work["mes"] == month]

    if serie_norm is not None:
        if isinstance(serie_norm, str):
            serie_norm = [serie_norm]
        serie_norm = [str(x).strip().upper() for x in serie_norm]
        work = work[work["serie_norm"].isin(serie_norm)]

    if exclude_series_norm is not None:
        if isinstance(exclude_series_norm, str):
            exclude_series_norm = [exclude_series_norm]
        exclude_series_norm = [str(x).strip().upper() for x in exclude_series_norm]
        work = work[~work["serie_norm"].isin(exclude_series_norm)]

    work = work.dropna(subset=["valor_num"])

    if work.empty:
        return None

    return float(work["valor_num"].sum())


def latest_and_previous_month(df, serie_norm=None, exclude_series_norm=None):
    work = df.copy()

    if serie_norm is not None:
        if isinstance(serie_norm, str):
            serie_norm = [serie_norm]
        serie_norm = [str(x).strip().upper() for x in serie_norm]
        work = work[work["serie_norm"].isin(serie_norm)]

    if exclude_series_norm is not None:
        if isinstance(exclude_series_norm, str):
            exclude_series_norm = [exclude_series_norm]
        exclude_series_norm = [str(x).strip().upper() for x in exclude_series_norm]
        work = work[~work["serie_norm"].isin(exclude_series_norm)]

    work = work.dropna(subset=["mes", "valor_num"]).sort_values("mes")

    if work.empty:
        return None, None

    months = []
    for m in work["mes"].tolist():
        if m not in months:
            months.append(m)

    latest = months[-1] if months else None
    previous = months[-2] if len(months) >= 2 else None
    return latest, previous


def calc_delta_pct(current, previous):
    if current is None or previous is None:
        return None
    if pd.isna(current) or pd.isna(previous):
        return None
    if previous == 0:
        return None
    return ((current - previous) / previous) * 100


def build_kpi_context(df, serie_norm=None, exclude_series_norm=None, meta_series="META"):
    latest_month, previous_month = latest_and_previous_month(
        df,
        serie_norm=serie_norm,
        exclude_series_norm=exclude_series_norm
    )

    current_value = metric_sum(
        df,
        serie_norm=serie_norm,
        exclude_series_norm=exclude_series_norm,
        month=latest_month
    )

    previous_value = metric_sum(
        df,
        serie_norm=serie_norm,
        exclude_series_norm=exclude_series_norm,
        month=previous_month
    )

    total_value = metric_sum(
        df,
        serie_norm=serie_norm,
        exclude_series_norm=exclude_series_norm
    )

    meta_value = metric_sum(
        df,
        serie_norm=meta_series,
        month=latest_month
    )

    return {
        "latest_month": latest_month,
        "previous_month": previous_month,
        "latest_month_label": MESES_LABEL.get(latest_month, str(latest_month) if latest_month else "-"),
        "current": current_value,
        "previous": previous_value,
        "total": total_value,
        "meta": meta_value,
        "delta_pct": calc_delta_pct(current_value, previous_value),
    }

def format_delta_pct(delta):
    if delta is None or pd.isna(delta):
        return "—"
    return f"{delta:+.1f}%".replace(".", ",")

def delta_arrow(delta):
    if delta is None or pd.isna(delta):
        return "•"
    if delta > 0:
        return "↑"
    if delta < 0:
        return "↓"
    return "→"

def delta_color(delta, inverse=False):
    """
    inverse=False: maior é melhor
    inverse=True: menor é melhor
    """
    if delta is None or pd.isna(delta):
        return SEMANTIC_COLORS["neutral"]

    if inverse:
        if delta < 0:
            return SEMANTIC_COLORS["success"]
        if delta > 0:
            return SEMANTIC_COLORS["danger"]
        return SEMANTIC_COLORS["warning"]

    if delta > 0:
        return SEMANTIC_COLORS["success"]
    if delta < 0:
        return SEMANTIC_COLORS["danger"]
    return SEMANTIC_COLORS["warning"]

def format_meta_line(current=None, meta=None):
    if current is None or meta is None or pd.isna(current) or pd.isna(meta):
        return "Meta: —"

    diff = current - meta
    status = "acima"
    if diff < 0:
        status = "abaixo"
    elif diff == 0:
        status = "em linha"

    return (
        f"Meta: {clean_card_value(meta)}"
        f" • {status} em {clean_card_value(abs(diff)) if diff != 0 else '0'}"
    )



def format_pct_br(x):
    if x is None or pd.isna(x):
        return "-"
    return f"{x:,.1f}%".replace(",", "X").replace(".", ",").replace("X", ".")


def format_compact_number(x):
    if x is None or pd.isna(x):
        return "-"
    x = float(x)
    if abs(x) >= 1000000:
        return f"{x / 1000000:,.1f} mi".replace(",", "X").replace(".", ",").replace("X", ".")
    if abs(x) >= 1000:
        return f"{x:,.0f}".replace(",", ".")
    if x.is_integer():
        return str(int(x))
    return f"{x:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")

def format_hours_hms(value):
    if value is None or pd.isna(value):
        return "-"

    total_seconds = int(round(float(value) * 3600))

    sign = "-" if total_seconds < 0 else ""
    total_seconds = abs(total_seconds)

    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"


def time_tick_values(max_value):
    if max_value is None or pd.isna(max_value) or max_value <= 0:
        return [0, 0.25, 0.5, 0.75, 1.0]

    if max_value <= 1:
        step = 10 / 60  # 10 min
    elif max_value <= 3:
        step = 20 / 60  # 20 min
    elif max_value <= 6:
        step = 30 / 60  # 30 min
    elif max_value <= 12:
        step = 1.0      # 1h
    elif max_value <= 24:
        step = 2.0      # 2h
    else:
        step = 6.0      # 6h

    ticks = []
    current = 0.0
    limit = float(max_value) * 1.08

    while current <= limit + 1e-9:
        ticks.append(round(current, 6))
        current += step

    if not ticks:
        ticks = [0.0, round(float(max_value), 6)]

    return ticks


def line_time_chart(
    df,
    title,
    main_series=None,
    prefix="time_line",
    unidade=None,
    subtitle=None,
):
    work = df.dropna(subset=["valor_num"]).copy()
    if work.empty:
        st.info("Sem dados para este gráfico.")
        return

    fig = go.Figure()

    if main_series:
        main_norm = normalize_text(main_series)
        main = work[work["serie_norm"] == main_norm]

        if not main.empty:
            fig.add_trace(
                go.Scatter(
                    x=main["mes_label"],
                    y=main["valor_num"],
                    mode="lines+markers",
                    name=str(main_series),
                    line=dict(color=SEMANTIC_COLORS["realizado"], width=3.5),
                    marker=dict(size=7, color=SEMANTIC_COLORS["realizado"]),
                    customdata=main["valor_num"].apply(format_hours_hms),
                    hovertemplate="<b>%{fullData.name}</b><br>Mês: %{x}<br>Tempo: %{customdata}<extra></extra>"
                )
            )

        others = work[
            (~work["serie_norm"].eq(main_norm)) &
            (~work["serie_norm"].eq("META"))
        ].copy()

        for serie in others["serie"].dropna().unique().tolist():
            temp = others[others["serie"] == serie].copy()
            serie_color = semantic_color(serie, default=SEMANTIC_COLORS["neutral"])

            fig.add_trace(
                go.Scatter(
                    x=temp["mes_label"],
                    y=temp["valor_num"],
                    mode="lines+markers",
                    name=str(serie),
                    line=dict(color=serie_color, width=2.4),
                    marker=dict(size=6, color=serie_color),
                    customdata=temp["valor_num"].apply(format_hours_hms),
                    hovertemplate="<b>%{fullData.name}</b><br>Mês: %{x}<br>Tempo: %{customdata}<extra></extra>"
                )
            )
    else:
        series = work["serie"].dropna().unique().tolist()
        color_map = build_semantic_color_map(series)

        for serie in series:
            temp = work[work["serie"] == serie].copy()
            serie_color = semantic_color(serie, default=color_map.get(serie, SEMANTIC_COLORS["neutral"]))

            fig.add_trace(
                go.Scatter(
                    x=temp["mes_label"],
                    y=temp["valor_num"],
                    mode="lines+markers",
                    name=str(serie),
                    line=dict(color=serie_color, width=3 if "MÉDIA GERAL" in str(serie).upper() or "MEDIA GERAL" in str(serie).upper() else 2.4),
                    marker=dict(size=6, color=serie_color),
                    customdata=temp["valor_num"].apply(format_hours_hms),
                    hovertemplate="<b>%{fullData.name}</b><br>Mês: %{x}<br>Tempo: %{customdata}<extra></extra>"
                )
            )

    meta = work[work["serie_norm"] == "META"].copy()
    if not meta.empty:
        fig.add_trace(
            go.Scatter(
                x=meta["mes_label"],
                y=meta["valor_num"],
                mode="lines+markers",
                name="Meta",
                line=dict(color=SEMANTIC_COLORS["meta"], width=2, dash="dash"),
                marker=dict(size=5, color=SEMANTIC_COLORS["meta"]),
                customdata=meta["valor_num"].apply(format_hours_hms),
                hovertemplate="<b>Meta</b><br>Mês: %{x}<br>Tempo: %{customdata}<extra></extra>"
            )
        )

    fig = apply_plotly_theme(
        fig,
        title=title,
        subtitle=chart_subtitle(work, unidade) if subtitle is None else subtitle,
        yaxis_title="Tempo (HH:MM:SS)",
        height=360,
        legend=True,
        legend_orientation="h"
    )

    max_y = work["valor_num"].max()
    ticks = time_tick_values(max_y)

    fig.update_yaxes(
        tickmode="array",
        tickvals=ticks,
        ticktext=[format_hours_hms(v) for v in ticks]
    )

    fig = apply_month_axis_order(fig, work)
    plot(fig, prefix)

def percent_atingido(executado, meta):
    if executado is None or meta is None or pd.isna(executado) or pd.isna(meta) or meta == 0:
        return None
    return (executado / meta) * 100


def status_meta(executado, meta):
    pct = percent_atingido(executado, meta)
    if pct is None:
        return "Sem base", SEMANTIC_COLORS["neutral"], None
    if executado > meta:
        return "Acima da meta", SEMANTIC_COLORS["success"], ((executado - meta) / meta) * 100
    if executado < meta:
        return "Abaixo da meta", SEMANTIC_COLORS["warning"], ((meta - executado) / meta) * 100
    return "Meta atingida", SEMANTIC_COLORS["info"], 0.0


def compute_executado_for_categoria(data, categoria, mes=None):
    work = data.copy()

    if mes is not None:
        work = work[work["mes"] == mes]

    work = work.dropna(subset=["valor_num"])
    if work.empty:
        return 0.0

    categoria_norm = str(categoria).strip().upper()

    def sum_mask(mask):
        subset = work[mask & work["valor_num"].notna()].copy()
        if subset.empty:
            return 0.0

        subset = subset[
            ~subset["serie_norm"].isin(
                ["META", "MÉDIA DIÁRIA", "MEDIA DIÁRIA", "MEDIA DIARIA", "TOTAL"]
            )
        ]
        return float(subset["valor_num"].sum()) if not subset.empty else 0.0

    painel_upper = work["painel"].astype(str).str.upper()
    painel_norm_upper = work["painel_norm"].astype(str).str.upper() if "painel_norm" in work.columns else painel_upper
    serie_upper = work["serie"].astype(str).str.upper() if "serie" in work.columns else work["serie_norm"].astype(str).str.upper()
    serie_norm_upper = work["serie_norm"].astype(str).str.upper()
    unidade_upper = work["unidade"].astype(str).str.upper()
    unidade_norm_upper = work["unidade_norm"].astype(str).str.upper() if "unidade_norm" in work.columns else unidade_upper

    # ATENÇÃO PRIMÁRIA
    if categoria_norm in ["ATENÇÃO PRIMÁRIA", "ATENCAO PRIMARIA"]:
        return sum_mask(
            unidade_upper.eq("ATENÇÃO PRIMÁRIA") |
            unidade_upper.eq("ATENCAO PRIMARIA") |
            unidade_norm_upper.eq("ATENÇÃO PRIMÁRIA") |
            unidade_norm_upper.eq("ATENCAO PRIMARIA")
        )

    # ATENÇÃO ESPECIALIZADA
    if categoria_norm in ["ATENÇÃO ESPECIALIZADA", "ATENCAO ESPECIALIZADA"]:
        return sum_mask(
            unidade_upper.eq("ATENÇÃO ESPECIALIZADA") |
            unidade_upper.eq("ATENCAO ESPECIALIZADA") |
            unidade_norm_upper.eq("ATENÇÃO ESPECIALIZADA") |
            unidade_norm_upper.eq("ATENCAO ESPECIALIZADA") |
            painel_upper.str.contains("ESPECIALIZ", na=False) |
            painel_norm_upper.str.contains("ESPECIALIZ", na=False)
        )

    # AÇÕES COLETIVA
    if categoria_norm == "AÇÕES COLETIVA":
        return sum_mask(
            painel_upper.str.contains("AÇÃO COLET", na=False) |
            painel_upper.str.contains("ACAO COLET", na=False) |
            painel_norm_upper.str.contains("AÇÃO COLET", na=False) |
            painel_norm_upper.str.contains("ACAO COLET", na=False) |
            serie_upper.str.contains("AÇÃO COLET", na=False) |
            serie_upper.str.contains("ACAO COLET", na=False) |
            serie_norm_upper.str.contains("AÇÃO COLET", na=False) |
            serie_norm_upper.str.contains("ACAO COLET", na=False)
        )

    # ODONTOLOGIA
    if categoria_norm == "ODONTOLOGIA":
        return sum_mask(
            painel_upper.str.contains("ODONTO", na=False) |
            painel_norm_upper.str.contains("ODONTO", na=False) |
            serie_upper.str.contains("ODONTO", na=False) |
            serie_norm_upper.str.contains("ODONTO", na=False)
        )

    # ENFERMAGEM
    if categoria_norm == "ENFERMAGEM":
        return sum_mask(
            painel_upper.str.contains("ENFERM", na=False) |
            painel_norm_upper.str.contains("ENFERM", na=False) |
            serie_upper.str.contains("ENFERM", na=False) |
            serie_norm_upper.str.contains("ENFERM", na=False)
        )

    # MÉDICOS
    if categoria_norm == "MÉDICOS":
        return sum_mask(
            painel_upper.str.contains("MÉDIC", na=False) |
            painel_upper.str.contains("MEDIC", na=False) |
            painel_norm_upper.str.contains("MÉDIC", na=False) |
            painel_norm_upper.str.contains("MEDIC", na=False) |
            serie_upper.str.contains("MÉDIC", na=False) |
            serie_upper.str.contains("MEDIC", na=False) |
            serie_norm_upper.str.contains("MÉDIC", na=False) |
            serie_norm_upper.str.contains("MEDIC", na=False) |
            serie_upper.str.contains("CONSULTAS MÉDICAS", na=False) |
            serie_upper.str.contains("CONSULTAS MEDICAS", na=False) |
            serie_norm_upper.str.contains("CONSULTAS MÉDICAS", na=False) |
            serie_norm_upper.str.contains("CONSULTAS MEDICAS", na=False)
        )

    # EQUIPE MULTIDISCIPLINAR (EXCETO MÉDICOS)
    if categoria_norm == "EQUIPE MULTIDISCIPLINAR (EXCETO MÉDICOS)":
        return sum_mask(
            painel_upper.eq("NÍVEL SUPERIOR (EXCETO MÉDICO)") |
            painel_upper.eq("NIVEL SUPERIOR (EXCETO MEDICO)") |
            painel_norm_upper.eq("NÍVEL SUPERIOR (EXCETO MÉDICO)") |
            painel_norm_upper.eq("NIVEL SUPERIOR (EXCETO MEDICO)") |
            serie_upper.str.contains("NUTRI", na=False) |
            serie_upper.str.contains("PSICOLOG", na=False) |
            serie_upper.str.contains("ASSISTENTE SOCIAL", na=False) |
            serie_upper.str.contains("FISIOTERAP", na=False) |
            serie_norm_upper.str.contains("NUTRI", na=False) |
            serie_norm_upper.str.contains("PSICOLOG", na=False) |
            serie_norm_upper.str.contains("ASSISTENTE SOCIAL", na=False) |
            serie_norm_upper.str.contains("FISIOTERAP", na=False)
        )

    return 0.0


def build_metas_panel(data, metas_df):
    if metas_df is None or metas_df.empty:
        return pd.DataFrame()

    painel = metas_df.copy()
    painel["executado"] = pd.to_numeric(painel["executado"], errors="coerce").fillna(0.0)
    painel["meta"] = pd.to_numeric(painel["meta"], errors="coerce").fillna(0.0)

    painel["atingido_pct"] = painel.apply(
        lambda x: percent_atingido(x["executado"], x["meta"]),
        axis=1
    )
    painel["saldo"] = painel["executado"] - painel["meta"]
    painel["saldo_pct"] = painel.apply(
        lambda x: ((x["saldo"] / x["meta"]) * 100)
        if pd.notna(x["meta"]) and x["meta"] not in [0, None]
        else None,
        axis=1
    )

    return painel


def meta_status_badge(executado, meta):
    label, color, variacao_pct = status_meta(executado, meta)
    if variacao_pct is None:
        detalhe = "Sem comparativo"
    elif executado > meta:
        detalhe = f"+{format_pct_br(abs(variacao_pct))}"
    elif executado < meta:
        detalhe = f"Falta {format_pct_br(abs(variacao_pct))}"
    else:
        detalhe = "100,0%"

    return label, color, detalhe


def render_meta_card(categoria, executado, meta, atingido_pct, saldo_pct):
    status_label, status_color, detalhe = meta_status_badge(executado, meta)

    if saldo_pct is None:
        saldo_texto = "Sem cálculo"
    elif saldo_pct > 0:
        saldo_texto = f"Excedeu {format_pct_br(abs(saldo_pct))}"
    elif saldo_pct < 0:
        saldo_texto = f"Falta {format_pct_br(abs(saldo_pct))}"
    else:
        saldo_texto = "Meta exata"

    st.markdown(
        f"""
        <div style="
            background: linear-gradient(180deg, #FFFFFF 0%, #F8FAFC 100%);
            border: 1px solid #E2E8F0;
            border-radius: 22px;
            padding: 18px 18px 16px 18px;
            box-shadow: 0 10px 24px rgba(15, 23, 42, 0.08);
            min-height: 210px;
            margin-bottom: 14px;
        ">
            <div style="display:flex; justify-content:space-between; align-items:flex-start; gap:10px; margin-bottom:12px;">
                <div style="font-size:14px; font-weight:800; color:#0F172A; line-height:1.3;">{categoria}</div>
                <div style="background:{status_color}; color:#FFFFFF; font-size:11px; font-weight:700; padding:6px 10px; border-radius:999px; white-space:nowrap;">{status_label}</div>
            </div>
            <div style="display:grid; grid-template-columns:1fr 1fr; gap:10px; margin-bottom:12px;">
                <div style="background:#F8FAFC; border:1px solid #E2E8F0; border-radius:16px; padding:10px;">
                    <div style="font-size:11px; color:#64748B; text-transform:uppercase; font-weight:700;">Executado</div>
                    <div style="font-size:26px; font-weight:800; color:#0F172A; margin-top:4px;">{format_compact_number(executado)}</div>
                </div>
                <div style="background:#F8FAFC; border:1px solid #E2E8F0; border-radius:16px; padding:10px;">
                    <div style="font-size:11px; color:#64748B; text-transform:uppercase; font-weight:700;">Meta</div>
                    <div style="font-size:26px; font-weight:800; color:#0F172A; margin-top:4px;">{format_compact_number(meta)}</div>
                </div>
            </div>
            <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
                <div style="font-size:13px; color:#64748B; font-weight:700;">% atingido</div>
                <div style="font-size:20px; font-weight:800; color:#0F6CBD;">{format_pct_br(atingido_pct)}</div>
            </div>
            <div style="height:8px; background:#E2E8F0; border-radius:999px; overflow:hidden; margin-bottom:10px;">
                <div style="width:{0 if atingido_pct is None else min(max(atingido_pct,0),100)}%; height:100%; background:{status_color};"></div>
            </div>
            <div style="display:flex; justify-content:space-between; align-items:center; gap:10px;">
                <div style="font-size:12px; color:#64748B;">{saldo_texto}</div>
                <div style="font-size:12px; color:{status_color}; font-weight:700;">{detalhe}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )


def render_metas_page(data, metas_df, total_geral_map=None, meses_filtrados=None):
    st.subheader("Metas do Plano")

    if metas_df is None or metas_df.empty:
        st.warning("A aba 'METAS DO PLANO DE TRABALHO' não foi encontrada ou está vazia.")
        return

    painel_metas = build_metas_panel(data, metas_df)
    if painel_metas.empty:
        st.warning("Não foi possível montar o painel de metas com a base atual.")
        return

    resumo = (
        painel_metas.groupby("categoria", as_index=False)
        .agg({
            "executado": "sum",
            "meta": "sum",
        })
    )
    resumo["atingido_pct"] = resumo.apply(lambda x: percent_atingido(x["executado"], x["meta"]), axis=1)
    resumo["saldo"] = resumo["executado"] - resumo["meta"]
    resumo["saldo_pct"] = resumo.apply(
        lambda x: ((x["saldo"] / x["meta"]) * 100) if pd.notna(x["meta"]) and x["meta"] not in [0, None] else None,
        axis=1
    )

    total_meta = resumo["meta"].sum()
    total_executado_soma = float(resumo["executado"].sum())

    total_geral_por_mes = pd.Series(dtype=float)
    if total_geral_map:
        total_geral_por_mes = pd.Series(total_geral_map, dtype=float)
    elif "executado_total_geral" in metas_df.columns:
        total_geral_por_mes = (
            metas_df[["mes", "mes_label", "executado_total_geral"]]
            .dropna(subset=["mes_label", "executado_total_geral"])
            .groupby("mes_label", as_index=True)["executado_total_geral"]
            .max()
            .astype(float)
        )

    # Regra solicitada: usar somente o TOTAL GERAL do mes de referencia.
    # O mes de referencia deve seguir exatamente o filtro selecionado na sidebar.
    ordem_meses_label = [MESES_LABEL[m] for m in MESES]
    meses_ref = [m for m in ordem_meses_label if m in (meses_filtrados or [])]
    mes_referencia = meses_ref[-1] if meses_ref else None

    # Fallback defensivo: se nao houver filtro valido, usa o ultimo mes presente em metas_df.
    if mes_referencia is None and "mes" in metas_df.columns:
        meses_presentes = set(metas_df["mes"].dropna().astype(str).tolist())
        meses_disponiveis_ref = [m for m in MESES if m in meses_presentes]
        mes_referencia = MESES_LABEL.get(meses_disponiveis_ref[-1], meses_disponiveis_ref[-1]) if meses_disponiveis_ref else None

    total_executado = 0.0
    if mes_referencia and not total_geral_por_mes.empty:
        valor_mes_ref = total_geral_por_mes.get(mes_referencia)
        if valor_mes_ref is not None and pd.notna(valor_mes_ref):
            total_executado = float(valor_mes_ref)

    total_pct = percent_atingido(total_executado, total_meta)
    total_saldo_pct = ((total_executado - total_meta) / total_meta) * 100 if total_meta else None

    if mes_referencia:
        subtitle_executado_total = f"▲ total geral da planilha em {mes_referencia}"
    else:
        subtitle_executado_total = "▲ total geral da planilha (mês referência indisponível)"

    # Regras visuais solicitadas para os KPIs de metas
    if total_pct is not None and not pd.isna(total_pct) and total_pct > 99.99:
        pct_subtitle = "▲ executado em relação à meta"
        pct_subtitle_color = "#16A34A"
        pct_accent_color = "#22C55E"
    else:
        pct_subtitle = "▲ executado em relação à meta"
        pct_subtitle_color = "#EA580C"
        pct_accent_color = "#F97316"

    if total_saldo_pct is None or pd.isna(total_saldo_pct):
        saldo_subtitle = "• sem base de comparação"
        saldo_subtitle_color = "#64748B"
        saldo_accent_color = "#94A3B8"
    elif total_saldo_pct > 0:
        saldo_subtitle = "▲ acima da meta"
        saldo_subtitle_color = "#16A34A"
        saldo_accent_color = "#22C55E"
    elif total_saldo_pct < 0:
        saldo_subtitle = "▼ abaixo da meta"
        saldo_subtitle_color = "#DC2626"
        saldo_accent_color = "#EF4444"
    else:
        saldo_subtitle = "• em linha com a meta"
        saldo_subtitle_color = "#2563EB"
        saldo_accent_color = "#3B82F6"

    section_start("", "")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        top_kpi_card("Executado total", format_compact_number(total_executado), icon="📌", subtitle=subtitle_executado_total, accent_color="#22C55E", subtitle_color="#16A34A")
    with c2:
        top_kpi_card("Meta total", format_compact_number(total_meta), icon="🎯", subtitle="▲ somatório das metas", accent_color="#3B82F6", subtitle_color="#2563EB")
    with c3:
        top_kpi_card("% atingido", format_pct_br(total_pct), icon="📈", subtitle=pct_subtitle, accent_color=pct_accent_color, subtitle_color=pct_subtitle_color)
    with c4:
        top_kpi_card("Saldo percentual", format_pct_br(total_saldo_pct), icon="⚖️", subtitle=saldo_subtitle, accent_color=saldo_accent_color, subtitle_color=saldo_subtitle_color)
    section_end()

    serie_grafico = resumo.sort_values("atingido_pct", ascending=False).copy()
    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=serie_grafico["categoria"],
            y=serie_grafico["executado"],
            name="Executado",
            marker_color=SEMANTIC_COLORS["realizado"],
            hovertemplate="<b>Executado</b><br>%{x}<br>%{y:,.0f}<extra></extra>"
        )
    )
    fig.add_trace(
        go.Bar(
            x=serie_grafico["categoria"],
            y=serie_grafico["meta"],
            name="Meta",
            marker_color=SEMANTIC_COLORS["meta"],
            hovertemplate="<b>Meta</b><br>%{x}<br>%{y:,.0f}<extra></extra>"
        )
    )
    fig = apply_plotly_theme(
        fig,
        title="",
        subtitle="",
        yaxis_title="Quantidade",
        height=430,
        legend=True,
        legend_orientation="h",
        tick_angle=-25
    )
    fig.update_layout(barmode="group")
    st.markdown(
        """
        <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.7rem;
            padding: 0.15rem 0 0.65rem 0;
            overflow-x: auto;
        }

        .stTabs [data-baseweb="tab"] {
            background: linear-gradient(180deg, #F5FAFF 0%, #EDF5FC 100%);
            border: 1px solid #BDD5EA;
            border-radius: 14px;
            padding: 0.48rem 1.05rem;
            min-height: 48px;
            box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
            transition: all 0.18s ease;
        }

        .stTabs [data-baseweb="tab"]:hover {
            background: linear-gradient(180deg, #F0F7FF 0%, #E6F0FA 100%);
            border-color: #8CB8DF;
        }

        .stTabs [data-baseweb="tab"] p {
            font-size: 0.98rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            color: #0B3B69;
        }

        .stTabs [aria-selected="true"] {
            background: linear-gradient(180deg, #FFFFFF 0%, #F4F9FE 100%);
            border: 2px solid #0F6CBD;
            box-shadow: inset 0 -3px 0 #EF4444, 0 8px 18px rgba(15, 108, 189, 0.14);
            transform: translateY(-1px);
        }

        .stTabs [aria-selected="true"] p {
            color: #083055;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    tab_painel_meta, tab_metas_comparativo = st.tabs([
        "🎯 Painel por meta",
        "📊 Executado x Meta por categoria"
    ])
    with tab_painel_meta:
        cols = st.columns(2)
        for idx, row in enumerate(resumo.itertuples(index=False)):
            with cols[idx % 2]:
                render_meta_card(row.categoria, row.executado, row.meta, row.atingido_pct, row.saldo_pct)
    with tab_metas_comparativo:
        plot(fig, "metas_comparativo")

    tabela = resumo.copy()
    tabela["Executado"] = tabela["executado"].apply(format_compact_number)
    tabela["Meta"] = tabela["meta"].apply(format_compact_number)
    tabela["% Atingido"] = tabela["atingido_pct"].apply(format_pct_br)
    tabela["Saldo %"] = tabela["saldo_pct"].apply(format_pct_br)
    tabela = tabela[["categoria", "Executado", "Meta", "% Atingido", "Saldo %"]]
    tabela.columns = ["Meta do plano", "Executado", "Meta", "% atingido", "Saldo %"]

    with st.expander("Detalhamento das metas"):
        st.table(tabela.reset_index(drop=True))
        st.caption(
            f"Auditoria do executado total: total exibido = {format_compact_number(total_executado)} | "
            f"fonte = TOTAL GERAL | "
            f"mês referência = {mes_referencia if mes_referencia else '-'} | "
            f"soma categorias (somente conferência) = {format_compact_number(total_executado_soma)}"
        )
        st.caption("Observação: o executado é calculado com base nos dados disponíveis na planilha carregada. Categorias sem produção correspondente na base atual permanecem zeradas.")
def card(title, value, icon="📊", subtitle="Indicador consolidado"):
    value = clean_card_value(value)

    html = (
        '<div style="'
        'background: linear-gradient(135deg, #FFFFFF 0%, #F8FAFC 100%);'
        'border: 1px solid #E2E8F0;'
        'border-radius: 20px;'
        'padding: 18px 18px 16px 18px;'
        'box-shadow: 0 10px 24px rgba(15, 23, 42, 0.08);'
        'min-height: 130px;'
        'display: flex;'
        'flex-direction: column;'
        'justify-content: space-between;'
        '">'
            '<div style="display:flex; justify-content:space-between; align-items:flex-start; margin-bottom:10px;">'
                '<div>'
                    f'<div style="font-size: 13px; font-weight: 600; color: #64748B; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 6px;">{title}</div>'
                    f'<div style="font-size: 12px; color: #94A3B8;">{subtitle}</div>'
                '</div>'
                f'<div style="width: 42px; height: 42px; border-radius: 12px; background: linear-gradient(135deg, #DBEAFE 0%, #BFDBFE 100%); display:flex; align-items:center; justify-content:center; font-size: 20px;">{icon}</div>'
            '</div>'
            f'<div style="font-size: 32px; font-weight: 800; color: #0F172A; line-height: 1; margin-top: 8px;">{value}</div>'
        '</div>'
    )

    st.markdown(html, unsafe_allow_html=True)

def top_kpi_card(
    title,
    value,
    icon="📊",
    subtitle="Indicador consolidado",
    accent_color="#22C55E",
    subtitle_color="#64748B"
):
    value = clean_card_value(value)

    html = (
        '<div style="'
        'background: #F8FAFC;'
        'border: 1px solid #E2E8F0;'
        'border-top: 4px solid ' + accent_color + ';'
        'border-radius: 16px;'
        'padding: 14px 16px 12px 16px;'
        'box-shadow: 0 8px 18px rgba(15, 23, 42, 0.06);'
        'min-height: 148px;'
        '">'
            '<div style="'
            'width: 36px;'
            'height: 36px;'
            'border-radius: 10px;'
            'background: #EEF2FF;'
            'display: flex;'
            'align-items: center;'
            'justify-content: center;'
            'font-size: 18px;'
            'margin-bottom: 10px;'
            '">' + icon + '</div>'
            '<div style="font-size: 14.4px; letter-spacing: 1.1px; text-transform: uppercase; color: #475569; font-weight: 700;">'
            + title +
            '</div>'
            '<div style="font-size: 40px; font-weight: 800; color: #0F172A; line-height: 1.05; margin-top: 8px;">'
            + value +
            '</div>'
            '<div style="font-size: 15px; color: ' + subtitle_color + '; margin-top: 6px; font-weight: 600;">'
            + subtitle +
            '</div>'
        '</div>'
    )

    st.markdown(html, unsafe_allow_html=True)


def command_center_kpi_card(title, value, subtitle="", accent="#22C55E", icon="◎"):
    value = clean_card_value(value)
    html_kpi = (
        '<div style="'
        'background: linear-gradient(145deg, #0F172A 0%, #132A46 100%);'
        'border: 1px solid rgba(148,163,184,0.25);'
        'border-radius: 18px;'
        'padding: 14px 15px 12px 15px;'
        'box-shadow: 0 14px 28px rgba(2, 6, 23, 0.30);'
        'position: relative;'
        'overflow: hidden;'
        'min-height: 124px;'
        '">'
            '<div style="position:absolute; top:0; left:0; right:0; height:3px; background:' + accent + '; opacity:0.95;"></div>'
            '<div style="display:flex; align-items:center; justify-content:space-between; gap:10px; margin-bottom:10px;">'
                '<div style="font-size:11px; font-weight:700; letter-spacing:1px; text-transform:uppercase; color:#9FB6D3;">' + title + '</div>'
                '<div style="width:32px; height:32px; border-radius:10px; background:rgba(255,255,255,0.08); color:#E2E8F0; display:flex; align-items:center; justify-content:center; font-size:16px;">' + icon + '</div>'
            '</div>'
            '<div style="font-size:30px; font-weight:800; line-height:1; color:#F8FAFC; letter-spacing:-0.5px;">' + value + '</div>'
            '<div style="margin-top:8px; font-size:12px; color:#B8CAE0; line-height:1.35;">' + subtitle + '</div>'
        '</div>'
    )
    st.markdown(html_kpi, unsafe_allow_html=True)

def render_elegant_table(
    df,
    column_config=None,
    hide_index=True,
    use_container_width=True,
    row_height=42,
    key=None,
    status_columns=None,
    emphasis_columns=None,
    bar_columns=None,
    heatmap_columns=None,
    progress_columns=None,
    critical_condition=None,
):
    if df is None or df.empty:
        st.info("Sem dados para exibir.")
        return

    display_df = df.copy()
    status_columns = status_columns or []
    emphasis_columns = emphasis_columns or []
    bar_columns = bar_columns or []
    heatmap_columns = heatmap_columns or []
    progress_columns = progress_columns or []

    numeric_cache = {}
    for col in display_df.columns:
        numeric_cache[col] = pd.to_numeric(display_df[col], errors="coerce")

    def _to_float(value):
        try:
            return float(value)
        except Exception:
            return None

    def _format_value(col_name, value):
        if pd.isna(value):
            return "-"
        num_val = _to_float(value)
        if num_val is None:
            return str(value)

        col_norm = normalize_text(col_name)
        if "%" in col_name or "PARTICIPACAO" in col_norm:
            return f"{num_val:.1f}%"
        if "INDICE" in col_norm:
            return f"{num_val:.2f}"
        if "MEDIA" in col_norm or "IDEAIS" in col_norm or "GAP" in col_norm:
            return f"{num_val:.1f}"
        if abs(num_val) >= 1000:
            return format_int(int(round(num_val)))
        if float(num_val).is_integer():
            return str(int(num_val))
        return f"{num_val:.1f}"

    def _status_badge_class(value):
        raw = normalize_text(str(value or ""))
        if raw.startswith("DEFICIT") or raw.startswith("CRITICA"):
            return "badge-critical"
        if raw.startswith("EQUILIBR"):
            return "badge-ok"
        if raw.startswith("EXCESSO"):
            return "badge-warn"
        if raw.startswith("MODERADA") or raw.startswith("MEDIA"):
            return "badge-medium"
        if raw.startswith("BAIXA"):
            return "badge-low"
        return "badge-info"

    def _is_critical_row(row):
        if critical_condition is None:
            return False
        try:
            return bool(critical_condition(row))
        except Exception:
            return False

    def _norm_value(col_name, value):
        series = numeric_cache.get(col_name)
        if series is None:
            return 0.0
        valid = series.dropna()
        val = _to_float(value)
        if val is None or valid.empty:
            return 0.0
        vmin = float(valid.min())
        vmax = float(valid.max())
        if abs(vmax - vmin) < 1e-9:
            return 0.5
        return min(1.0, max(0.0, (val - vmin) / (vmax - vmin)))

    table_dom_id = (key or f"territorial_table_{abs(hash(tuple(display_df.columns)))}")
    table_dom_id = re.sub(r"[^A-Za-z0-9_-]", "_", str(table_dom_id))

    st.markdown(
        f"""
        <style>
        .analytics-table-shell#{table_dom_id} {{
            border: 1px solid #DFE7F3;
            border-radius: 18px;
            background: linear-gradient(180deg, #FFFFFF 0%, #F8FBFF 100%);
            box-shadow: 0 12px 28px rgba(15, 23, 42, 0.08);
            overflow: hidden;
            margin-top: 0.35rem;
            margin-bottom: 0.35rem;
        }}

        .analytics-table-shell#{table_dom_id} .analytics-table-scroll {{
            max-height: 560px;
            overflow: auto;
        }}

        .analytics-table-shell#{table_dom_id} table {{
            width: 100%;
            border-collapse: separate;
            border-spacing: 0;
            table-layout: auto;
        }}

        .analytics-table-shell#{table_dom_id} th {{
            position: sticky;
            top: 0;
            z-index: 2;
            background: linear-gradient(180deg, #F4F8FF 0%, #EAF1FC 100%);
            color: #0F1E33;
            text-transform: uppercase;
            letter-spacing: 0.07em;
            font-size: 11px;
            font-weight: 800;
            text-align: left;
            padding: 14px 16px;
            border-bottom: 1px solid #D7E2F1;
            white-space: nowrap;
        }}

        .analytics-table-shell#{table_dom_id} td {{
            padding: 14px 16px;
            border-bottom: 1px solid #ECF2FA;
            color: #0F172A;
            font-size: 13px;
            line-height: 1.35;
            vertical-align: middle;
            transition: background 180ms ease, transform 180ms ease;
        }}

        .analytics-table-shell#{table_dom_id} tbody tr {{
            animation: tblFadeIn 220ms ease both;
        }}

        .analytics-table-shell#{table_dom_id} tbody tr:nth-child(even) td {{
            background: #FBFDFF;
        }}

        .analytics-table-shell#{table_dom_id} tbody tr:hover td {{
            background: #EEF4FF;
        }}

        .analytics-table-shell#{table_dom_id} tbody tr.row-critical td {{
            background: #FFF7ED;
        }}

        .analytics-table-shell#{table_dom_id} tbody tr.row-critical td:first-child {{
            box-shadow: inset 4px 0 0 #DC2626;
        }}

        .analytics-table-shell#{table_dom_id} .td-emphasis {{
            font-weight: 800;
            color: #0B1F33;
        }}

        .analytics-table-shell#{table_dom_id} .td-muted {{
            color: #64748B;
            font-weight: 500;
        }}

        .analytics-table-shell#{table_dom_id} .cell-number-strong {{
            font-weight: 800;
            font-size: 13.4px;
            color: #0F172A;
        }}

        .analytics-table-shell#{table_dom_id} .badge {{
            display: inline-flex;
            align-items: center;
            justify-content: center;
            border-radius: 999px;
            padding: 5px 11px;
            font-size: 11.5px;
            font-weight: 800;
            letter-spacing: 0.01em;
            border: 1px solid transparent;
            box-shadow: 0 1px 4px rgba(15, 23, 42, 0.08);
            white-space: nowrap;
        }}

        .analytics-table-shell#{table_dom_id} .badge-critical {{
            background: #FEE2E2;
            color: #991B1B;
            border-color: #FCA5A5;
            box-shadow: 0 0 0 1px rgba(220, 38, 38, 0.14), 0 0 14px rgba(220, 38, 38, 0.16);
        }}
        .analytics-table-shell#{table_dom_id} .badge-ok {{
            background: #DCFCE7;
            color: #166534;
            border-color: #86EFAC;
        }}
        .analytics-table-shell#{table_dom_id} .badge-warn {{
            background: #FEF3C7;
            color: #92400E;
            border-color: #FCD34D;
        }}
        .analytics-table-shell#{table_dom_id} .badge-medium {{
            background: #FFEDD5;
            color: #9A3412;
            border-color: #FDBA74;
        }}
        .analytics-table-shell#{table_dom_id} .badge-low {{
            background: #E2E8F0;
            color: #334155;
            border-color: #CBD5E1;
        }}
        .analytics-table-shell#{table_dom_id} .badge-info {{
            background: #DBEAFE;
            color: #1D4ED8;
            border-color: #93C5FD;
        }}

        .analytics-table-shell#{table_dom_id} .metric-wrap {{
            display: flex;
            flex-direction: column;
            gap: 6px;
            min-width: 120px;
        }}

        .analytics-table-shell#{table_dom_id} .metric-value {{
            font-weight: 800;
            color: #0F172A;
            font-size: 12.5px;
        }}

        .analytics-table-shell#{table_dom_id} .metric-track {{
            height: 8px;
            background: #E5ECF6;
            border-radius: 999px;
            overflow: hidden;
        }}

        .analytics-table-shell#{table_dom_id} .metric-fill {{
            height: 100%;
            border-radius: 999px;
            transition: width 240ms ease;
            background: linear-gradient(90deg, #0EA5E9 0%, #2563EB 100%);
        }}

        .analytics-table-shell#{table_dom_id} .metric-fill.neg {{
            background: linear-gradient(90deg, #F97316 0%, #DC2626 100%);
        }}

        .analytics-table-shell#{table_dom_id} .progress-wrap {{
            display: flex;
            align-items: center;
            gap: 8px;
            min-width: 130px;
        }}

        .analytics-table-shell#{table_dom_id} .progress-track {{
            flex: 1;
            height: 8px;
            background: #E2E8F0;
            border-radius: 999px;
            overflow: hidden;
        }}

        .analytics-table-shell#{table_dom_id} .progress-fill {{
            height: 100%;
            border-radius: 999px;
            background: linear-gradient(90deg, #22C55E 0%, #0EA5E9 100%);
        }}

        .analytics-table-shell#{table_dom_id} .progress-fill.warn {{
            background: linear-gradient(90deg, #F59E0B 0%, #F97316 100%);
        }}

        .analytics-table-shell#{table_dom_id} .progress-fill.critical {{
            background: linear-gradient(90deg, #F97316 0%, #DC2626 100%);
        }}

        .analytics-table-shell#{table_dom_id} .table-footer {{
            padding: 10px 14px;
            color: #64748B;
            font-size: 11px;
            border-top: 1px solid #E6EDF7;
            background: #FAFCFF;
        }}

        @keyframes tblFadeIn {{
            from {{ opacity: 0; transform: translateY(2px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    headers_html = "".join([f"<th>{html.escape(str(col))}</th>" for col in display_df.columns])
    rows_html = []

    for _, row in display_df.iterrows():
        row_class = "row-critical" if _is_critical_row(row) else ""
        cell_parts = []

        for col in display_df.columns:
            raw_value = row[col]
            formatted = _format_value(col, raw_value)
            safe_text = html.escape(str(formatted))
            cell_classes = []
            cell_style = ""
            content = safe_text

            is_secondary = col not in emphasis_columns and col not in status_columns and col not in bar_columns and col not in progress_columns

            if col in emphasis_columns:
                cell_classes.append("td-emphasis")
            elif is_secondary:
                cell_classes.append("td-muted")

            if col in status_columns:
                badge_class = _status_badge_class(raw_value)
                content = f'<span class="badge {badge_class}">{html.escape(str(raw_value))}</span>'

            elif col in bar_columns:
                num_val = _to_float(raw_value)
                if num_val is not None:
                    valid = numeric_cache[col].dropna()
                    vmax = max(abs(float(valid.min() if not valid.empty else 0.0)), abs(float(valid.max() if not valid.empty else 0.0)), 1.0)
                    pct = min(100.0, max(0.0, (abs(num_val) / vmax) * 100.0))
                    fill_class = "metric-fill neg" if num_val < 0 else "metric-fill"
                    content = (
                        '<div class="metric-wrap">'
                        f'<div class="metric-value">{safe_text}</div>'
                        '<div class="metric-track">'
                        f'<div class="{fill_class}" style="width:{pct:.1f}%;"></div>'
                        '</div>'
                        '</div>'
                    )
                else:
                    content = '<span class="td-muted">-</span>'

            elif col in progress_columns:
                num_val = _to_float(raw_value)
                if num_val is not None:
                    normalized = min(1.0, max(0.0, num_val))
                    if num_val < 0.85:
                        p_class = "progress-fill critical"
                    elif num_val > 1.15:
                        p_class = "progress-fill warn"
                    else:
                        p_class = "progress-fill"
                    content = (
                        '<div class="progress-wrap">'
                        '<div class="progress-track">'
                        f'<div class="{p_class}" style="width:{normalized * 100.0:.1f}%;"></div>'
                        '</div>'
                        f'<span class="metric-value">{safe_text}</span>'
                        '</div>'
                    )

            elif col in heatmap_columns:
                intensity = _norm_value(col, raw_value)
                blue_alpha = 0.08 + (0.22 * intensity)
                red_alpha = 0.05 + (0.18 * intensity)
                if "GAP" in normalize_text(col) and _to_float(raw_value) is not None and _to_float(raw_value) >= 0:
                    cell_style = f"background: linear-gradient(90deg, rgba(239,68,68,{red_alpha:.3f}) 0%, rgba(255,255,255,0) 82%);"
                else:
                    cell_style = f"background: linear-gradient(90deg, rgba(14,165,233,{blue_alpha:.3f}) 0%, rgba(255,255,255,0) 82%);"
                cell_classes.append("cell-number-strong")

            elif _to_float(raw_value) is not None and col in ["Atendimentos", "Gap colaboradores", "Indice adequacao"]:
                cell_classes.append("cell-number-strong")

            cls_str = " ".join(cell_classes).strip()
            class_attr = f' class="{cls_str}"' if cls_str else ""
            style_attr = f' style="{cell_style}"' if cell_style else ""
            cell_parts.append(f"<td{class_attr}{style_attr}>{content}</td>")

        rows_html.append(f"<tr class=\"{row_class}\">{''.join(cell_parts)}</tr>")

    table_html = (
        f'<div id="{table_dom_id}" class="analytics-table-shell">'
        '<div class="analytics-table-scroll">'
        '<table>'
        f'<thead><tr>{headers_html}</tr></thead>'
        f'<tbody>{"".join(rows_html)}</tbody>'
        '</table>'
        '</div>'
        f'<div class="table-footer">{len(display_df)} registros | leitura executiva territorial</div>'
        '</div>'
    )

    st.markdown(table_html, unsafe_allow_html=True)


def section_start(title, subtitle="", theme=None):
    theme_class = f" section-card--{theme}" if theme else ""
    st.markdown(
        f"""
        <div class="section-card{theme_class}">
            <div class="section-title">{title}</div>
            <div class="section-subtitle">{subtitle}</div>
        """,
        unsafe_allow_html=True
    )


def section_end():
    st.markdown("</div>", unsafe_allow_html=True)


def hero_header(page_title, source_name, meses_selecionados):
    st.markdown(
        """
        <style>
        .hero-wrap {
            background: linear-gradient(135deg, #0F172A 0%, #12324A 50%, #0F6CBD 100%);
            border: 1px solid rgba(255,255,255,0.05);
            border-radius: 24px;
            padding: 1.15rem 1.25rem;
            margin-top: 0.2rem;
            margin-bottom: 1.1rem;
            box-shadow: 0 16px 36px rgba(15, 23, 42, 0.16);
        }

        .hero-title {
            color: #FFFFFF;
            font-size: 1.9rem;
            font-weight: 800;
            letter-spacing: -0.8px;
            margin-bottom: 0.2rem;
        }

        .hero-subtitle {
            color: rgba(255,255,255,0.82);
            font-size: 0.98rem;
            margin-bottom: 1rem;
        }

        .logo-slot {
            display: flex;
            align-items: center;
            justify-content: center;
            height: 100%;
            padding-top: 0.35rem;
        }
        .logo-left {
            display: flex;
            align-items: center;
            justify-content: center;
            height: 100%;
            margin-top: 80px;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    col1, col2, col3 = st.columns([1.2, 6, 1.2])

    with col1:
        st.markdown('<div class="logo-left">', unsafe_allow_html=True)
        try:
            st.image(str(LOGO_PATRIS), width=315)
        except Exception:
            st.empty()
        st.markdown("</div>", unsafe_allow_html=True)

    with col2:
        st.markdown(
            f"""
            <div class="hero-wrap">
                <div class="hero-title" style="width: 100%; text-align: center;">Painel de Gestão Patris</div>
                <div class="hero-subtitle" style="width: 100%; text-align: center;">
                   Gestão estratégica da produção assistencial e desempenho operacional
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

    with col3:
        st.markdown('<div class="logo-slot">', unsafe_allow_html=True)
        try:
            st.image(str(LOGO_PREFEITURA), width=315)
        except Exception:
            st.empty()
        st.markdown("</div>", unsafe_allow_html=True)
SEMANTIC_COLORS = {
    # identidade base
    "primary": "#0F6CBD",
    "primary_soft": "#93C5FD",
    "secondary": "#0F172A",

    # estados
    "success": "#16A34A",
    "warning": "#F59E0B",
    "danger": "#DC2626",
    "neutral": "#64748B",
    "info": "#0891B2",

    # leitura executiva
    "realizado": "#0F6CBD",
    "realizado_soft": "#93C5FD",
    "media": "#38BDF8",
    "meta": "#94A3B8",
    "alerta": "#DC2626",
    "bom": "#16A34A",
    "atencao": "#F59E0B",
    "critico": "#DC2626",

    # apoio visual
    "grid": "rgba(148,163,184,0.14)",
    "axis": "#076BF7",
    "text": "#CFD7E2",
    "title": "#F6F7FB",
    "plot_bg": "#071224",

    # séries neutras
    "series_1": "#0F6CBD",
    "series_2": "#16A34A",
    "series_3": "#F59E0B",
    "series_4": "#DC2626",
    "series_5": "#7C3AED",
    "series_6": "#0891B2",
    "series_7": "#64748B",
}

APP_COLORS = {
    "primary": SEMANTIC_COLORS["primary"],
    "primary_soft": SEMANTIC_COLORS["primary_soft"],
    "secondary": SEMANTIC_COLORS["secondary"],
    "success": SEMANTIC_COLORS["success"],
    "warning": SEMANTIC_COLORS["warning"],
    "danger": SEMANTIC_COLORS["danger"],
    "neutral": SEMANTIC_COLORS["neutral"],
    "grid": SEMANTIC_COLORS["grid"],
    "axis": SEMANTIC_COLORS["axis"],
    "text": SEMANTIC_COLORS["text"],
    "title": SEMANTIC_COLORS["title"],
    "plot_bg": SEMANTIC_COLORS["plot_bg"],
}

DEFAULT_CHART_COLORS = [
    SEMANTIC_COLORS["series_1"],
    SEMANTIC_COLORS["series_2"],
    SEMANTIC_COLORS["series_3"],
    SEMANTIC_COLORS["series_4"],
    SEMANTIC_COLORS["series_5"],
    SEMANTIC_COLORS["series_6"],
    SEMANTIC_COLORS["series_7"],
]

def apply_visual_theme(theme_name):
    themes = {
        "Portal Clínico (Azul)": {
            "palette": {
                "primary": "#0F6CBD",
                "primary_soft": "#93C5FD",
                "success": "#16A34A",
                "warning": "#F59E0B",
                "danger": "#DC2626",
                "neutral": "#64748B",
                "realizado": "#0F6CBD",
                "realizado_soft": "#93C5FD",
                "media": "#38BDF8",
                "meta": "#94A3B8",
                "grid": "rgba(148,163,184,0.14)",
                "axis": "#076BF7",
                "text": "#CFD7E2",
                "title": "#F6F7FB",
                "plot_bg": "#071224",
                "series": ["#0F6CBD", "#16A34A", "#F59E0B", "#DC2626", "#7C3AED", "#0891B2", "#64748B"],
            },
            "css": f"""
            <style>
            [data-testid="stAppViewContainer"] {{
                background-image: linear-gradient(rgba(239, 248, 255, 0.72), rgba(239, 248, 255, 0.82)), url("data:image/png;base64,{BACKGROUND_BASE64}") !important;
                background-size: cover !important;
                background-position: center !important;
                background-attachment: fixed !important;
                background-color: #EEF7FC !important;
            }}
            section[data-testid="stSidebar"] {{
                background: linear-gradient(180deg, #0F4C81 0%, #0B2E4E 100%) !important;
            }}
            section[data-testid="stSidebar"] * {{ color: #F8FAFC !important; }}
            .hero-wrap {{ background: linear-gradient(135deg, #0F172A 0%, #12324A 50%, #0F6CBD 100%) !important; }}
            div[data-testid="stPlotlyChart"] {{ background: #071224 !important; }}
            </style>
            """,
        },
        "Pro Analytics (Escuro)": {
            "palette": {
                "primary": "#00C2FF",
                "primary_soft": "#67E8F9",
                "success": "#00E5A0",
                "warning": "#F59E0B",
                "danger": "#EF4444",
                "neutral": "#94A3B8",
                "realizado": "#00C2FF",
                "realizado_soft": "#67E8F9",
                "media": "#22D3EE",
                "meta": "#CBD5E1",
                "grid": "rgba(148,163,184,0.18)",
                "axis": "#60A5FA",
                "text": "#E2E8F0",
                "title": "#F8FAFC",
                "plot_bg": "#0D1321",
                "series": ["#00C2FF", "#00E5A0", "#F59E0B", "#EF4444", "#A78BFA", "#22D3EE", "#94A3B8"],
            },
            "css": """
            <style>
            [data-testid="stAppViewContainer"] {
                background: radial-gradient(circle at 20% 0%, #16233A 0%, #0A0E1A 45%, #090D18 100%) !important;
            }
            [data-testid="stMain"] {
                background: transparent !important;
            }
            section[data-testid="stSidebar"] {
                background: linear-gradient(180deg, #0D1321 0%, #0A0E1A 100%) !important;
                border-right: 1px solid rgba(255,255,255,0.08) !important;
            }
            section[data-testid="stSidebar"] * { color: #E2E8F0 !important; }
            h1, h2, h3, p, label, .stMarkdown, .stCaption { color: #E2E8F0 !important; }
            .hero-wrap {
                background: linear-gradient(125deg, #0D1A2E 0%, #0F2340 55%, #0A66CC 100%) !important;
                border: 1px solid rgba(0,194,255,0.22) !important;
                box-shadow: 0 0 36px rgba(0,194,255,0.14) !important;
            }
            .hero-subtitle, .hero-chip { color: #E2E8F0 !important; }
            div[data-testid="stMetric"] {
                background: #0D1321 !important;
                border: 1px solid rgba(255,255,255,0.08) !important;
            }
            details {
                background: #0D1321 !important;
                border: 1px solid rgba(255,255,255,0.10) !important;
            }
            div[data-testid="stPlotlyChart"] {
                background: #0D1321 !important;
                border: 1px solid rgba(255,255,255,0.08) !important;
            }
            </style>
            """,
        },
        "Healthcare Clean (Verde)": {
            "palette": {
                "primary": "#0EA472",
                "primary_soft": "#86EFAC",
                "success": "#16A34A",
                "warning": "#F59E0B",
                "danger": "#DC2626",
                "neutral": "#64748B",
                "realizado": "#0EA472",
                "realizado_soft": "#86EFAC",
                "media": "#10B981",
                "meta": "#94A3B8",
                "grid": "rgba(148,163,184,0.20)",
                "axis": "#0EA472",
                "text": "#334155",
                "title": "#0F172A",
                "plot_bg": "#F8FBF9",
                "series": ["#0EA472", "#3B82F6", "#16A34A", "#F59E0B", "#DC2626", "#14B8A6", "#64748B"],
            },
            "css": f"""
            <style>
            [data-testid="stAppViewContainer"] {{
                background-image: linear-gradient(rgba(246, 252, 248, 0.94), rgba(246, 252, 248, 0.97)), url("data:image/png;base64,{BACKGROUND_BASE64}") !important;
                background-size: cover !important;
                background-position: center !important;
                background-color: #F3FBF6 !important;
            }}
            section[data-testid="stSidebar"] {{
                background: linear-gradient(180deg, #0B7A5A 0%, #065F46 100%) !important;
            }}
            section[data-testid="stSidebar"] * {{ color: #ECFDF5 !important; }}
            .hero-wrap {{
                background: linear-gradient(120deg, #065F46 0%, #0EA472 60%, #10B981 100%) !important;
                box-shadow: 0 12px 30px rgba(6,95,70,0.22) !important;
            }}
            div[data-testid="stMetric"] {{
                background: linear-gradient(180deg, #FFFFFF 0%, #F8FFFB 100%) !important;
                border: 1px solid #DCFCE7 !important;
            }}
            details {{
                background: #FFFFFF !important;
                border: 1px solid #DCFCE7 !important;
            }}
            div[data-testid="stPlotlyChart"] {{
                background: #F8FBF9 !important;
                border: 1px solid #DCFCE7 !important;
            }}
            </style>
            """,
        },
    }

    selected = themes.get(theme_name, themes["Portal Clínico (Azul)"])
    palette = selected["palette"]

    SEMANTIC_COLORS.update({
        "primary": palette["primary"],
        "primary_soft": palette["primary_soft"],
        "success": palette["success"],
        "warning": palette["warning"],
        "danger": palette["danger"],
        "neutral": palette["neutral"],
        "realizado": palette["realizado"],
        "realizado_soft": palette["realizado_soft"],
        "media": palette["media"],
        "meta": palette["meta"],
        "grid": palette["grid"],
        "axis": palette["axis"],
        "text": palette["text"],
        "title": palette["title"],
        "plot_bg": palette["plot_bg"],
        "series_1": palette["series"][0],
        "series_2": palette["series"][1],
        "series_3": palette["series"][2],
        "series_4": palette["series"][3],
        "series_5": palette["series"][4],
        "series_6": palette["series"][5],
        "series_7": palette["series"][6],
    })

    APP_COLORS.update({
        "primary": SEMANTIC_COLORS["primary"],
        "primary_soft": SEMANTIC_COLORS["primary_soft"],
        "secondary": SEMANTIC_COLORS["secondary"],
        "success": SEMANTIC_COLORS["success"],
        "warning": SEMANTIC_COLORS["warning"],
        "danger": SEMANTIC_COLORS["danger"],
        "neutral": SEMANTIC_COLORS["neutral"],
        "grid": SEMANTIC_COLORS["grid"],
        "axis": SEMANTIC_COLORS["axis"],
        "text": SEMANTIC_COLORS["text"],
        "title": SEMANTIC_COLORS["title"],
        "plot_bg": SEMANTIC_COLORS["plot_bg"],
    })

    DEFAULT_CHART_COLORS[:] = palette["series"]
    st.markdown(selected["css"], unsafe_allow_html=True)

def semantic_color(name, default=None):
    if not name:
        return default or SEMANTIC_COLORS["neutral"]

    key = str(name).strip().upper()

    # meta / referência
    if "META" in key:
        return SEMANTIC_COLORS["meta"]

    # médias
    if "MÉDIA" in key or "MEDIA" in key:
        return SEMANTIC_COLORS["media"]

    # alertas / eventos críticos
    if "ÓBITO" in key or "OBITO" in key:
        return SEMANTIC_COLORS["danger"]

    # risco
    if key in RISK_COLORS:
        return RISK_COLORS[key]

    # séries principais comuns
    if "ATENDIMENTOS MÉDICOS" in key:
        return SEMANTIC_COLORS["realizado"]

    if "PACIENTES RECEPCIONADOS" in key:
        return SEMANTIC_COLORS["realizado_soft"]

    if "MÉDIA GERAL" in key or "MEDIA GERAL" in key:
        return SEMANTIC_COLORS["media"]

    return default or SEMANTIC_COLORS["neutral"]

def build_semantic_color_map(series_list):
    palette = [
        SEMANTIC_COLORS["series_1"],
        SEMANTIC_COLORS["series_2"],
        SEMANTIC_COLORS["series_3"],
        SEMANTIC_COLORS["series_4"],
        SEMANTIC_COLORS["series_5"],
        SEMANTIC_COLORS["series_6"],
        SEMANTIC_COLORS["series_7"],
    ]

    color_map = {}
    fallback_idx = 0

    for serie in series_list:
        forced = semantic_color(serie, default=None)
        if forced is not None and forced != SEMANTIC_COLORS["neutral"]:
            color_map[serie] = forced
        else:
            color_map[serie] = palette[fallback_idx % len(palette)]
            fallback_idx += 1

    return color_map

def apply_plotly_theme(
    fig,
    title=None,
    subtitle=None,
    yaxis_title="",
    height=360,
    legend=True,
    legend_orientation="h",
    tick_angle=0
):
    full_title = ""
    if title:
        full_title = f"<span style='font-weight:800'>{title}</span>"
        if subtitle:
            full_title += f"<br><span style='font-size:12px; color:#64748B; font-weight:400'>{subtitle}</span>"

    fig.update_layout(
        title=dict(
            text=full_title,
            x=0.0,
            xanchor="left",
            y=0.97,
            yanchor="top"
        ),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor=APP_COLORS["plot_bg"],
        font=dict(
            family="Inter, Segoe UI, Arial, sans-serif",
            color=APP_COLORS["text"],
            size=12
        ),
        title_font=dict(
            color=APP_COLORS["title"],
            size=20
        ),
        colorway=DEFAULT_CHART_COLORS,
        height=height,
        margin=dict(l=30, r=18, t=84, b=72),
        hoverlabel=dict(
            bgcolor="#FFFFFF",
            bordercolor="#CBD5E1",
            font=dict(color="#0F172A", size=12)
        ),
        bargap=0.30,
        bargroupgap=0.10
    )

    first_x_len = 0
    try:
        if fig.data and hasattr(fig.data[0], "x") and fig.data[0].x is not None:
            first_x_len = len(fig.data[0].x)
    except Exception:
        first_x_len = 0

    auto_tick_angle = 0 if first_x_len <= 5 else -45

    fig.update_xaxes(
        title_text="",
        showgrid=False,
        showline=False,
        zeroline=False,
        tickfont=dict(color="#64748B", size=10.5),
        tickangle=auto_tick_angle if tick_angle == 0 else tick_angle,
        automargin=True,
        ticklabeloverflow="allow"
    )

    fig.update_yaxes(
        title_text=yaxis_title,
        showgrid=True,
        gridcolor=APP_COLORS["grid"],
        gridwidth=1,
        zeroline=False,
        showline=False,
        tickfont=dict(color="#64748B", size=11),
        title_font=dict(color="#64748B", size=12),
        automargin=True
    )

    if legend:
        fig.update_layout(
            showlegend=True,
            legend=dict(
                title_text="",
                orientation=legend_orientation,
                yanchor="bottom",
                y=1.02,
                xanchor="left",
                x=0,
                font=dict(size=11, color="#64748B"),
                traceorder="normal",
                itemsizing="constant"
            )
        )
    else:
        fig.update_layout(showlegend=False)

    return fig
def chart_subtitle(df, unidade=None):
    meses = [m for m in df.get("mes_label", pd.Series(dtype=str)).dropna().unique().tolist()]
    if not meses:
        periodo_txt = "Sem período"
    elif len(meses) == 1:
        periodo_txt = meses[0]
    else:
        periodo_txt = f"{meses[0]} a {meses[-1]}"

    if unidade:
        return f"{unidade} • {periodo_txt}"
    return periodo_txt

def ordered_month_labels(df):
    if df is None or df.empty or "mes" not in df.columns:
        return []

    meses_validos = (
        df["mes"]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    meses_ordenados = [m for m in MESES if m in meses_validos]
    return [MESES_LABEL.get(m, m) for m in meses_ordenados]


def apply_month_axis_order(fig, df):
    ordered_labels = ordered_month_labels(df)
    if not ordered_labels:
        return fig

    fig.update_xaxes(
        type="category",
        categoryorder="array",
        categoryarray=ordered_labels,
        tickmode="array",
        tickvals=ordered_labels,
        ticktext=ordered_labels
    )
    return fig

def truncate_series_name(name, max_len=28):
    name = str(name)
    return name if len(name) <= max_len else name[:max_len-3] + "..."


def clean_trace_names(fig):
    return fig


def smart_legend_visibility(df, max_series_horizontal=5):
    n = df["serie"].dropna().nunique() if "serie" in df.columns else 0
    return n > 1, ("h" if n <= max_series_horizontal else "h") 
def line_with_optional_meta(
    df,
    title,
    main_series=None,
    unit_suffix="",
    prefix="line",
    unidade=None,
    subtitle=None,
):
    work = df.dropna(subset=["valor_num"]).copy()
    if work.empty:
        st.info("Sem dados para este gráfico.")
        return

    fig = go.Figure()

    if main_series:
        main = work[work["serie_norm"] == str(main_series).upper()]
        if not main.empty:
            main_color = semantic_color(main_series, default=SEMANTIC_COLORS["realizado"])

            fig.add_trace(
                go.Scatter(
                    x=main["mes_label"],
                    y=main["valor_num"],
                    mode="lines+markers",
                    name=str(main_series).title(),
                    line=dict(color=main_color, width=3.5),
                    marker=dict(size=7, color=main_color),
                    customdata=main["valor_num"].apply(format_hours_hms),
                    hovertemplate="<b>%{fullData.name}</b><br>Mês: %{x}<br>Tempo: %{customdata}<extra></extra>"
                )
            )

        others = work[
            (~work["serie_norm"].eq(str(main_series).upper())) &
            (~work["serie_norm"].eq("META"))
        ]

        for serie in others["serie"].dropna().unique().tolist():
            temp = others[others["serie"] == serie]
            serie_color = semantic_color(serie, default=SEMANTIC_COLORS["neutral"])

            fig.add_trace(
                go.Scatter(
                    x=temp["mes_label"],
                    y=temp["valor_num"],
                    mode="lines+markers",
                    name=str(serie),
                    line=dict(color=serie_color, width=2),
                    marker=dict(size=5, color=serie_color),
                    opacity=0.65,
                    hovertemplate="<b>%{fullData.name}</b><br>Mês: %{x}<br>Valor: %{y:,.1f}<extra></extra>"
                )
            )
    else:
        series = work["serie"].dropna().unique().tolist()
        color_map = build_semantic_color_map(series)

        for serie in series:
            temp = work[work["serie"] == serie]
            serie_color = color_map.get(serie, SEMANTIC_COLORS["neutral"])

            fig.add_trace(
                go.Scatter(
                    x=temp["mes_label"],
                    y=temp["valor_num"],
                    mode="lines+markers",
                    name=str(serie),
                    line=dict(color=serie_color, width=2.4),
                    marker=dict(size=6, color=serie_color),
                    customdata=temp["valor_num"].apply(format_hours_hms),
                    hovertemplate="<b>%{fullData.name}</b><br>Mês: %{x}<br>Tempo: %{customdata}<extra></extra>"
                )
            )

    meta = work[work["serie_norm"] == "META"]
    if not meta.empty:
        meta_color = SEMANTIC_COLORS["meta"]
        fig.add_trace(
            go.Scatter(
                x=meta["mes_label"],
                y=meta["valor_num"],
                mode="lines+markers",
                name="Meta",
                line=dict(color=meta_color, width=2, dash="dash"),
                marker=dict(size=5, color=meta_color),
                customdata=meta["valor_num"].apply(format_hours_hms),
                hovertemplate="<b>Meta</b><br>Mês: %{x}<br>Tempo: %{customdata}<extra></extra>"
            )
        )

    fig = apply_plotly_theme(
        fig,
        title=title,
        subtitle=chart_subtitle(work, unidade) if subtitle is None else subtitle,
        yaxis_title=unit_suffix,
        height=350,
        legend=True,
        legend_orientation="h"
    )

    fig = apply_month_axis_order(fig, work)

    plot(fig, prefix)


def grouped_bar(
    df,
    title,
    color_map=None,
    barmode="group",
    unit_suffix="",
    prefix="bar",
    unidade=None,
    subtitle=None,
    show_status_chip=True,
):
    work = df.dropna(subset=["valor_num"]).copy()
    if work.empty:
        st.info("Sem dados para este gráfico.")
        return

    if "serie" in work.columns:
        serie_txt = work["serie"].astype(str).str.strip()
        invalid_names = serie_txt.str.lower().isin(["", "nan", "none", "undefined"])
        serie_txt = serie_txt.mask(invalid_names, "Sem classificação")
        work["serie"] = serie_txt

    fig = px.bar(
        work,
        x="mes_label",
        y="valor_num",
        color="serie",
        barmode=barmode,
        color_discrete_map=color_map or {}
    )

    fig.update_traces(
        marker_line_width=0,
        hovertemplate="<b>%{fullData.name}</b><br>Mês: %{x}<br>Valor: %{y:,.0f}<extra></extra>"
    )

    fig = apply_plotly_theme(
        fig,
        title=title,
        subtitle=chart_subtitle(work, unidade) if subtitle is None else subtitle,
        yaxis_title=unit_suffix,
        height=380,
        legend=True,
        legend_orientation="h"
    )

    fig = apply_month_axis_order(fig, work)

    plot(fig, prefix, show_status_chip=show_status_chip)


def stacked_bar(
    df,
    title,
    color_map=None,
    as_percent=False,
    prefix="stack",
    unidade=None
):
    work = df.dropna(subset=["valor_num"]).copy()
    if work.empty:
        st.info("Sem dados para este gráfico.")
        return

    pivot = work.pivot_table(
        index="mes_label",
        columns="serie",
        values="valor_num",
        aggfunc="sum"
    ).fillna(0)

    if as_percent:
        pivot = pivot.div(pivot.sum(axis=1).replace(0, pd.NA), axis=0) * 100

    fig = go.Figure()
    for serie in pivot.columns:
        fig.add_trace(
            go.Bar(
                x=pivot.index,
                y=pivot[serie],
                name=str(serie),
                marker_color=(color_map or {}).get(serie),
                hovertemplate=f"<b>{serie}</b><br>Mês: %{{x}}<br>Valor: %{{y:.1f}}{'%' if as_percent else ''}<extra></extra>"
            )
        )

    fig = apply_plotly_theme(
        fig,
        title=title,
        subtitle=chart_subtitle(work, unidade),
        yaxis_title="Percentual (%)" if as_percent else "Quantidade",
        height=390,
        legend=True,
        legend_orientation="h"
    )

    fig.update_layout(barmode="stack")

    if as_percent:
        fig.update_yaxes(range=[0, 100])

    fig = apply_month_axis_order(fig, work)

    plot(fig, prefix)


def pie_latest(df, title, color_map=None, prefix="pie", unidade=None):
    work = df.dropna(subset=["valor_num"]).copy()
    if work.empty:
        st.info("Sem dados para este gráfico.")
        return

    latest_mes = work["mes"].dropna().max()
    latest = work[work["mes"] == latest_mes].copy()
    if latest.empty:
        st.info("Sem dados para este gráfico.")
        return

    fig = px.pie(
        latest,
        names="serie",
        values="valor_num",
        color="serie",
        color_discrete_map=color_map or {}
    )

    fig.update_traces(
        textposition="inside",
        textinfo="percent",
        hole=0.45,
        sort=False,
        hovertemplate="<b>%{label}</b><br>Valor: %{value:,.0f}<br>Participação: %{percent}<extra></extra>"
    )

    fig = apply_plotly_theme(
        fig,
        title=title,
        subtitle=f"{unidade + ' • ' if unidade else ''}{MESES_LABEL.get(latest_mes, latest_mes)}",
        height=380,
        legend=True,
        legend_orientation="h"
    )

    plot(fig, prefix)
def render_upa_page(df, unidade, meses_filtrados=None):
    st.markdown(
        """
        <style>
        .upa-title-card {
            background: #FFFFFF;
            border-radius: 12px;
            border: 1px solid #D6E4F0;
            border-left: 6px solid #0F6CBD;
            padding: 0.75rem 1rem;
            margin: 0.1rem 0 0.9rem 0;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.06);
        }

        .upa-title-main {
            color: #0F2A43;
            font-size: 1.22rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            margin: 0;
        }

        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        f"""
        <div class="upa-title-card">
            <div class="upa-title-main">🚑 {unidade}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    def _mf(panel_df):
        if panel_df is None or panel_df.empty:
            return panel_df
        if meses_filtrados and "mes_label" in panel_df.columns:
            return panel_df[panel_df["mes_label"].isin(meses_filtrados)].copy()
        return panel_df

    recep = _mf(filter_panel(df, unidade, "PACIENTES RECEPCIONADOS"))
    atend_med = _mf(filter_panel(df, unidade, "ATENDIMENTOS MÉDICOS"))
    risco = _mf(filter_panel(df, unidade, "ATENDIMENTOS POR CLASSIFICAÇÃO DE RISCO"))
    perc_risco = _mf(filter_panel(df, unidade, "PERCENTUAL DE ATENDIMENTOS POR CLASSIFICAÇÃO DE RISCOS"))
    espera = _mf(filter_panel(df, unidade, "TEMPO DE ESPERA PARA CLASSIFICAÇÃO DE RISCO"))
    tempo_med = _mf(filter_panel(df, unidade, "TEMPO MÉDIO DE ESPERA DE ATENDIMENTO MÉDICO POR CLASSIFICAÇÃO DE RISCO"))
    intern = _mf(filter_panel(df, unidade, "TEMPO DE PERMANÊNCIA DE PACIENTES INTERNADOS"))
    semint = _mf(filter_panel(df, unidade, "TEMPO DE PERMANÊNCIA DE PACIENTES SEM INTERNAÇÃO"))
    transf = _mf(filter_panel(df, unidade, "TRANSFERÊNCIAS (REMOÇÕES)"))
    exames = _mf(filter_panel(df, unidade, "EXAMES INTERNOS"))
    faixa = _mf(filter_panel(df, unidade, "ATENDIMENTOS DIVIDIDOS POR FAIXA ETARIA"))
    origem = _mf(filter_panel(df, unidade, "ATENDIMENTOS DE  PACIENTES"))
    obitos = _mf(filter_panel(df, unidade, "ÓBITOS"))

    section_start("", "")
    c1, c2, c3, c4 = st.columns(4)

    with c1:
        total_recep = recep[
           recep["serie_norm"].isin([
               "PACIENTES RECEPCIONADOS"
            ])
        ]["valor_num"].sum()

        top_kpi_card(
            "Pacientes recepcionados",
            format_int(total_recep),
            icon="👥",
            subtitle="▲ volume total no período",
            accent_color="#22C55E",
            subtitle_color="#16A34A"
        )

    with c2:
        total_atend_med = metric_sum(
            atend_med,
            exclude_series_norm=["META", "MÉDIA DIÁRIA", "MEDIA DIÁRIA", "MEDIA DIARIA", "TOTAL"]
        )

        top_kpi_card(
            "Atendimentos médicos",
            format_int(total_atend_med or 0),
            icon="🩺",
            subtitle="▲ produção médica consolidada",
            accent_color="#3B82F6",
            subtitle_color="#2563EB"
        )

    with c3:
        top_kpi_card(
            "Óbitos",
            format_int(obitos["valor_num"].sum()),
            icon="⚠️",
            subtitle="▼ ocorrências registradas",
            accent_color="#EF4444",
            subtitle_color="#DC2626"
        )

    with c4:
        top_kpi_card(
            "Exames internos",
            format_int(exames[~exames["serie_norm"].eq("TOTAL")]["valor_num"].sum()),
            icon="🧪",
            subtitle="▲ procedimentos realizados",
            accent_color="#F97316",
            subtitle_color="#EA580C"
        )
    section_end()

    st.markdown(
        """
        <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.6rem;
            padding: 0.2rem 0.1rem 0.5rem 0.1rem;
        }

        .stTabs [data-baseweb="tab"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.12) 0%, rgba(15,108,189,0.05) 100%);
            border: 1px solid rgba(15,108,189,0.35);
            border-radius: 12px;
            padding: 0.5rem 0.95rem;
            min-height: 44px;
            box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
        }

        .stTabs [data-baseweb="tab"] p {
            font-size: 1rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            color: #0b3b69;
        }

        .stTabs [aria-selected="true"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.24) 0%, rgba(15,108,189,0.12) 100%);
            border: 2px solid #0F6CBD;
            box-shadow: 0 0 0 1px rgba(15,108,189,0.15), 0 8px 18px rgba(15, 108, 189, 0.18);
            transform: translateY(-1px);
        }

        .stTabs [aria-selected="true"] p {
            color: #083055;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    tab_prod, tab_risco, tab_perm, tab_exames = st.tabs([
        "Produção assistencial",
        "Risco e tempo assistencial",
        "Permanência, apoio e desfechos",
        "Exames internos e faixa etária",
    ])

    with tab_prod:
        section_start("Produção assistencial", "")
        col1, col2 = st.columns(2)

        with col1:
            fig = go.Figure()

            main = recep[recep["serie_norm"] == "PACIENTES RECEPCIONADOS"]
            avg = recep[recep["serie_norm"].isin(["MÉDIA DIÁRIA", "MEDIA DIÁRIA", "MEDIA DIARIA"])]

            if not main.empty:
                fig.add_trace(
                    go.Bar(
                        x=main["mes_label"],
                        y=main["valor_num"],
                        name="Pacientes recepcionados",
                        marker_color=APP_COLORS["primary_soft"],
                        hovertemplate="<b>Pacientes recepcionados</b><br>Mês: %{x}<br>Total: %{y:,.0f}<extra></extra>"
                )
            )

            if not avg.empty:
                fig.add_trace(
                    go.Scatter(
                        x=avg["mes_label"],
                        y=avg["valor_num"],
                        mode="lines+markers",
                        name="Média diária",
                        line=dict(color=APP_COLORS["primary"], width=3),
                        marker=dict(color=APP_COLORS["primary"], size=7),
                        hovertemplate="<b>Média diária</b><br>Mês: %{x}<br>Valor: %{y:,.1f}<extra></extra>"
                    )
                )

            fig = clean_trace_names(fig)

            fig = apply_plotly_theme(
                fig,
                title="Pacientes recepcionados e média diária",
                subtitle="",
                yaxis_title="Quantidade",
                height=380,
                legend=True,
                legend_orientation="h"
            )

            fig = apply_month_axis_order(fig, recep)

            plot(fig, f"{unidade}_recep_media")
        with col2:
            line_with_optional_meta(
                atend_med,
                "Atendimentos médicos vs meta",
                main_series="ATENDIMENTOS MÉDICOS",
                prefix=f"{unidade}_atend_med",
                unidade=unidade,
                subtitle=""
            )
        section_end()

    with tab_risco:
        section_start("Risco e tempo assistencial", "")
        risco_plot = risco[
            ~risco["serie_norm"].str.contains("TOTAL", na=False)
        ].copy()

        # remove meses totalmente zerados
        if not risco_plot.empty:
            risco_plot_original = risco_plot.copy()
            soma_mes_risco = risco_plot.groupby("mes_label")["valor_num"].sum(min_count=1)
            meses_validos_risco = soma_mes_risco[soma_mes_risco.fillna(0) > 0].index.tolist()
            risco_plot = risco_plot[risco_plot["mes_label"].isin(meses_validos_risco)].copy()
            if risco_plot.empty:
                risco_plot = risco_plot_original

        grouped_bar(
            risco_plot,
            "Atendimentos por classificação de risco",
            color_map=RISK_COLORS,
            unit_suffix="Quantidade",
            prefix=f"{unidade}_risco_qtd",
            unidade=unidade,
            subtitle=""
        )

        perc_plot = perc_risco[
            ~perc_risco["serie_norm"].str.contains("TOTAL", na=False)
        ].copy()

        # fallback para bases que só trazem linha TOTAL no painel percentual
        if perc_plot.empty:
            perc_plot = perc_risco.copy()

        # remove erros e meses vazios
        perc_plot = perc_plot[perc_plot["valor_num"].notna()].copy()

        if not perc_plot.empty:
            perc_plot_original = perc_plot.copy()
            soma_mes_perc = perc_plot.groupby("mes_label")["valor_num"].sum(min_count=1)
            meses_validos_perc = soma_mes_perc[soma_mes_perc.fillna(0) > 0].index.tolist()
            perc_plot = perc_plot[perc_plot["mes_label"].isin(meses_validos_perc)].copy()
            if perc_plot.empty:
                perc_plot = perc_plot_original

            # Excel percentual vem como fração (ex.: 0.65) -> converter para 65
            perc_plot["valor_num"] = perc_plot["valor_num"] * 100

        grouped_bar(
            perc_plot,
            "Percentual de atendimentos por classificação de risco",
            color_map=RISK_COLORS,
            unit_suffix="Percentual (%)",
            prefix=f"{unidade}_risco_perc",
            unidade=unidade,
            subtitle=""
        )

        line_time_chart(
            espera,
            "Tempo de espera para classificação de risco vs meta",
            main_series="MÉDIA GERAL",
            prefix=f"{unidade}_espera_class",
            unidade=unidade,
            subtitle=""
        )

        line_time_chart(
            tempo_med,
            "Tempo médio de espera de atendimento médico por classificação de risco",
            prefix=f"{unidade}_tempo_med_risco",
            unidade=unidade,
            subtitle=""
        )
        section_end()

    with tab_perm:
        section_start("Permanência, apoio e desfechos", "")
        col1, col2 = st.columns(2)
        with col1:
            line_time_chart(
                intern,
                "Tempo de permanência de pacientes internados",
                prefix=f"{unidade}_intern",
                unidade=unidade,
                subtitle=""
            )
        with col2:
            line_time_chart(
                semint,
                "Tempo de permanência de pacientes sem internação",
                prefix=f"{unidade}_semintern",
                unidade=unidade,
                subtitle=""
            )

        col1, col2 = st.columns(2)
        with col1:
            grouped_bar(
                transf,
                "Transferências (remoções)",
                prefix=f"{unidade}_transf",
                unidade=unidade,
                subtitle=""
            )
        with col2:
            grouped_bar(
                obitos,
                "Óbitos",
                prefix=f"{unidade}_obitos",
                unidade=unidade,
                subtitle=""
            )

        col1, col2 = st.columns(2)
        with col1:
            grouped_bar(
                origem[~origem["serie_norm"].eq("TOTAL")],
                "Distribuição mais recente de pacientes por origem",
                prefix=f"{unidade}_origem_bar",
                unidade=unidade,
                subtitle=""
            )
        with col2:
            pie_latest(
                origem[~origem["serie_norm"].eq("TOTAL")],
                "Distribuição mais recente de pacientes por origem",
                prefix=f"{unidade}_origem_pie",
                unidade=unidade
            )
        section_end()

    with tab_exames:
        section_start("Exames internos", "")
        grouped_bar(
            exames[~exames["serie_norm"].eq("TOTAL")],
            "Exames internos",
            prefix=f"{unidade}_exames",
            unidade=unidade,
            subtitle=""
        )

        grouped_bar(
            faixa[~faixa["serie_norm"].eq("TOTAL")],
            "Atendimentos divididos por faixa etária",
            prefix=f"{unidade}_faixa",
            unidade=unidade,
            subtitle=""
        )
        section_end()

def render_hmji(df, meses_filtrados=None):
    unidade = "HMJI"
    st.markdown(
        """
        <style>
        .hmji-title-card {
            background: #FFFFFF;
            border-radius: 12px;
            border: 1px solid #D6E4F0;
            border-left: 6px solid #0F766E;
            padding: 0.75rem 1rem;
            margin: 0.1rem 0 0.9rem 0;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.06);
        }

        .hmji-title-main {
            color: #0F2A43;
            font-size: 1.22rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            margin: 0;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="hmji-title-card">
            <div class="hmji-title-main">🏥 HMJI</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    unit_df = df[df["unidade"] == unidade].copy()
    if meses_filtrados and "mes_label" in unit_df.columns:
        unit_df = unit_df[unit_df["mes_label"].isin(meses_filtrados)].copy()

    clin = df[
        (df["unidade"] == unidade) &
        (
            df["painel_norm"].str.contains("PACIENTES CLINICOS", na=False) |
            df["serie_norm"].str.contains("PACIENTES CLINICOS", na=False)
        )
    ].copy()
    if meses_filtrados and "mes_label" in clin.columns:
        clin = clin[clin["mes_label"].isin(meses_filtrados)].copy()

    meses_base = [m for m in unit_df["mes"].dropna().tolist() if pd.notna(m)]
    meses_base = list(dict.fromkeys(meses_base))

    if not meses_base:
        meses_base = [m for m in df["mes"].dropna().tolist() if pd.notna(m)]
        meses_base = list(dict.fromkeys(meses_base))

    def hmji_block(series_map, include_total=True):
        work = unit_df.copy()
        if work.empty:
            return pd.DataFrame()

        serie_upper = work["serie"].astype(str).str.strip().str.upper()

        aliases = {}
        for label, alias_list in series_map.items():
            aliases[label] = [str(x).strip().upper() for x in alias_list]

        selected_aliases = [item for values in aliases.values() for item in values]
        matched = work[serie_upper.isin(selected_aliases)].copy()

        if not matched.empty:
            matched["serie_canonica"] = matched["serie"].astype(str).str.strip().str.upper()
            for canonical, alias_list in aliases.items():
                matched.loc[
                    matched["serie_canonica"].isin(alias_list),
                    "serie_canonica"
                ] = canonical
        else:
            matched["serie_canonica"] = pd.Series(dtype=str)

        if "TOTAL" in aliases:
            if include_total:
                matched_total = matched[matched["serie_canonica"] == "TOTAL"].copy()
            else:
                matched_total = pd.DataFrame(columns=matched.columns)
            matched = matched[matched["serie_canonica"] != "TOTAL"].copy()
            if include_total and not matched_total.empty:
                matched = pd.concat([matched, matched_total], ignore_index=True)

        if meses_base:
            grid = pd.MultiIndex.from_product(
                [meses_base, list(series_map.keys())],
                names=["mes", "serie_canonica"]
            ).to_frame(index=False)
            grid["mes_label"] = grid["mes"].map(MESES_LABEL)
            base = matched.groupby(["mes", "mes_label", "serie_canonica"], as_index=False)["valor_num"].sum()
            merged = grid.merge(base, on=["mes", "mes_label", "serie_canonica"], how="left")
        else:
            merged = matched.groupby(["mes", "mes_label", "serie_canonica"], as_index=False)["valor_num"].sum()

        merged["valor_num"] = merged["valor_num"].fillna(0.0)
        merged["unidade"] = unidade
        merged["serie"] = merged["serie_canonica"]
        merged["serie_norm"] = merged["serie_canonica"]
        return merged.sort_values(["mes", "serie"])

    obitos = filter_panel(df, unidade, "ÓBITOS")
    if meses_filtrados and "mes_label" in obitos.columns:
        obitos = obitos[obitos["mes_label"].isin(meses_filtrados)].copy()
    obitos = obitos[obitos["serie_norm"].isin(["TOTAL", "ÓBITOS", "OBITOS"])].copy()
    esp = hmji_block({
        "CIRURGIA GERAL": ["CIRURGIA GERAL"],
        "UROLOGIA": ["UROLOGIA"],
        "GINECOLOGIA": ["GINECOLOGIA"],
    }, include_total=False)
    exames = hmji_block({
        "RAIO-X": ["RAIO-X"],
        "MAMOGRAFIAS": ["MAMOGRAFIAS"],
        "ULTRASOM": ["ULTRASOM"],
        "ELETROCARDIOGRAMA": ["ELETROCARDIOGRAMA"],
        "TOTAL": ["TOTAL"],
    }, include_total=True)
    cir = hmji_block({
        "CIRURGIAS GRANDES": ["CIRURGIAS GRANDES"],
        "BIÓPSIAS": ["BIÓPSIAS"],
        "VASECTOMIAS": ["VASECTOMIAS"],
        "PEQUENAS CIRURGIAS": ["PEQUENAS CIRURGIAS"],
    }, include_total=False)
    anes = hmji_block({
        "RAQUIANESTESIA": ["RAQUIANESTESIA"],
        "ANESTESIA GERAL": ["ANESTESIA GERAL"],
        "BLOQUEIO": ["BLOQUEIO", "BLOQUEIO "],
        "ANESTESIA LOCAL": ["ANESTESIA LOCAL"],
    }, include_total=False)

    c1, c2, c3 = st.columns(3)

    with c1:
        total_clin = clin[
            ~clin["serie_norm"].isin([
                "MÉDIA DIÁRIA",
                "MEDIA DIÁRIA",
                "MEDIA DIARIA",
                "TOTAL"
            ])
        ]["valor_num"].sum()

        top_kpi_card(
            "Pacientes clínicos",
            format_int(total_clin),
            icon="🏥",
            subtitle="▲ atendimentos no período",
            accent_color="#22C55E",
            subtitle_color="#16A34A"
        )

    with c2:
        total_obitos = obitos["valor_num"].sum()

        top_kpi_card(
            "Óbitos",
            format_int(total_obitos),
            icon="⚠️",
            subtitle="▼ apenas total de óbitos",
            accent_color="#EF4444",
            subtitle_color="#DC2626"
        )

    with c3:
        top_kpi_card(
            "Procedimentos cirúrgicos",
            format_int(cir["valor_num"].sum()),
            icon="🩹",
            subtitle="▲ produção cirúrgica consolidada",
            accent_color="#3B82F6",
            subtitle_color="#2563EB"
        )

    st.markdown(
        """
        <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.6rem;
            padding: 0.2rem 0.1rem 0.5rem 0.1rem;
        }

        .stTabs [data-baseweb="tab"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.12) 0%, rgba(15,108,189,0.05) 100%);
            border: 1px solid rgba(15,108,189,0.35);
            border-radius: 12px;
            padding: 0.5rem 0.95rem;
            min-height: 44px;
            box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
        }

        .stTabs [data-baseweb="tab"] p {
            font-size: 1rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            color: #0b3b69;
        }

        .stTabs [aria-selected="true"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.24) 0%, rgba(15,108,189,0.12) 100%);
            border: 2px solid #0F6CBD;
            box-shadow: 0 0 0 1px rgba(15,108,189,0.15), 0 8px 18px rgba(15, 108, 189, 0.18);
            transform: translateY(-1px);
        }

        .stTabs [aria-selected="true"] p {
            color: #083055;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    (
        tab_pacientes,
        tab_obitos,
        tab_esp,
        tab_exames,
        tab_cir,
        tab_anes,
    ) = st.tabs(
        [
            "Pacientes clínicos",
            "Óbitos",
            "Consultas especializadas",
            "Exames internos",
            "Procedimentos cirúrgicos",
            "Anestesias",
        ]
    )

    with tab_pacientes:
        fig = go.Figure()

        main = clin[clin["serie_norm"].isin([
             "PACIENTES CLINICOS ATENDIDOS"
        ])]

        avg = clin[clin["serie_norm"].isin([
            "MÉDIA DIÁRIA",
            "MEDIA DIÁRIA",
            "MEDIA DIARIA"
        ])]

        if not main.empty:
            fig.add_trace(
                go.Bar(
                    x=main["mes_label"],
                    y=main["valor_num"],
                    name="Pacientes clínicos",
                    marker_color=SEMANTIC_COLORS["realizado_soft"],
                    hovertemplate="<b>Pacientes clínicos</b><br>Mês: %{x}<br>Total: %{y:,.0f}<extra></extra>"
                )
            )

        if not avg.empty:
            fig.add_trace(
                go.Scatter(
                    x=avg["mes_label"],
                    y=avg["valor_num"],
                    mode="lines+markers",
                    name="Média diária",
                    line=dict(color=SEMANTIC_COLORS["realizado"], width=3),
                    marker=dict(color=SEMANTIC_COLORS["realizado"], size=7),
                    hovertemplate="<b>Média diária</b><br>Mês: %{x}<br>Valor: %{y:,.1f}<extra></extra>"
                )
            )

        fig = apply_plotly_theme(
            fig,
            title="Pacientes clínicos atendidos / média diária",
            subtitle="",
            yaxis_title="Quantidade",
            height=360,
            legend=True,
            legend_orientation="h"
        )
        fig = apply_month_axis_order(fig, clin)
        plot(fig, f"{unidade}_pacientes", show_status_chip=False)

    with tab_obitos:
        grouped_bar(
            obitos,
            "Óbitos",
            prefix=f"{unidade}_obitos",
            unidade=unidade,
            subtitle="",
            show_status_chip=False,
        )

    with tab_esp:
        grouped_bar(esp, "Consultas especializadas", prefix=f"{unidade}_esp", unidade=unidade, subtitle="", show_status_chip=False)

    with tab_exames:
        grouped_bar(exames, "Exames internos", prefix=f"{unidade}_exames", unidade=unidade, subtitle="", show_status_chip=False)

    with tab_cir:
        grouped_bar(cir, "Procedimentos cirúrgicos", prefix=f"{unidade}_cir", unidade=unidade, subtitle="", show_status_chip=False)

    with tab_anes:
        grouped_bar(anes, "Anestesias", prefix=f"{unidade}_anes", unidade=unidade, subtitle="", show_status_chip=False)

def render_generic(df, unidade, paineis):
    st.subheader(unidade)
    for i,painel in enumerate(paineis, start=1):
        grouped_bar(filter_panel(df, unidade, painel), painel.title(), prefix=f"{unidade}_{i}")


def render_atencao_primaria_tabs(df):
    unidade = "ATENÇÃO PRIMÁRIA"
    paineis = [
        "CONSULTAS MÉDICAS",
        "NÍVEL SUPERIOR (EXCETO MÉDICO)",
    ]
    titulos = [
        "Consultas Médicas",
        "Nível Superior (Exceto Médico)",
    ]

    st.subheader(unidade)
    st.markdown(
        """
        <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.6rem;
            padding: 0.2rem 0.1rem 0.5rem 0.1rem;
        }

        .stTabs [data-baseweb="tab"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.12) 0%, rgba(15,108,189,0.05) 100%);
            border: 1px solid rgba(15,108,189,0.35);
            border-radius: 12px;
            padding: 0.5rem 0.95rem;
            min-height: 44px;
            box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
        }

        .stTabs [data-baseweb="tab"] p {
            font-size: 1rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            color: #0b3b69;
        }

        .stTabs [aria-selected="true"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.24) 0%, rgba(15,108,189,0.12) 100%);
            border: 2px solid #0F6CBD;
            box-shadow: 0 0 0 1px rgba(15,108,189,0.15), 0 8px 18px rgba(15, 108, 189, 0.18);
            transform: translateY(-1px);
        }

        .stTabs [aria-selected="true"] p {
            color: #083055;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    tabs = st.tabs(titulos)
    for i, (tab, painel, titulo) in enumerate(zip(tabs, paineis, titulos), start=1):
        with tab:
            grouped_bar(
                filter_panel(df, unidade, painel),
                titulo,
                prefix=f"{unidade}_tab_{i}",
            )


def render_atencao_secundaria_tabs(df):
    unidade = "ATENÇÃO SECUNDÁRIA"
    paineis = [
        "CONSULTAS ESPECIALIZADAS (CAIS)",
        "CONSULTAS ESPECIALIZADAS (MATERNO INFANTIL)",
        "CONSULTAS ESPECIALIZADAS (FARMÁCIA CENTRAL)",
    ]
    titulos = [
        "Consultas Especializadas (CAIS)",
        "Consultas Especializadas (Materno Infantil)",
        "Consultas Especializadas (Farmácia Central)",
    ]

    st.subheader(unidade)
    st.markdown(
        """
        <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.6rem;
            padding: 0.2rem 0.1rem 0.5rem 0.1rem;
        }

        .stTabs [data-baseweb="tab"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.12) 0%, rgba(15,108,189,0.05) 100%);
            border: 1px solid rgba(15,108,189,0.35);
            border-radius: 12px;
            padding: 0.5rem 0.95rem;
            min-height: 44px;
            box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
        }

        .stTabs [data-baseweb="tab"] p {
            font-size: 1rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            color: #0b3b69;
        }

        .stTabs [aria-selected="true"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.24) 0%, rgba(15,108,189,0.12) 100%);
            border: 2px solid #0F6CBD;
            box-shadow: 0 0 0 1px rgba(15,108,189,0.15), 0 8px 18px rgba(15, 108, 189, 0.18);
            transform: translateY(-1px);
        }

        .stTabs [aria-selected="true"] p {
            color: #083055;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    tabs = st.tabs(titulos)
    for i, (tab, painel, titulo) in enumerate(zip(tabs, paineis, titulos), start=1):
        with tab:
            grouped_bar(
                filter_panel(df, unidade, painel),
                titulo,
                prefix=f"{unidade}_tab_{i}",
            )


def render_saude_mental_tabs(df):
    unidade = "SAÚDE MENTAL"
    paineis = [
        "CONSULTAS ESPECIALIZADAS (CAPS II)",
        "CONSULTAS ESPECIALIZADAS (CAPS AD III)",
        "CONSULTAS ESPECIALIZADAS (CLÍNICA PSICOLOGIA)",
    ]
    titulos = [
        "Consultas Especializadas (CAPS II)",
        "Consultas Especializadas (CAPS AD III)",
        "Consultas Especializadas (Clínica Psicologia)",
    ]

    st.subheader(unidade)
    st.markdown(
        """
        <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.6rem;
            padding: 0.2rem 0.1rem 0.5rem 0.1rem;
        }

        .stTabs [data-baseweb="tab"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.12) 0%, rgba(15,108,189,0.05) 100%);
            border: 1px solid rgba(15,108,189,0.35);
            border-radius: 12px;
            padding: 0.5rem 0.95rem;
            min-height: 44px;
            box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
        }

        .stTabs [data-baseweb="tab"] p {
            font-size: 1rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            color: #0b3b69;
        }

        .stTabs [aria-selected="true"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.24) 0%, rgba(15,108,189,0.12) 100%);
            border: 2px solid #0F6CBD;
            box-shadow: 0 0 0 1px rgba(15,108,189,0.15), 0 8px 18px rgba(15, 108, 189, 0.18);
            transform: translateY(-1px);
        }

        .stTabs [aria-selected="true"] p {
            color: #083055;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    tabs = st.tabs(titulos)
    for i, (tab, painel, titulo) in enumerate(zip(tabs, paineis, titulos), start=1):
        with tab:
            grouped_bar(
                filter_panel(df, unidade, painel),
                titulo,
                prefix=f"{unidade}_tab_{i}",
            )

def rh_get_latest_month(panel_df):
    if panel_df is None or panel_df.empty:
        return None

    work = panel_df.copy()

    # considera apenas linhas com mês e valor numérico preenchido
    work = work.dropna(subset=["mes"]).copy()
    work = work[work["valor_num"].notna()].copy()

    if work.empty:
        return None

    return work["mes"].max()


def rh_get_value_and_meta(panel_df):
    if panel_df is None or panel_df.empty:
        return {
            "mes": None,
            "valor": None
        }

    latest_mes = rh_get_latest_month(panel_df)
    if latest_mes is None:
        return {
            "mes": None,
            "valor": None
        }

    recorte = panel_df[panel_df["mes"] == latest_mes].copy()
    if recorte.empty:
        return {
            "mes": latest_mes,
            "valor": None
        }

    valor_df = recorte[
        ~recorte["serie_norm"].isin(["META", "MÉDIA DIÁRIA", "MEDIA DIÁRIA", "MEDIA DIARIA", "TOTAL"])
    ].copy()

    # fallback: se o indicador vier com a própria série igual ao painel
    if valor_df.empty:
        valor_df = recorte.copy()

    valor_df = valor_df[valor_df["valor_num"].notna()].copy()

    valor = valor_df["valor_num"].sum() if not valor_df.empty else None

    if valor is not None and pd.isna(valor):
        valor = None

    return {
        "mes": latest_mes,
        "valor": float(valor) if valor is not None else None
    }


def rh_is_lower_better(nome_indicador):
    nome_norm = normalize_text(nome_indicador) or ""
    indicadores_menor_melhor = {
        "TAXA DE TURNOVER",
        "ABSENTEISMO",
        "ACIDENTES DE TRABALHO",
    }
    return nome_norm in indicadores_menor_melhor


def rh_compute_status(nome_indicador, valor, meta):
    """
    Regras:
    - sem meta -> neutro
    - maior é melhor:
        >=100% da meta = verde
        entre 85% e 99,9% = amarelo
        abaixo de 85% = vermelho
    - menor é melhor:
        <=100% da meta = verde
        até 115% da meta = amarelo
        acima de 115% = vermelho
    """
    if valor is None or meta is None or pd.isna(valor) or pd.isna(meta) or meta == 0:
        return {
            "status": "Sem meta",
            "cor": "#64748B",
            "pct": None,
            "comparacao": "Sem comparativo"
        }

    menor_melhor = rh_is_lower_better(nome_indicador)
    pct = (valor / meta) * 100

    if menor_melhor:
        if valor <= meta:
            status = "Meta atingida"
            cor = "#16A34A"
        elif valor <= meta * 1.15:
            status = "Atenção"
            cor = "#F59E0B"
        else:
            status = "Abaixo da meta"
            cor = "#DC2626"
    else:
        if valor >= meta:
            status = "Meta atingida"
            cor = "#16A34A"
        elif valor >= meta * 0.85:
            status = "Atenção"
            cor = "#F59E0B"
        else:
            status = "Abaixo da meta"
            cor = "#DC2626"

    diferenca = valor - meta
    if diferenca > 0:
        comparacao = f"+{rh_format_value(nome_indicador, abs(diferenca))} vs meta"
    elif diferenca < 0:
        comparacao = f"-{rh_format_value(nome_indicador, abs(diferenca))} vs meta"
    else:
        comparacao = "Em linha com a meta"

    return {
        "status": status,
        "cor": cor,
        "pct": pct,
        "comparacao": comparacao
    }

def rh_format_value(nome_indicador, valor):
    if valor is None or pd.isna(valor):
        return "-"

    nome_norm = normalize_text(nome_indicador) or ""

    indicadores_percentuais = {
        "TAXA DE TURNOVER",
        "ABSENTEISMO",
    }

    if nome_norm in indicadores_percentuais:
        return f"{valor * 100:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")

    if float(valor).is_integer():
        return f"{int(valor):,}".replace(",", ".")

    return f"{valor:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")
RH_ICONS = {
    "TOTAL DE COLABORADORES CLT": "👥",
    "TOTAL DE MÉDICOS": "🩺",
    "TOTAL DE ENFERMAGEM": "💉",
    "ADMISSÕES": "📥",
    "DESLIGAMENTOS": "📤",
    "TAXA DE TURNOVER": "🔄",
    "ABSENTEÍSMO": "⏱️",
    "AFASTAMENTOS": "🏥",
    "ACIDENTES DE TRABALHO": "⚠️",
}

def render_rh_indicator_card(nome_indicador, panel_df):
    info = rh_get_value_and_meta(panel_df)

    valor = info["valor"]
    mes = info["mes"]

    icone = RH_ICONS.get(nome_indicador, "📊")
    valor_fmt = rh_format_value(nome_indicador, valor)
    mes_fmt = MESES_LABEL.get(mes, "-") if mes is not None else "-"

    top_kpi_card(
        title=nome_indicador,
        value=valor_fmt,
        icon=icone,
        subtitle=f"Ref: {mes_fmt}",
        accent_color="#0F6CBD",
        subtitle_color="#64748B",
    )

def render_rh_page(df, meses_filtrados, file_bytes=None, _mtime=None):
    unidade = "RH"
    st.subheader("Gestão de Pessoas")

    work_df = df.copy()

    # respeita o filtro lateral de período já existente no app
    if "mes_label" in work_df.columns and meses_filtrados:
        work_df = work_df[work_df["mes_label"].isin(meses_filtrados)].copy()

    indicadores_rh = [
        "TOTAL DE COLABORADORES CLT",
        "TOTAL DE MÉDICOS",
        "TOTAL DE ENFERMAGEM",
        "ADMISSÕES",
        "DESLIGAMENTOS",
        "TAXA DE TURNOVER",
        "ABSENTEÍSMO",
        "AFASTAMENTOS",
        "ACIDENTES DE TRABALHO",
]

    paineis = {
        indicador: filter_panel(work_df, unidade, indicador)
        for indicador in indicadores_rh
    }

    st.markdown(
        """
        <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.6rem;
            padding: 0.2rem 0.1rem 0.5rem 0.1rem;
        }

        .stTabs [data-baseweb="tab"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.12) 0%, rgba(15,108,189,0.05) 100%);
            border: 1px solid rgba(15,108,189,0.35);
            border-radius: 12px;
            padding: 0.5rem 0.95rem;
            min-height: 44px;
            box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
        }

        .stTabs [data-baseweb="tab"] p {
            font-size: 1rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            color: #0b3b69;
        }

        .stTabs [aria-selected="true"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.24) 0%, rgba(15,108,189,0.12) 100%);
            border: 2px solid #0F6CBD;
            box-shadow: 0 0 0 1px rgba(15,108,189,0.15), 0 8px 18px rgba(15, 108, 189, 0.18);
            transform: translateY(-1px);
        }

        .stTabs [aria-selected="true"] p {
            color: #083055;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    tab_indicadores, tab_colaboradores = st.tabs(["📊 Indicadores RH", "👥 Colaboradores"])

    with tab_indicadores:
        section_start(
            "<strong>Painel de indicadores de RH</strong>",
            ""
        )

        cols = st.columns(3)
        for idx, indicador in enumerate(indicadores_rh):
            with cols[idx % 3]:
                render_rh_indicator_card(indicador, paineis[indicador])

        section_end()

    with tab_colaboradores:
        section_start(
            "Base completa de colaboradores",
            ""
        )

        colab_full = load_colaboradores_sheet(file_bytes=file_bytes, _mtime=_mtime)
        if colab_full is None or colab_full.empty:
            st.info("A aba COLABORADORES não foi encontrada ou está sem dados na base atual.")
            section_end()
            return

        def _pick_col(df_cols, aliases):
            cols_norm = {normalize_text(c): c for c in df_cols}
            for alias in aliases:
                if alias in cols_norm:
                    return cols_norm[alias]
            for col in df_cols:
                col_norm = normalize_text(col) or ""
                if any(alias in col_norm for alias in aliases):
                    return col
            return None

        col_nome = _pick_col(colab_full.columns, {"COLABORADOR", "NOME"})
        col_situacao = _pick_col(colab_full.columns, {"SITUACAO", "SITUAÇÃO", "STATUS"})
        col_cargo = _pick_col(colab_full.columns, {"CARGO"})
        col_un1 = _pick_col(colab_full.columns, {"UNIDADE"})
        col_un2 = _pick_col(colab_full.columns, {"UNIDADE 2", "UNIDADE2"})
        col_un3 = _pick_col(colab_full.columns, {"UNIDADE 3", "UNIDADE3"})

        filtro_cols = st.columns([1.6, 1.2, 1.2, 1.2])
        busca_nome = filtro_cols[0].text_input("Buscar colaborador", value="", key="rh_colab_busca")

        situacao_options = []
        if col_situacao:
            situacao_options = sorted(
                colab_full[col_situacao].dropna().astype(str).str.strip().replace("", pd.NA).dropna().unique().tolist()
            )
        filtro_situacao = filtro_cols[1].multiselect(
            "Situação",
            options=situacao_options,
            default=[],
            key="rh_colab_situacao",
        )

        cargo_options = []
        if col_cargo:
            cargo_options = sorted(
                colab_full[col_cargo].dropna().astype(str).str.strip().replace("", pd.NA).dropna().unique().tolist()
            )
        filtro_cargo = filtro_cols[2].multiselect(
            "Cargo",
            options=cargo_options,
            default=[],
            key="rh_colab_cargo",
        )

        unidade_options = set()
        for uc in [col_un1, col_un2, col_un3]:
            if not uc:
                continue
            unidade_options.update(
                colab_full[uc].dropna().astype(str).str.strip().replace("", pd.NA).dropna().unique().tolist()
            )
        filtro_unidade = filtro_cols[3].multiselect(
            "Unidade",
            options=sorted(unidade_options),
            default=[],
            key="rh_colab_unidade",
        )

        colab_view = colab_full.copy()
        if busca_nome and col_nome:
            colab_view = colab_view[
                colab_view[col_nome].astype(str).str.contains(busca_nome, case=False, na=False)
            ]
        if filtro_situacao and col_situacao:
            colab_view = colab_view[colab_view[col_situacao].astype(str).isin(filtro_situacao)]
        if filtro_cargo and col_cargo:
            colab_view = colab_view[colab_view[col_cargo].astype(str).isin(filtro_cargo)]
        if filtro_unidade:
            unidade_mask = pd.Series(False, index=colab_view.index)
            for uc in [col_un1, col_un2, col_un3]:
                if uc:
                    unidade_mask = unidade_mask | colab_view[uc].astype(str).isin(filtro_unidade)
            colab_view = colab_view[unidade_mask]

        st.caption(f"Registros exibidos: {len(colab_view):,}".replace(",", "."))
        st.dataframe(colab_view, width="stretch", hide_index=True)

        section_end()


def render_produtividade_medica_page():
    prod = load_produtividade_data(_mtime=_samu_file_mtime())
    kd = prod["kpi_diario"].copy()
    ku = prod["kpi_diario_unidade"].copy()
    ks = prod["kpi_semanal"].copy()
    rk = prod["ranking"].copy()
    t0 = prod["top5_geral"].copy()
    t2 = prod["top5_upa2"].copy()
    t1 = prod["top5_upa1"].copy()

    if kd.empty and rk.empty:
        st.warning("Arquivo urgencia_tratado_validado.xlsx não encontrado na pasta do app.")
        return

    st.markdown("## 📈 Produtividade Médica UPAs")
    st.markdown("#### Filtros")

    data_min = kd["Data"].dropna().min().date() if "Data" in kd.columns and not kd["Data"].dropna().empty else None
    data_max = kd["Data"].dropna().max().date() if "Data" in kd.columns and not kd["Data"].dropna().empty else None

    cf1, cf2 = st.columns([1, 2])
    with cf1:
        unid = st.selectbox(
            "Unidade",
            ["Todas", "UPA II DE LUZIÂNIA", "UPA I JARDIM INGÁ"],
            key="pm_unid",
        )
    with cf2:
        if data_min and data_max:
            periodo = st.date_input(
                "Período",
                value=(data_min, data_max),
                min_value=data_min,
                max_value=data_max,
                key="pm_periodo",
            )
        else:
            periodo = None
    st.divider()

    ini = fim = None
    if isinstance(periodo, (list, tuple)) and len(periodo) == 2:
        ini, fim = pd.to_datetime(periodo[0]), pd.to_datetime(periodo[1])
    elif isinstance(periodo, dt.date):
        ini = fim = pd.to_datetime(periodo)

    if ini is not None:
        if "Data" in kd.columns:
            kd = kd[(kd["Data"] >= ini) & (kd["Data"] <= fim)].copy()
        if "Data" in ku.columns:
            ku = ku[(ku["Data"] >= ini) & (ku["Data"] <= fim)].copy()
        if "Semana_Inicio" in ks.columns and "Semana_Fim" in ks.columns:
            ks = ks[(ks["Semana_Fim"] >= ini) & (ks["Semana_Inicio"] <= fim)].copy()

    unidade_col_map = {
        "UPA II DE LUZIÂNIA": "UPA II DE LUZIÂNIA",
        "UPA I JARDIM INGÁ": "UPA I JARDIM INGÁ",
    }
    serie_coluna = "Total_Geral_24h"
    if unid != "Todas":
        serie_coluna = unidade_col_map.get(unid, "Total_Geral_24h")

    if unid != "Todas":
        if "Unidade" in ku.columns:
            ku = ku[ku["Unidade"] == unid].copy()
        if "Unidade" in rk.columns:
            rk = rk[rk["Unidade"] == unid].copy()
        top5_ref = t2 if unid == "UPA II DE LUZIÂNIA" else (t1 if unid == "UPA I JARDIM INGÁ" else t0)
    else:
        top5_ref = t0

    # KPIs seguem exatamente a unidade selecionada no filtro.
    if "Data" in kd.columns and serie_coluna in kd.columns:
        kpi_df = kd[["Data", serie_coluna]].copy().rename(columns={serie_coluna: "valor"}).dropna(subset=["Data", "valor"])
    else:
        kpi_df = pd.DataFrame(columns=["Data", "valor"])

    serie = pd.to_numeric(kpi_df.get("valor", pd.Series(dtype=float)), errors="coerce").dropna()
    total = float(serie.sum()) if not serie.empty else 0.0
    media = float(serie.mean()) if not serie.empty else 0.0
    melhor = float(serie.max()) if not serie.empty else 0.0
    pior = float(serie.min()) if not serie.empty else 0.0
    melhor_dia = pior_dia = "-"
    if not kpi_df.empty:
        melhor_dia = kpi_df.loc[kpi_df["valor"].idxmax(), "Data"].strftime("%d/%m/%Y")
        pior_dia = kpi_df.loc[kpi_df["valor"].idxmin(), "Data"].strftime("%d/%m/%Y")

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        top_kpi_card("Total do período", f"{int(total):,}".replace(",", "."), icon="📈", subtitle="Soma diária", accent_color=SEMANTIC_COLORS["success"], subtitle_color=SEMANTIC_COLORS["success"])
    with k2:
        top_kpi_card("Média diária", f"{media:,.1f}".replace(",", "."), icon="📆", subtitle="Média do período", accent_color=SEMANTIC_COLORS["primary"], subtitle_color=SEMANTIC_COLORS["primary"])
    with k3:
        top_kpi_card("Melhor dia", f"{int(melhor):,}".replace(",", "."), icon="🏆", subtitle=f"Data: {melhor_dia}", accent_color=SEMANTIC_COLORS["warning"], subtitle_color=SEMANTIC_COLORS["warning"])
    with k4:
        top_kpi_card("Pior dia", f"{int(pior):,}".replace(",", "."), icon="📉", subtitle=f"Data: {pior_dia}", accent_color=SEMANTIC_COLORS["danger"], subtitle_color=SEMANTIC_COLORS["danger"])

    st.markdown(
        """
        <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.7rem;
            padding: 0.15rem 0 0.65rem 0;
            overflow-x: auto;
        }

        .stTabs [data-baseweb="tab"] {
            background: linear-gradient(180deg, #F5FAFF 0%, #EDF5FC 100%);
            border: 1px solid #BDD5EA;
            border-radius: 14px;
            padding: 0.48rem 1.05rem;
            min-height: 48px;
            box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
            transition: all 0.18s ease;
        }

        .stTabs [data-baseweb="tab"]:hover {
            background: linear-gradient(180deg, #F0F7FF 0%, #E6F0FA 100%);
            border-color: #8CB8DF;
        }

        .stTabs [data-baseweb="tab"] p {
            font-size: 0.98rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            color: #0B3B69;
        }

        .stTabs [aria-selected="true"] {
            background: linear-gradient(180deg, #FFFFFF 0%, #F4F9FE 100%);
            border: 2px solid #0F6CBD;
            box-shadow: inset 0 -3px 0 #EF4444, 0 8px 18px rgba(15, 108, 189, 0.14);
            transform: translateY(-1px);
        }

        .stTabs [aria-selected="true"] p {
            color: #083055;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    tab_evolucao, tab_unidade, tab_semanal, tab_top5, tab_ranking = st.tabs([
        "📈 Evolução diária",
        "🏥 Produção por unidade",
        "🗓️ Produção semanal",
        "🥇 Top 5 médicos",
        "📋 Ranking completo",
    ])

    with tab_evolucao:
        section_start("", "")
        if not kpi_df.empty:
            ln = kpi_df.sort_values("Data")
            fig = go.Figure()
            fig.add_trace(
                go.Scatter(
                    x=ln["Data"],
                    y=ln["valor"],
                    mode="lines+markers",
                    line=dict(color=SEMANTIC_COLORS["primary"], width=3),
                    marker=dict(size=6),
                    hovertemplate="<b>%{x|%d/%m/%Y}</b><br>Total: %{y:,.0f}<extra></extra>",
                )
            )
            fig = apply_plotly_theme(fig, title="Atendimentos diários", subtitle="", yaxis_title="Atendimentos", height=360, legend=False)
            plot(fig, "pm_evolucao")
        else:
            st.info("Sem dados para o período selecionado.")
        section_end()

    with tab_unidade:
        section_start("", "")
        ucols = [c for c in ["UPA II DE LUZIÂNIA", "UPA I JARDIM INGÁ"] if c in kd.columns]
        if ucols and not kd.empty and "Data" in kd.columns:
            plot_cols = ucols if unid == "Todas" else [c for c in ucols if c == unid]
            if plot_cols:
                lng = kd[["Data"] + plot_cols].melt(id_vars="Data", var_name="Unidade", value_name="Atendimentos").dropna(subset=["Atendimentos", "Data"]).sort_values("Data")
                fig2 = px.line(
                    lng,
                    x="Data",
                    y="Atendimentos",
                    color="Unidade",
                    markers=True,
                    color_discrete_sequence=[SEMANTIC_COLORS["series_1"], SEMANTIC_COLORS["series_2"], SEMANTIC_COLORS["series_3"]],
                )
                fig2.update_traces(hovertemplate="<b>%{fullData.name}</b><br>%{x|%d/%m/%Y}<br>%{y:,.0f}<extra></extra>")
                fig2 = apply_plotly_theme(fig2, title="Atendimentos por unidade", subtitle="", yaxis_title="Atendimentos", height=360, legend=True, legend_orientation="h")
                plot(fig2, "pm_unidades")
        else:
            st.info("Sem dados por unidade.")
        section_end()

    with tab_semanal:
        section_start("", "")
        semanal_col_map = {
            "Todas": "Total_Semana_Geral",
            "UPA II DE LUZIÂNIA": "Total_Semana_UPA_II",
            "UPA I JARDIM INGÁ": "Total_Semana_UPA_I",
        }
        semanal_col = semanal_col_map.get(unid, "Total_Semana_Geral")
        semanal_titulo = ""

        if not ks.empty and semanal_col in ks.columns and "Semana_Inicio" in ks.columns and "Semana_Fim" in ks.columns:
            sp = ks.sort_values("Semana_Inicio").copy()
            sp["Semana"] = sp.apply(
                lambda r: f"{r['Semana_Inicio'].strftime('%d/%m')} - {r['Semana_Fim'].strftime('%d/%m')}" if pd.notna(r.get("Semana_Inicio")) and pd.notna(r.get("Semana_Fim")) else "-",
                axis=1,
            )
            fig3 = px.bar(sp, x="Semana", y=semanal_col, color_discrete_sequence=[SEMANTIC_COLORS["primary_soft"]])
            fig3.update_traces(marker_line_width=0, hovertemplate="<b>%{x}</b><br>%{y:,.0f}<extra></extra>")
            fig3 = apply_plotly_theme(fig3, title=semanal_titulo, subtitle="", yaxis_title="Atendimentos", height=340, legend=False)
            plot(fig3, "pm_semanal")
        else:
            st.info("Sem dados semanais.")
        section_end()

    with tab_top5:
        section_start("", "")
        if not top5_ref.empty and "Médico" in top5_ref.columns and "Total_Atendimentos" in top5_ref.columns:
            fig4 = px.bar(top5_ref, y="Médico", x="Total_Atendimentos", orientation="h", color="Total_Atendimentos", color_continuous_scale=[SEMANTIC_COLORS["primary_soft"], SEMANTIC_COLORS["primary"]])
            fig4.update_traces(hovertemplate="<b>%{y}</b><br>%{x:,.0f} atendimentos<extra></extra>")
            fig4 = apply_plotly_theme(fig4, title="Top 5 por atendimentos", subtitle="", yaxis_title="", height=340, legend=False)
            fig4.update_xaxes(title_text="Atendimentos")
            plot(fig4, "pm_top5")
        else:
            st.info("Sem dados de Top 5.")
        section_end()

    with tab_ranking:
        section_start("", "")
        if not rk.empty:
            rcols = [c for c in ["Médico", "Unidade", "Total_Atendimentos", "Plantoes", "Media_por_Plantao", "Media_por_Hora"] if c in rk.columns]
            rv = rk.sort_values("Total_Atendimentos", ascending=False) if "Total_Atendimentos" in rk.columns else rk
            st.dataframe(rv[rcols].reset_index(drop=True), use_container_width=True)
        else:
            st.info("Sem dados de ranking.")
        section_end()


def render_samu_page():
    samu = load_samu_data(_mtime=_samu_file_mtime())
    diario = samu["diario"].copy()
    resumo = samu["resumo"].copy()
    titulo = samu.get("titulo", "SAMU")

    st.markdown(
        """
        <style>
        .samu-title-card {
            background: #FFFFFF;
            border-radius: 12px;
            border: 1px solid #D6E4F0;
            border-left: 6px solid #DC2626;
            padding: 0.75rem 1rem;
            margin: 0.1rem 0 0.9rem 0;
            box-shadow: 0 2px 10px rgba(15, 23, 42, 0.06);
        }

        .samu-title-main {
            color: #0F2A43;
            font-size: 1.22rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            margin: 0;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div class="samu-title-card">
            <div class="samu-title-main">🚨 SAMU</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if diario.empty and resumo.empty:
        st.warning("A aba SAMU não foi encontrada ou está vazia na planilha.")
        return

    data_min = diario["Data"].dropna().min().date() if "Data" in diario.columns and not diario["Data"].dropna().empty else None
    data_max = diario["Data"].dropna().max().date() if "Data" in diario.columns and not diario["Data"].dropna().empty else None

    st.markdown(
        """
        <div style="margin: 2px 0 4px 0; font-size: 11px; font-weight: 600; letter-spacing: 0.04em; text-transform: uppercase; color: #64748B;">
            Filtros
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <style>
        div[data-testid="stDateInput"] {
            max-width: 320px;
        }
        div[data-testid="stDateInput"] input {
            min-height: 34px !important;
            padding: 0.24rem 0.62rem !important;
            font-size: 0.90rem !important;
            border-radius: 10px !important;
            border: 1px solid #CBD5E1 !important;
            background: #F8FAFC !important;
            box-shadow: none !important;
        }
        div[data-testid="stDateInput"] input:focus {
            border-color: #94A3B8 !important;
            box-shadow: 0 0 0 1px rgba(148, 163, 184, 0.18) !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    filtro_col, _ = st.columns([0.36, 0.64])
    with filtro_col:
        if data_min and data_max:
            periodo = st.date_input(
                "Período",
                value=(data_min, data_max),
                key="samu_periodo",
                label_visibility="collapsed",
            )
        else:
            periodo = None

    st.caption(f"Fonte: {titulo}")
    st.divider()

    diario_filtrado = diario.copy()
    if periodo is not None and not diario_filtrado.empty:
        if isinstance(periodo, (list, tuple)) and len(periodo) == 2:
            ini = pd.to_datetime(periodo[0])
            fim = pd.to_datetime(periodo[1])
            diario_filtrado = diario_filtrado[
                (diario_filtrado["Data"] >= ini) & (diario_filtrado["Data"] <= fim)
            ].copy()
        elif isinstance(periodo, dt.date):
            alvo = pd.to_datetime(periodo)
            diario_filtrado = diario_filtrado[diario_filtrado["Data"] == alvo].copy()

    if diario_filtrado.empty:
        st.info("Sem dados do SAMU para o período selecionado.")
        return

    diario_total = diario_filtrado.groupby("Data", as_index=False)["Atendimentos"].sum().sort_values("Data")
    procedimentos_total = (
        diario_filtrado.groupby(["Descricao", "Codigo_SIGTAP"], as_index=False)["Atendimentos"]
        .sum()
        .sort_values("Atendimentos", ascending=False)
    )

    total_periodo = float(diario_filtrado["Atendimentos"].sum())
    media_diaria = float(diario_total["Atendimentos"].mean()) if not diario_total.empty else 0.0
    melhor_dia = float(diario_total["Atendimentos"].max()) if not diario_total.empty else 0.0
    melhor_data = diario_total.loc[diario_total["Atendimentos"].idxmax(), "Data"].strftime("%d/%m/%Y") if not diario_total.empty else "-"
    procedimentos_ativos = int(procedimentos_total["Descricao"].nunique())

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        top_kpi_card(
            "Total no período",
            f"{int(total_periodo):,}".replace(",", "."),
            icon="📈",
            subtitle="Soma dos atendimentos",
            accent_color=SEMANTIC_COLORS["success"],
            subtitle_color=SEMANTIC_COLORS["success"],
        )
    with k2:
        top_kpi_card(
            "Média diária",
            f"{media_diaria:,.1f}".replace(",", "."),
            icon="📆",
            subtitle="",
            accent_color=SEMANTIC_COLORS["primary"],
            subtitle_color=SEMANTIC_COLORS["primary"],
        )
    with k3:
        top_kpi_card(
            "Melhor dia",
            f"{int(melhor_dia):,}".replace(",", "."),
            icon="🏆",
            subtitle=f"Data: {melhor_data}",
            accent_color=SEMANTIC_COLORS["warning"],
            subtitle_color=SEMANTIC_COLORS["warning"],
        )
    with k4:
        top_kpi_card(
            "Procedimentos ativos",
            f"{procedimentos_ativos}",
            icon="🧾",
            subtitle="Com produção no período",
            accent_color=SEMANTIC_COLORS["danger"],
            subtitle_color=SEMANTIC_COLORS["danger"],
        )

    st.markdown(
        """
        <style>
        .stTabs [data-baseweb="tab-list"] {
            gap: 0.6rem;
            padding: 0.2rem 0.1rem 0.5rem 0.1rem;
        }

        .stTabs [data-baseweb="tab"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.12) 0%, rgba(15,108,189,0.05) 100%);
            border: 1px solid rgba(15,108,189,0.35);
            border-radius: 12px;
            padding: 0.5rem 0.95rem;
            min-height: 44px;
            box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
        }

        .stTabs [data-baseweb="tab"] p {
            font-size: 1rem;
            font-weight: 800;
            letter-spacing: 0.01em;
            color: #0b3b69;
        }

        .stTabs [aria-selected="true"] {
            background: linear-gradient(180deg, rgba(15,108,189,0.24) 0%, rgba(15,108,189,0.12) 100%);
            border: 2px solid #0F6CBD;
            box-shadow: 0 0 0 1px rgba(15,108,189,0.15), 0 8px 18px rgba(15, 108, 189, 0.18);
            transform: translateY(-1px);
        }

        .stTabs [aria-selected="true"] p {
            color: #083055;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    tab_meta, tab_prod, tab_top, tab_inds, tab_resumo = st.tabs([
        "Metas mensais prioritárias",
        "Produção diária do SAMU",
        "Top 10 procedimentos",
        "Indicadores",
        "Resumo por procedimento",
    ])

    with tab_meta:
        section_start("", "")
        metas_samu = [
            {
                "descricao": "ATENDIMENTO REALIZADO PELA USA TERRESTRE (COM ENVIO DA VIATURA)",
                "meta_mensal": 60.5,
                "termos_chave": [
                    "ATENDIMENTO REALIZADO",
                    "USA TERRESTRE",
                    "ENVIO DA VIATURA",
                ],
            },
            {
                "descricao": "ATENDIMENTO DAS CHAMADAS RECEBIDAS PELA CENTRAL DE REGULAÇÃO DAS URGÊNCIAS COM ORIENTAÇÃO (SEM ENVIO DE VIATURA)",
                "meta_mensal": 148.5,
                "termos_chave": [
                    "ATENDIMENTO DAS CHAMADAS RECEBIDAS",
                    "CENTRAL DE REGULACAO DAS URGENCIAS",
                    "ORIENTACAO",
                    "SEM ENVIO DE VIATURA",
                ],
            },
        ]

        col_meta_1, col_meta_2 = st.columns(2)
        for idx_meta, meta_cfg in enumerate(metas_samu):
            container = col_meta_1 if idx_meta == 0 else col_meta_2
            with container:
                desc_norm = procedimentos_total["Descricao"].fillna("").astype(str).map(normalize_text)
                mask = desc_norm.map(
                    lambda d: all(term in d for term in meta_cfg["termos_chave"])
                )
                realizado = float(procedimentos_total.loc[mask, "Atendimentos"].sum()) if mask.any() else 0.0
                meta_mensal = float(meta_cfg["meta_mensal"])
                atingimento_pct = ((realizado / meta_mensal) * 100) if meta_mensal > 0 else None
                saldo_pct = (((realizado - meta_mensal) / meta_mensal) * 100) if meta_mensal > 0 else None

                render_meta_card(
                    meta_cfg["descricao"],
                    realizado,
                    meta_mensal,
                    atingimento_pct,
                    saldo_pct,
                )
        section_end()

    with tab_prod:
        section_start("", "")
        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=diario_total["Data"],
                y=diario_total["Atendimentos"],
                mode="lines+markers",
                line=dict(color=SEMANTIC_COLORS["primary"], width=3),
                marker=dict(size=6),
                hovertemplate="<b>%{x|%d/%m/%Y}</b><br>Atendimentos: %{y:,.0f}<extra></extra>",
            )
        )
        fig = apply_plotly_theme(
            fig,
            title="Produção diária do SAMU",
            subtitle="",
            yaxis_title="Atendimentos",
            height=350,
            legend=False,
        )
        st.plotly_chart(fig, use_container_width=True, key="samu_evolucao_diaria")
        section_end()

    with tab_top:
        section_start("", "")
        top_proc = procedimentos_total.head(10).copy()
        if not top_proc.empty:
            fig2 = px.bar(
                top_proc.sort_values("Atendimentos", ascending=True),
                x="Atendimentos",
                y="Descricao",
                orientation="h",
                color_discrete_sequence=[SEMANTIC_COLORS["primary_soft"]],
            )
            fig2.update_traces(hovertemplate="<b>%{y}</b><br>%{x:,.0f}<extra></extra>")
            fig2 = apply_plotly_theme(
                fig2,
                title="Top 10 procedimentos",
                subtitle="",
                yaxis_title="",
                height=420,
                legend=False,
            )
            fig2.update_xaxes(title_text="Atendimentos")
            st.plotly_chart(fig2, use_container_width=True, key="samu_top_procedimentos")
        else:
            st.info("Sem dados de procedimentos para o período selecionado.")
        section_end()

    with tab_inds:
        section_start("", "")
        indicadores_ordenados = procedimentos_total[["Descricao", "Codigo_SIGTAP", "Atendimentos"]].copy()
        indicadores_ordenados = indicadores_ordenados.sort_values("Atendimentos", ascending=False).reset_index(drop=True)

        if indicadores_ordenados.empty:
            st.info("Sem indicadores com produção no período selecionado.")
        else:
            for idx, row in indicadores_ordenados.iterrows():
                descricao = str(row.get("Descricao", "Indicador"))

                serie_indicador = (
                    diario_filtrado[diario_filtrado["Descricao"] == descricao]
                    .groupby("Data", as_index=False)["Atendimentos"]
                    .sum()
                    .sort_values("Data")
                )

                if serie_indicador.empty:
                    continue

                fig_ind = go.Figure()
                fig_ind.add_trace(
                    go.Scatter(
                        x=serie_indicador["Data"],
                        y=serie_indicador["Atendimentos"],
                        mode="lines+markers",
                        line=dict(color=SEMANTIC_COLORS["series_2"], width=2.8),
                        marker=dict(size=6, color=SEMANTIC_COLORS["series_2"]),
                        hovertemplate="<b>%{x|%d/%m/%Y}</b><br>Atendimentos: %{y:,.0f}<extra></extra>",
                    )
                )
                fig_ind = apply_plotly_theme(
                    fig_ind,
                    title=f"{descricao}",
                    subtitle="",
                    yaxis_title="Atendimentos",
                    height=300,
                    legend=False,
                )
                st.plotly_chart(fig_ind, use_container_width=True, key=f"samu_indicador_{idx}")
                st.divider()

        section_end()

    with tab_resumo:
        section_start("", "")
        resumo_periodo = procedimentos_total.rename(columns={"Atendimentos": "Total_Periodo"})
        tabela_resumo = resumo_periodo.merge(
            resumo[["Descricao", "Codigo_SIGTAP", "Meta", "Falta", "Eficacia"]],
            on=["Descricao", "Codigo_SIGTAP"],
            how="left",
        )
        tabela_resumo = tabela_resumo.sort_values("Total_Periodo", ascending=False).reset_index(drop=True)

        if "Eficacia" in tabela_resumo.columns:
            tabela_resumo["Eficacia_pct"] = (pd.to_numeric(tabela_resumo["Eficacia"], errors="coerce") * 100).round(1)
        else:
            tabela_resumo["Eficacia_pct"] = pd.NA

        st.dataframe(
            tabela_resumo[
                [
                    "Descricao",
                    "Codigo_SIGTAP",
                    "Total_Periodo",
                    "Meta",
                    "Falta",
                    "Eficacia_pct",
                ]
            ].rename(columns={
                "Descricao": "Descrição",
                "Codigo_SIGTAP": "Cód. SIGTAP",
                "Total_Periodo": "Total no período",
                "Meta": "Meta",
                "Falta": "Falta",
                "Eficacia_pct": "% Eficácia",
            }),
            use_container_width=True,
        )
        section_end()


# ---------------------------------------------------------------------------
# MAPA DE CALOR — página dedicada
# ---------------------------------------------------------------------------

_DIAS_SEMANA_PT = ["Domingo", "Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado"]
_DIAS_SEMANA_ORDER = {d: i for i, d in enumerate(_DIAS_SEMANA_PT)}

_HEATMAP_COLORSCALE = [
    [0.0,  "#63BE7B"],   # verde  — mínimo  (igual Excel)
    [0.5,  "#FFEB84"],   # amarelo — p50    (igual Excel)
    [1.0,  "#F8696B"],   # vermelho — máximo (igual Excel)
]


def _heatmap_fig(matrix_df, title, x_label, y_label, colorscale=None, fmt=".0f", height=320):
    """Gera go.Heatmap padronizado a partir de um DataFrame pivotado (linhas=y, colunas=x)."""
    import plotly.graph_objects as go

    z = matrix_df.values
    x = list(matrix_df.columns)
    y = list(matrix_df.index)

    text = [[f"{v:{fmt}}" if pd.notna(v) else "" for v in row] for row in z]

    # Calcula zmin/zmax excluindo a linha TOTAL (para não comprimir a escala de cores)
    mask = [lbl != "TOTAL" for lbl in y]
    z_no_total = [row for row, keep in zip(z, mask) if keep]
    import numpy as np
    flat = [v for row in z_no_total for v in row if v is not None and not (isinstance(v, float) and np.isnan(v))]
    zmin_val = min(flat) if flat else None
    zmax_val = max(flat) if flat else None

    fig = go.Figure(
        go.Heatmap(
            z=z,
            x=x,
            y=y,
            text=text,
            texttemplate="%{text}",
            textfont={"size": 11, "color": "#1a1a1a"},
            colorscale=colorscale or _HEATMAP_COLORSCALE,
            zmin=zmin_val,
            zmax=zmax_val,
            hoverongaps=False,
            showscale=True,
        )
    )
    fig.update_layout(
        title={"text": title, "font": {"size": 15, "color": "#e0e0e0"}, "x": 0.02},
        xaxis_title=x_label,
        yaxis_title=y_label,
        height=height,
        margin={"l": 155, "r": 60, "t": 50, "b": 60},
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font={"color": "#e0e0e0"},
        xaxis={"gridcolor": "rgba(0,0,0,0)", "tickfont": {"color": "#e0e0e0"}, "side": "top"},
        yaxis={"gridcolor": "rgba(0,0,0,0)", "autorange": "reversed", "tickfont": {"color": "#e0e0e0"}},
    )
    return fig


def render_heatmap_page():
    """
    Página dedicada: Mapa de Calor — dados transacionais CELK (hora × dia da semana).
    Fallback para dados KPI agregados quando o arquivo CELK não está disponível.
    """
    import plotly.graph_objects as go

    _DOW_ORDER = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta", "Sábado", "Domingo"]

    # ── Tenta carregar dados CELK (ricos, por hora) ──────────────────────
    celk = load_celk_data(_mtime=_celk_mtime())
    has_celk = celk is not None and not celk.empty

    if has_celk:
        # Unidades detectadas no CELK
        unidades_celk = sorted(
            [u for u in celk["UNIDADE_PAINEL"].dropna().unique() if _is_upa_unit(u)],
            key=lambda x: (0 if "II" in (normalize_text(x) or "") else 1, x)
        )
        # Ordena meses cronologicamente e filtra outliers (>= 50 registros no mês)
        _mes_counts = celk["MES_LABEL"].value_counts()
        _mes_validos = _mes_counts[_mes_counts >= 50].index.tolist()
        meses_celk = sorted(
            [m for m in celk["MES_LABEL"].dropna().unique() if m in _mes_validos],
            key=lambda m: pd.to_datetime(m, format="%b/%y", errors="coerce")
        )
        # Padrão: preferir "Mar/26" se disponível, senão o mês com mais dados
        _todos_meses_list = ["Todos"] + meses_celk
        if "Mar/26" in meses_celk:
            _idx_mes_default = _todos_meses_list.index("Mar/26")
        else:
            _mes_dominante = _mes_counts.idxmax() if not _mes_counts.empty else None
            _idx_mes_default = _todos_meses_list.index(_mes_dominante) if _mes_dominante in meses_celk else len(meses_celk)
    else:
        st.info("📁 Arquivo CELK não encontrado em `PowerBI/data_raw/`. Exibindo dados agregados.")

    # ════════════════════════════════════════════════════════════════════
    # SEÇÃO 1 — Hora × Dia da Semana  (CELK)
    # ════════════════════════════════════════════════════════════════════
    if has_celk:
        st.markdown(
            """
            <style>
            .stTabs [data-baseweb="tab-list"] {
                gap: 0.6rem;
                padding: 0.2rem 0.1rem 0.5rem 0.1rem;
            }

            .stTabs [data-baseweb="tab"] {
                background: linear-gradient(180deg, rgba(15,108,189,0.12) 0%, rgba(15,108,189,0.05) 100%);
                border: 1px solid rgba(15,108,189,0.35);
                border-radius: 12px;
                padding: 0.5rem 0.95rem;
                min-height: 44px;
                box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
            }

            .stTabs [data-baseweb="tab"] p {
                font-size: 1rem;
                font-weight: 800;
                letter-spacing: 0.01em;
                color: #0b3b69;
            }

            .stTabs [aria-selected="true"] {
                background: linear-gradient(180deg, rgba(15,108,189,0.24) 0%, rgba(15,108,189,0.12) 100%);
                border: 2px solid #0F6CBD;
                box-shadow: 0 0 0 1px rgba(15,108,189,0.15), 0 8px 18px rgba(15, 108, 189, 0.18);
                transform: translateY(-1px);
            }

            .stTabs [aria-selected="true"] p {
                color: #083055;
            }

            .st-key-hm_celk_mes_hora,
            .st-key-hm_celk_unidades_hora,
            .st-key-hm_celk_metrica_hora {
                margin-top: 0.08rem;
            }

            .st-key-hm_celk_mes_hora label p,
            .st-key-hm_celk_unidades_hora label p,
            .st-key-hm_celk_metrica_hora label p {
                font-size: 0.70rem !important;
                font-weight: 700 !important;
                text-transform: uppercase;
                letter-spacing: 0.06em;
                color: rgba(148, 163, 184, 0.88) !important;
                margin-bottom: 0.22rem !important;
            }

            .st-key-hm_celk_mes_hora [data-baseweb="select"] > div,
            .st-key-hm_celk_unidades_hora [data-baseweb="select"] > div,
            .st-key-hm_celk_metrica_hora [data-baseweb="select"] > div {
                min-height: 2.14rem;
                border-radius: 12px;
                border: 1px solid rgba(148, 163, 184, 0.24);
                background: linear-gradient(180deg, rgba(15, 23, 42, 0.22) 0%, rgba(15, 23, 42, 0.15) 100%);
                box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.03), 0 1px 3px rgba(15, 23, 42, 0.08);
                transition: border-color 0.2s ease, box-shadow 0.2s ease, background 0.2s ease;
            }

            .st-key-hm_celk_mes_hora [data-baseweb="select"] > div:focus-within,
            .st-key-hm_celk_unidades_hora [data-baseweb="select"] > div:focus-within,
            .st-key-hm_celk_metrica_hora [data-baseweb="select"] > div:focus-within {
                border-color: rgba(56, 189, 248, 0.46);
                box-shadow: 0 0 0 1px rgba(56, 189, 248, 0.14);
                background: linear-gradient(180deg, rgba(15, 23, 42, 0.28) 0%, rgba(15, 23, 42, 0.18) 100%);
            }

            .st-key-hm_celk_mes_hora [data-baseweb="select"] span,
            .st-key-hm_celk_unidades_hora [data-baseweb="select"] span,
            .st-key-hm_celk_metrica_hora [data-baseweb="select"] span {
                font-size: 0.84rem !important;
                color: rgba(226, 232, 240, 0.96) !important;
            }

            .st-key-hm_celk_unidades_hora [data-baseweb="tag"] {
                height: 1.40rem;
                border-radius: 999px;
                padding: 0 0.34rem;
                margin: 0.08rem 0.22rem 0.08rem 0;
                border: 1px solid rgba(148, 163, 184, 0.28);
                background: linear-gradient(180deg, rgba(241, 245, 249, 0.14) 0%, rgba(226, 232, 240, 0.08) 100%);
                box-shadow: none;
                transition: border-color 0.18s ease, background 0.18s ease;
            }

            .st-key-hm_celk_unidades_hora [data-baseweb="tag"]:hover {
                border-color: rgba(148, 163, 184, 0.44);
                background: linear-gradient(180deg, rgba(241, 245, 249, 0.20) 0%, rgba(226, 232, 240, 0.12) 100%);
            }

            .st-key-hm_celk_unidades_hora [data-baseweb="tag"] span {
                font-size: 0.74rem !important;
                font-weight: 600;
                letter-spacing: 0.01em;
                color: rgba(226, 232, 240, 0.94) !important;
            }

            .st-key-hm_celk_unidades_hora [data-baseweb="tag"] svg {
                width: 12px;
                height: 12px;
                opacity: 0.60;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )

        _grupos_cfg = [
            ("Atenção Básica",     "💊", "hm_bas",   True),
            ("Atenção Secundária", "🏨", "hm_sec",   False),
            ("Odontologia",        "🦷", "hm_odont", False),
        ]
        _grupos_disponiveis = []
        for _gnome, _gicon, _gpfx, _default_one in _grupos_cfg:
            _gunids = sorted(
                celk[celk["GRUPO_PAINEL"] == _gnome]["UNIDADE_PAINEL"].dropna().unique().tolist()
            )
            if _gunids:
                _grupos_disponiveis.append((_gnome, _gicon, _gpfx, _default_one, _gunids))

        _tab_labels = ["🔥 Mapa de Calor — Hora × Dia da Semana"]
        _tab_labels.extend([f"{_gicon} {_gnome}" for _gnome, _gicon, _, _, _ in _grupos_disponiveis])
        _tabs = st.tabs(_tab_labels)
        tab_fluxo_horario = _tabs[0]
        with tab_fluxo_horario:
            section_start(
                "",
                "<strong>🕐 Heatmap de Fluxo Horário — Hora × Dia da Semana</strong>"
            )

            with st.expander("Filtros", expanded=False):
                col_f1, col_f2, col_f3 = st.columns([1.6, 2.4, 1.2])
                mes_hora = col_f1.selectbox(
                    "Mês",
                    ["Todos"] + meses_celk,
                    index=_idx_mes_default,  # mês com mais dados por padrão
                    key="hm_celk_mes_hora",
                )
                _upas_opcoes = unidades_celk if unidades_celk else []
                unidades_hora = col_f2.multiselect(
                    "Unidades",
                    options=_upas_opcoes,
                    default=_upas_opcoes[:1],
                    key="hm_celk_unidades_hora",
                )
                metrica_hora = col_f3.selectbox(
                    "Métrica",
                    ["Média/dia", "Total", "Pacientes únicos"],
                    key="hm_celk_metrica_hora",
                )

            df_h = celk[celk["UNIDADE_PAINEL"].isin(unidades_hora)].copy()
            if mes_hora != "Todos":
                df_h = df_h[df_h["MES_LABEL"] == mes_hora]

            if not unidades_hora:
                st.warning("Selecione ao menos uma unidade.")
            elif df_h.empty:
                st.warning("Sem dados para o filtro selecionado.")
            else:
                if metrica_hora == "Pacientes únicos":
                    raw = (
                        df_h.groupby(["HORA", "DIA_SEMANA"])["PACIENTE"]
                        .nunique()
                        .reset_index()
                        .pivot(index="HORA", columns="DIA_SEMANA", values="PACIENTE")
                    )
                    fmt_str = ".0f"
                elif metrica_hora == "Média/dia":
                    dias_por_dow = df_h.groupby("DIA_SEMANA")["DATA"].apply(lambda s: s.dt.date.nunique())
                    contagem = (
                        df_h.groupby(["HORA", "DIA_SEMANA"])
                        .size()
                        .reset_index(name="QTD")
                        .pivot(index="HORA", columns="DIA_SEMANA", values="QTD")
                    )
                    raw = contagem.div(dias_por_dow).round(1)
                    fmt_str = ".1f"
                else:
                    raw = (
                        df_h.groupby(["HORA", "DIA_SEMANA"])
                        .size()
                        .reset_index(name="QTD")
                        .pivot(index="HORA", columns="DIA_SEMANA", values="QTD")
                    )
                    fmt_str = ".0f"

                raw = raw.reindex(columns=_DOW_ORDER)
                raw = raw.reindex(sorted(raw.index))

                total_row = raw.sum(axis=0)
                pivot_hora = pd.concat([raw, total_row.rename("TOTAL").to_frame().T])

                hora_labels = [
                    f"{int(h):02d}:00 às {int(h):02d}:59" if h != "TOTAL" else "TOTAL"
                    for h in pivot_hora.index
                ]
                pivot_hora.index = hora_labels
                pivot_hora.index.name = "Hora"
                pivot_hora.columns.name = "Dia da Semana"

                _lbl_upas = ", ".join(unidades_hora) if len(unidades_hora) <= 2 else f"{len(unidades_hora)} unidades"
                titulo = f"UPAs ({_lbl_upas}) — {metrica_hora} por Hora × Dia da Semana"
                if mes_hora != "Todos":
                    titulo += f"  ({mes_hora})"

                fig = _heatmap_fig(pivot_hora, titulo, "Dia da Semana", "Hora do Dia", height=700, fmt=fmt_str)
                st.plotly_chart(fig, use_container_width=True)

                pivot_sem_total = pivot_hora.drop("TOTAL", errors="ignore")
                k1, k2, k3, k4 = st.columns(4)
                pico_hora_lbl = pivot_sem_total.sum(axis=1).idxmax()
                pico_dia_lbl = pivot_sem_total.sum(axis=0).idxmax()
                total_atend = int(raw.sum().sum())
                media_hora = round(raw.sum(axis=1).mean(), 1)
                k1.metric("🔺 Hora de pico", pico_hora_lbl)
                k2.metric("📅 Dia de pico", pico_dia_lbl)
                k3.metric("🔢 Total no período", f"{total_atend:,}")
                k4.metric("⌀ Média/hora", f"{media_hora:,.1f}")

            section_end()

        for _tab, (_gnome, _gicon, _gpfx, _default_one, _gunids) in zip(_tabs[1:], _grupos_disponiveis):
            with _tab:
                _gc1, _gc2, _gc3 = st.columns([2, 3, 1])
                _gmes = _gc1.selectbox(
                    "Mês", ["Todos"] + meses_celk,
                    index=_idx_mes_default,
                    key=f"{_gpfx}_mes",
                )
                _gsel = _gc2.multiselect(
                    "Unidades",
                    options=_gunids,
                    default=_gunids[:1] if _default_one else _gunids,
                    key=f"{_gpfx}_unidades",
                )
                _gmet = _gc3.selectbox(
                    "Métrica",
                    ["Média/dia", "Total", "Pacientes únicos"],
                    key=f"{_gpfx}_metrica",
                )

                _gdf = celk[celk["GRUPO_PAINEL"] == _gnome].copy()
                if _gmes != "Todos":
                    _gdf = _gdf[_gdf["MES_LABEL"] == _gmes]
                if _gsel:
                    _gdf = _gdf[_gdf["UNIDADE_PAINEL"].isin(_gsel)]

                if _gdf.empty or not _gsel:
                    st.warning("Selecione ao menos uma unidade.")
                else:
                    if _gmet == "Pacientes únicos":
                        _graw = (
                            _gdf.groupby(["HORA", "DIA_SEMANA"])["PACIENTE"]
                            .nunique().reset_index()
                            .pivot(index="HORA", columns="DIA_SEMANA", values="PACIENTE")
                        )
                        _gfmt = ".0f"
                    elif _gmet == "Média/dia":
                        _gdias = _gdf.groupby("DIA_SEMANA")["DATA"].apply(lambda s: s.dt.date.nunique())
                        _gcnt = (
                            _gdf.groupby(["HORA", "DIA_SEMANA"])
                            .size().reset_index(name="QTD")
                            .pivot(index="HORA", columns="DIA_SEMANA", values="QTD")
                        )
                        _graw = _gcnt.div(_gdias).round(1)
                        _gfmt = ".1f"
                    else:
                        _graw = (
                            _gdf.groupby(["HORA", "DIA_SEMANA"])
                            .size().reset_index(name="QTD")
                            .pivot(index="HORA", columns="DIA_SEMANA", values="QTD")
                        )
                        _gfmt = ".0f"

                    _graw = _graw.reindex(columns=_DOW_ORDER).reindex(sorted(_graw.index))
                    _gpivot = pd.concat([_graw, _graw.sum(axis=0).rename("TOTAL").to_frame().T])
                    _gpivot.index = [
                        f"{int(h):02d}:00 às {int(h):02d}:59" if h != "TOTAL" else "TOTAL"
                        for h in _gpivot.index
                    ]
                    _gpivot.index.name = "Hora"
                    _gpivot.columns.name = "Dia da Semana"

                    _glbl = ", ".join(_gsel) if len(_gsel) <= 2 else f"{len(_gsel)} unidades"
                    _gtit = f"{_gnome} ({_glbl}) — {_gmet} por Hora × Dia da Semana"
                    if _gmes != "Todos":
                        _gtit += f"  ({_gmes})"

                    st.plotly_chart(
                        _heatmap_fig(_gpivot, _gtit, "Dia da Semana", "Hora do Dia", height=700, fmt=_gfmt),
                        use_container_width=True,
                    )

                    _gpnt = _gpivot.drop("TOTAL", errors="ignore")
                    _gk1, _gk2, _gk3, _gk4 = st.columns(4)
                    _gk1.metric("🔺 Hora de pico",    _gpnt.sum(axis=1).idxmax())
                    _gk2.metric("📅 Dia de pico",      _gpnt.sum(axis=0).idxmax())
                    _gk3.metric("🔢 Total no período", f"{int(_graw.sum().sum()):,}")
                    _gk4.metric("⌀ Média/hora",        f"{round(_graw.sum(axis=1).mean(), 1):,.1f}")


    # ════════════════════════════════════════════════════════════════════
    # SEÇÃO 2 — Calendário diário (Semana × Dia da Semana)  (CELK)
    # ════════════════════════════════════════════════════════════════════
    if has_celk:
        section_start(
            "📅 Calendário Mensal de Atendimentos",
            "Volume por semana do mês × dia da semana — "
            "estilo calendário, útil para identificar semanas atípicas"
        )

        col_c1, col_c2 = st.columns([2, 2])
        mes_cal = col_c1.selectbox("Mês", meses_celk, index=len(meses_celk) - 1, key="hm_celk_mes_cal")
        if not unidades_celk:
            st.warning("Sem unidades UPA mapeadas para exibir o calendário mensal.")
            section_end()
            return

        unidade_cal = col_c2.selectbox("Unidade", unidades_celk, key="hm_celk_unidade_cal")

        df_cal = celk[(celk["UNIDADE_PAINEL"] == unidade_cal) & (celk["MES_LABEL"] == mes_cal)].copy()

        if df_cal.empty:
            st.warning("Sem dados para o filtro selecionado.")
        else:
            pivot_cal = (
                df_cal.groupby(["SEMANA_MES", "DIA_SEMANA"])
                .size()
                .reset_index(name="QTD")
                .pivot(index="SEMANA_MES", columns="DIA_SEMANA", values="QTD")
            )
            pivot_cal = pivot_cal.reindex(columns=[d for d in _DOW_ORDER if d in pivot_cal.columns])
            semana_order = [f"S{i}" for i in range(1, 7)]
            pivot_cal = pivot_cal.reindex([s for s in semana_order if s in pivot_cal.index])

            fig_cal = _heatmap_fig(
                pivot_cal,
                f"{unidade_cal} — Calendário {mes_cal}",
                "Dia da Semana",
                "Semana do Mês",
            )
            st.plotly_chart(fig_cal, use_container_width=True)

        section_end()


def render_mapa_territorial_page(file_bytes=None, _mtime=None):
    st.markdown(
        """
        <style>
        .territorial-wrap {
            background:
                radial-gradient(1200px 420px at 0% 0%, rgba(14, 116, 144, 0.10), rgba(14, 116, 144, 0.00) 60%),
                radial-gradient(1000px 360px at 100% 0%, rgba(30, 64, 175, 0.08), rgba(30, 64, 175, 0.00) 62%);
            border-radius: 24px;
            padding: 0.4rem 0.3rem 0.8rem 0.3rem;
        }

        .territorial-wrap .section-card--territorial {
            background: linear-gradient(180deg, #FFFFFF 0%, #F8FAFC 100%);
            border: 1px solid #E2E8F0;
            border-radius: 20px;
            padding: 1.0rem 1.05rem 1.05rem 1.05rem;
            margin-bottom: 1rem;
            box-shadow: 0 10px 24px rgba(15, 23, 42, 0.06);
            position: relative;
            overflow: hidden;
        }

        .territorial-wrap .section-card--territorial::before {
            content: "";
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 3px;
            background: linear-gradient(90deg, #0EA5E9 0%, #22C55E 48%, #F59E0B 100%);
            opacity: 0.82;
        }

        .territorial-wrap .section-card--territorial .section-title {
            font-size: 1.14rem;
            font-weight: 800;
            letter-spacing: -0.35px;
            color: #0B1F33;
            margin-bottom: 0.22rem;
            line-height: 1.2;
        }

        .territorial-wrap .section-card--territorial .section-subtitle {
            font-size: 0.85rem;
            font-weight: 500;
            color: #64748B;
            margin-bottom: 0.95rem;
            line-height: 1.35;
        }

        .territorial-wrap .section-card--territorial [data-testid="stDataFrame"],
        .territorial-wrap .section-card--territorial .stPlotlyChart,
        .territorial-wrap .section-card--territorial iframe {
            border-radius: 16px !important;
            overflow: hidden;
            border: 1px solid #E5EAF1;
            box-shadow: 0 6px 16px rgba(15, 23, 42, 0.05);
            background: #FFFFFF;
        }

        .territorial-wrap .section-card--territorial [data-testid="stCaptionContainer"] p {
            color: #6B7280 !important;
            font-size: 0.80rem !important;
        }

        .territorial-wrap .mt-control-card {
            border: 1px solid #E6ECF3;
            border-radius: 12px;
            background: linear-gradient(180deg, #FBFDFF 0%, #F6F9FC 100%);
            padding: 0.55rem 0.7rem 0.45rem 0.7rem;
            margin-bottom: 0.45rem;
        }

        .territorial-wrap .mt-control-title {
            font-size: 0.76rem;
            font-weight: 700;
            letter-spacing: 0.3px;
            text-transform: uppercase;
            color: #334155;
            margin-bottom: 0.15rem;
        }

        .territorial-wrap .mt-control-sub {
            font-size: 0.72rem;
            color: #64748B;
            line-height: 1.28;
        }

        .territorial-wrap .streamlit-expanderHeader {
            border: 1px solid #E6ECF3 !important;
            border-radius: 10px !important;
            background: linear-gradient(180deg, #F8FAFC 0%, #F1F5F9 100%) !important;
            color: #334155 !important;
            font-size: 0.8rem !important;
            font-weight: 700 !important;
            padding-top: 0.22rem !important;
            padding-bottom: 0.22rem !important;
        }

        .territorial-wrap .section-card--territorial [data-testid="stMultiSelect"] label,
        .territorial-wrap .section-card--territorial [data-testid="stSelectbox"] label,
        .territorial-wrap .section-card--territorial [data-testid="stSlider"] label,
        .territorial-wrap .section-card--territorial [data-testid="stCheckbox"] label {
            color: #64748B !important;
            font-size: 0.76rem !important;
            font-weight: 600 !important;
            letter-spacing: 0.1px;
            margin-bottom: 0.24rem !important;
        }

        .territorial-wrap .section-card--territorial [data-testid="stMultiSelect"] > div > div,
        .territorial-wrap .section-card--territorial [data-testid="stSelectbox"] > div > div {
            border-radius: 10px !important;
            border: 1px solid #E6ECF3 !important;
            background: linear-gradient(180deg, #F8FAFC 0%, #F1F5F9 100%) !important;
            min-height: 2.25rem !important;
            box-shadow: none !important;
            transition: border-color 0.18s ease, background 0.18s ease;
        }

        .territorial-wrap .section-card--territorial [data-testid="stMultiSelect"] > div > div:hover,
        .territorial-wrap .section-card--territorial [data-testid="stSelectbox"] > div > div:hover {
            border-color: #CBD5E1 !important;
            background: linear-gradient(180deg, #F8FAFC 0%, #EEF2F7 100%) !important;
        }

        .territorial-wrap .section-card--territorial [data-testid="stMultiSelect"] > div > div:focus-within,
        .territorial-wrap .section-card--territorial [data-testid="stSelectbox"] > div > div:focus-within {
            border-color: #94A3B8 !important;
            box-shadow: 0 0 0 2px rgba(148, 163, 184, 0.14) !important;
            background: #F8FAFC !important;
        }

        .territorial-wrap .section-card--territorial [data-testid="stMultiSelect"] [data-baseweb="tag"] {
            border-radius: 7px !important;
            border: 1px solid #D8E1EA !important;
            background: #EEF3F8 !important;
            color: #334155 !important;
            font-size: 0.72rem !important;
            font-weight: 600 !important;
            padding: 1px 6px !important;
        }

        .territorial-wrap .section-card--territorial [data-testid="stSlider"] > div {
            border-radius: 10px;
            border: 1px solid #E6ECF3;
            background: linear-gradient(180deg, #F8FAFC 0%, #F1F5F9 100%);
            padding: 0.42rem 0.56rem 0.2rem 0.56rem;
        }

        .territorial-wrap .section-card--territorial [data-testid="stSlider"] p {
            font-size: 0.73rem !important;
            color: #64748B !important;
        }

        .territorial-wrap .section-card--territorial [data-testid="stSlider"] [data-baseweb="slider"] div[role="slider"] {
            box-shadow: 0 0 0 1px #94A3B8;
            background: #FFFFFF !important;
            width: 12px !important;
            height: 12px !important;
            margin-top: -2px !important;
        }

        .territorial-wrap .section-card--territorial [data-testid="stSlider"] [data-baseweb="slider"] > div > div:nth-child(1) {
            background: #CBD5E1 !important;
            height: 0.16rem !important;
        }

        .territorial-wrap .section-card--territorial [data-testid="stSlider"] [data-baseweb="slider"] > div > div:nth-child(2) {
            background: #94A3B8 !important;
            height: 0.16rem !important;
        }

        .territorial-wrap .section-card--territorial [data-testid="stCheckbox"] {
            border-radius: 10px;
            border: 1px solid #E6ECF3;
            background: linear-gradient(180deg, #F8FAFC 0%, #F1F5F9 100%);
            padding: 0.45rem 0.6rem;
        }

        .territorial-wrap .section-card--territorial [data-testid="stCheckbox"] label p {
            color: #475569 !important;
            font-size: 0.78rem !important;
            line-height: 1.28 !important;
        }

        .territorial-wrap .section-card--territorial [data-testid="stMultiSelect"],
        .territorial-wrap .section-card--territorial [data-testid="stSelectbox"],
        .territorial-wrap .section-card--territorial [data-testid="stSlider"],
        .territorial-wrap .section-card--territorial [data-testid="stCheckbox"] {
            margin-bottom: 0.38rem;
        }

        .territorial-wrap .section-card--territorial [data-baseweb="button-group"] {
            gap: 0.34rem;
            flex-wrap: wrap;
        }

        .territorial-wrap .section-card--territorial [data-baseweb="button-group"] button {
            border-radius: 999px !important;
            border: 1px solid #D7E1EB !important;
            background: #F3F7FB !important;
            color: #475569 !important;
            min-height: 1.95rem !important;
            padding: 0.18rem 0.72rem !important;
            font-size: 0.74rem !important;
            font-weight: 700 !important;
            box-shadow: none !important;
            transition: all 0.18s ease;
        }

        .territorial-wrap .section-card--territorial [data-baseweb="button-group"] button:hover {
            border-color: #AFC3D8 !important;
            background: #EAF1F8 !important;
            color: #334155 !important;
        }

        .territorial-wrap .section-card--territorial [data-baseweb="button-group"] button[aria-pressed="true"] {
            border-color: #7FA6CC !important;
            background: linear-gradient(180deg, #DCEAF7 0%, #D2E4F5 100%) !important;
            color: #1E3A5F !important;
        }

        .territorial-wrap .section-card--territorial [data-baseweb="tab-list"] {
            gap: 0.45rem;
            background: linear-gradient(180deg, #F8FAFC 0%, #EEF3F8 100%);
            border: 1px solid #DCE6F1;
            border-radius: 14px;
            padding: 0.35rem;
            margin-bottom: 0.85rem;
        }

        .territorial-wrap .section-card--territorial [data-baseweb="tab-list"] button {
            border-radius: 10px !important;
            border: 1px solid transparent !important;
            background: transparent !important;
            color: #475569 !important;
            font-size: 0.86rem !important;
            font-weight: 800 !important;
            letter-spacing: 0.1px;
            min-height: 2.2rem !important;
            padding: 0.22rem 0.8rem !important;
            transition: all 0.18s ease;
        }

        .territorial-wrap .section-card--territorial [data-baseweb="tab-list"] button:hover {
            background: #E8F0F8 !important;
            border-color: #C8D8EA !important;
            color: #1E3A5F !important;
        }

        .territorial-wrap .section-card--territorial [data-baseweb="tab-list"] button[aria-selected="true"] {
            background: linear-gradient(180deg, #DBECFB 0%, #D3E7F8 100%) !important;
            border-color: #A7C4E3 !important;
            color: #0F2F55 !important;
            box-shadow: 0 6px 14px rgba(30, 64, 175, 0.10);
        }

        .territorial-wrap .mt-tab-highlight {
            border: 1px solid #D6E3F2;
            border-radius: 12px;
            background: linear-gradient(120deg, rgba(14,165,233,0.08) 0%, rgba(59,130,246,0.05) 48%, rgba(34,197,94,0.06) 100%);
            padding: 0.55rem 0.7rem;
            margin-bottom: 0.55rem;
        }

        .territorial-wrap .mt-tab-highlight-title {
            font-size: 0.78rem;
            font-weight: 900;
            text-transform: uppercase;
            letter-spacing: 0.32px;
            color: #1E3A5F;
            margin-bottom: 0.1rem;
        }

        .territorial-wrap .mt-tab-highlight-sub {
            font-size: 0.76rem;
            font-weight: 600;
            color: #475569;
            line-height: 1.25;
        }
        </style>
        <div class="territorial-wrap">
        """,
        unsafe_allow_html=True,
    )

    def territorial_heading(title):
        st.markdown(
            f"""
            <div style=\"margin: 8px 0 12px 0; font-size: 20px; font-weight: 900; letter-spacing: 0.2px; color: #0B1220; line-height: 1.2;\">{title}</div>
            """,
            unsafe_allow_html=True,
        )

    section_start("", "", theme="territorial")

    geo_df, ranking_df, colab_df, erros = load_mapa_territorial_data(file_bytes=file_bytes, _mtime=_mtime)

    if erros:
        for err in erros:
            st.warning(err)
        section_end()
        st.markdown("</div>", unsafe_allow_html=True)
        return

    if geo_df.empty:
        st.warning("Sem dados suficientes para renderizar o mapa territorial.")
        section_end()
        st.markdown("</div>", unsafe_allow_html=True)
        return

    section_start("", "", theme="territorial")
    tipos = sorted(geo_df["tipo"].dropna().unique().tolist())
    filtro_col_left, filtro_col_right = st.columns([0.72, 0.28])
    with filtro_col_right:
        st.markdown(
            """
            <div class="mt-control-card">
                <div class="mt-control-title" style="font-size:14px; font-weight:600;">Tipo de Unidade</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        try:
            tipos_sel = st.segmented_control(
                "Tipo de unidade",
                options=tipos,
                default=tipos,
                selection_mode="multi",
                key="mt_tipo_unidade",
                label_visibility="collapsed",
            )
        except Exception:
            tipos_sel = st.multiselect(
                "Tipo de unidade",
                options=tipos,
                default=tipos,
                key="mt_tipo_unidade",
                label_visibility="collapsed",
            )

    if tipos_sel is None:
        tipos_sel = []
    elif isinstance(tipos_sel, str):
        tipos_sel = [tipos_sel]
    else:
        tipos_sel = list(tipos_sel)

    section_end()

    mapa_df = geo_df[geo_df["tipo"].isin(tipos_sel)].copy() if tipos_sel else geo_df.iloc[0:0].copy()
    ranking_df_filtrado = ranking_df[ranking_df["tipo"].isin(tipos_sel)].copy() if tipos_sel else ranking_df.iloc[0:0].copy()
    if mapa_df.empty:
        st.info("Nenhuma unidade encontrada para o filtro selecionado.")
        section_end()
        st.markdown("</div>", unsafe_allow_html=True)
        return

    min_colab = int(mapa_df["qtd_colaboradores"].min())
    max_colab = int(mapa_df["qtd_colaboradores"].max())

    def territorial_help_badge(help_text):
        return (
            '<span title="' + html.escape(help_text, quote=True) + '" '
            'style="cursor:help; margin-left:4px; color:#334155; font-weight:700;">&#9432;</span>'
        )

    def territorial_info_line(label, help_text):
        # Mantido como no-op para deixar a página mais limpa, exibindo só os títulos principais.
        return

    def add_map_lock_control(map_obj, initially_locked=True):
        map_var = map_obj.get_name()
        script = f"""
        <script>
        (function() {{
            function boot(attempt) {{
                var map = window['{map_var}'];
                if (!map) {{
                    if (attempt < 24) setTimeout(function() {{ boot(attempt + 1); }}, 120);
                    return;
                }}

                var locked = {str(True if initially_locked else False).lower()};
                var buttonRef = null;

                function applyLockState() {{
                    if (locked) {{
                        if (map.dragging) map.dragging.disable();
                        if (map.touchZoom) map.touchZoom.disable();
                        if (map.doubleClickZoom) map.doubleClickZoom.disable();
                        if (map.scrollWheelZoom) map.scrollWheelZoom.disable();
                        if (map.boxZoom) map.boxZoom.disable();
                        if (map.keyboard) map.keyboard.disable();
                        if (map.tap) map.tap.disable();
                    }} else {{
                        if (map.dragging) map.dragging.enable();
                        if (map.touchZoom) map.touchZoom.enable();
                        if (map.doubleClickZoom) map.doubleClickZoom.enable();
                        if (map.scrollWheelZoom) map.scrollWheelZoom.enable();
                        if (map.boxZoom) map.boxZoom.enable();
                        if (map.keyboard) map.keyboard.enable();
                        if (map.tap) map.tap.enable();
                    }}

                    if (buttonRef) {{
                        buttonRef.innerHTML = locked ? '🔒 Mapa travado' : '🔓 Destravar ativo';
                        buttonRef.title = locked ? 'Clique para destravar o movimento do mapa' : 'Clique para travar novamente';
                        buttonRef.style.background = locked ? 'rgba(15,23,42,0.88)' : 'rgba(2,132,199,0.92)';
                    }}
                }}

                var LockToggleControl = L.Control.extend({{
                    options: {{ position: 'topright' }},
                    onAdd: function() {{
                        var container = L.DomUtil.create('div', 'leaflet-bar leaflet-control');
                        container.style.background = 'transparent';
                        container.style.border = 'none';
                        container.style.boxShadow = 'none';

                        var btn = L.DomUtil.create('button', '', container);
                        buttonRef = btn;
                        btn.type = 'button';
                        btn.style.height = '30px';
                        btn.style.padding = '0 10px';
                        btn.style.borderRadius = '8px';
                        btn.style.border = '1px solid rgba(255,255,255,0.35)';
                        btn.style.color = '#F8FAFC';
                        btn.style.fontSize = '12px';
                        btn.style.fontWeight = '700';
                        btn.style.cursor = 'pointer';
                        btn.style.boxShadow = '0 4px 10px rgba(15,23,42,0.28)';
                        btn.style.letterSpacing = '0.2px';
                        btn.style.backdropFilter = 'blur(2px)';

                        L.DomEvent.disableClickPropagation(container);
                        L.DomEvent.disableScrollPropagation(container);
                        L.DomEvent.on(btn, 'click', function(e) {{
                            L.DomEvent.stopPropagation(e);
                            L.DomEvent.preventDefault(e);
                            locked = !locked;
                            applyLockState();
                        }});

                        applyLockState();
                        return container;
                    }}
                }});

                map.addControl(new LockToggleControl());
                applyLockState();
            }}

            boot(0);
        }})();
        </script>
        """
        map_obj.get_root().html.add_child(folium.Element(script))

    def territorial_kpi_card(title, value, subtitle="", accent="#22C55E", icon="📊"):
        subtitle_tones = {
            "#38BDF8": "#0369A1",
            "#34D399": "#047857",
            "#F59E0B": "#B45309",
            "#F43F5E": "#BE123C",
            "#2563EB": "#1D4ED8",
            "#22C55E": "#15803D",
        }
        help_text = (
            f"Motivo: este KPI monitora {title.lower()} na leitura territorial. "
            f"Para que serve: {subtitle if subtitle else 'apoio direto a decisão operacional.'}"
        )
        top_kpi_card(
            title,
            value,
            icon=icon,
            subtitle="",
            accent_color=accent,
            subtitle_color=subtitle_tones.get(accent, "#475569"),
        )

    if max_colab == min_colab:
        mapa_df["marker_size"] = 40.0
    else:
        norm = (mapa_df["qtd_colaboradores"] - min_colab) / (max_colab - min_colab)
        mapa_df["marker_size"] = 10.0 + (norm.pow(0.50) * 80.0)

    maps_slot = st.container()

    section_start("", "", theme="territorial")
    territorial_heading("Mapa geográfico do município")
    mapa_render_df = mapa_df.copy().sort_values("qtd_colaboradores", ascending=True).reset_index(drop=True)
    mapa_render_df["quantidade_label"] = mapa_render_df["qtd_colaboradores"].astype(int).astype(str)
    mapa_render_df["tooltip_unidade"] = mapa_render_df["unidade"].fillna("Não informado")
    mapa_render_df["fill_color"] = mapa_render_df["tipo"].map({
        "UPA": [239, 68, 68, 220],
        "UBSF": [22, 163, 74, 220],
    })
    mapa_render_df["fill_color"] = mapa_render_df["fill_color"].apply(
        lambda color: color if isinstance(color, list) else [37, 99, 235, 220]
    )
    mapa_render_df["radius_m"] = mapa_render_df["marker_size"] * 22.0
    unidades_criticas_norm = {
        normalize_text("UPA II - UPA DE LUZIANIA"),
        normalize_text("UPA I - UPA DO JARDIM INGA"),
        normalize_text("UBSF - PARQUE ESTRELA DALVA IX FEIRINHA"),
        normalize_text("UBSF - MINGONE II A"),
        normalize_text("UBSF - PARQUE ALVORADA"),
        normalize_text("UBSF - VILA JURACY"),
        normalize_text("UBS - SETOR AEROPORTO"),
    }

    q_colab_60 = float(mapa_render_df["qtd_colaboradores"].quantile(0.60))
    q_colab_85 = float(mapa_render_df["qtd_colaboradores"].quantile(0.85))
    q_colab_95 = float(mapa_render_df["qtd_colaboradores"].quantile(0.95))
    unidades_criticas_mapa = int((mapa_render_df["qtd_colaboradores"] >= q_colab_85).sum())
    summary_slot = st.container()

    map_center = [
        float(mapa_render_df["latitude"].median()),
        float(mapa_render_df["longitude"].median()),
    ]
    initial_zoom_level = 11.5

    mapa_folium = folium.Map(
        location=map_center,
        zoom_start=initial_zoom_level,
        tiles="CartoDB Voyager",
        control_scale=True,
    )

    cluster_layer = MarkerCluster(
        name="Clusters operacionais",
        icon_create_function="""
        function(cluster) {
            var count = cluster.getChildCount();
            var tone = '#1D4ED8';
            if (count >= 8) {
                tone = '#B91C1C';
            } else if (count >= 5) {
                tone = '#B45309';
            } else if (count >= 3) {
                tone = '#0F766E';
            }
            return L.divIcon({
                html: '<div style="background:' + tone + ';border:2px solid rgba(255,255,255,0.85);width:34px;height:34px;border-radius:50%;display:flex;align-items:center;justify-content:center;color:#fff;font-weight:800;box-shadow:0 6px 14px rgba(15,23,42,0.30);">' + count + '</div>',
                className: 'territorial-cluster-icon',
                iconSize: [34, 34]
            });
        }
        """,
    )
    cluster_layer.add_to(mapa_folium)

    for _, row in mapa_render_df.iterrows():
        qtd = int(row["qtd_colaboradores"])
        if qtd >= q_colab_95:
            cor = "#b91c1c"
            nivel = "Critico"
        elif qtd >= q_colab_85:
            cor = "#ea580c"
            nivel = "Alto"
        elif qtd >= q_colab_60:
            cor = "#0284c7"
            nivel = "Moderado"
        else:
            cor = "#0f766e"
            nivel = "Base"

        raio = max(15, min(31, int(round(float(row["marker_size"]) / 3.2))))
        tooltip_html = (
            f"<b>{html.escape(str(row['unidade']))}</b><br>"
            f"Tipo: {html.escape(str(row['tipo']))}<br>"
            f"Lat/Lon: <b>{float(row['latitude']):.5f}, {float(row['longitude']):.5f}</b><br>"
            f"Colaboradores: <b>{qtd}</b><br>"
            f"Com CRM: <b>{int(row['qtd_medicos'])}</b><br>"
            f"Faixa operacional: <b>{nivel}</b>"
        )

        folium.CircleMarker(
            location=[float(row["latitude"]), float(row["longitude"])],
            radius=raio + 6,
            color=cor,
            weight=0,
            fill=True,
            fill_color=cor,
            fill_opacity=0.18,
            tooltip=folium.Tooltip(tooltip_html, sticky=True),
        ).add_to(mapa_folium)

        folium.CircleMarker(
            location=[float(row["latitude"]), float(row["longitude"])],
            radius=raio,
            color="#0F172A",
            weight=1.5,
            fill=True,
            fill_color=cor,
            fill_opacity=0.92,
            tooltip=folium.Tooltip(tooltip_html, sticky=True),
        ).add_to(mapa_folium)

        folium.Marker(
            location=[float(row["latitude"]), float(row["longitude"])],
            icon=folium.DivIcon(
                html=f'''
                <div style="
                    min-width: {raio * 1.15}px;
                    height: {raio * 0.95}px;
                    line-height: {raio * 0.95}px;
                    border-radius: 999px;
                    padding: 0 6px;
                    background: rgba(15, 23, 42, 0.70);
                    border: 1px solid rgba(255,255,255,0.35);
                    color: #F8FAFC;
                    text-align: center;
                    font-size: 11px;
                    font-weight: 800;
                    letter-spacing: 0.2px;
                    box-shadow: 0 3px 8px rgba(0,0,0,0.35);
                    transform: translate(-50%, -50%);
                ">{qtd}</div>
                '''
            ),
        ).add_to(mapa_folium)

        folium.Marker(
            location=[float(row["latitude"]), float(row["longitude"])],
            tooltip=folium.Tooltip(tooltip_html, sticky=True),
            icon=folium.DivIcon(
                html=(
                    "<div style='width:10px;height:10px;border-radius:50%;"
                    f"background:{cor};border:2px solid #ffffff;box-shadow:0 2px 8px rgba(15,23,42,0.35);'></div>"
                )
            ),
        ).add_to(cluster_layer)

    folium.LayerControl(collapsed=True).add_to(mapa_folium)
    add_map_lock_control(mapa_folium, initially_locked=True)
    mapa_geo_html = mapa_folium.get_root().render()

    unidades_criticas_df = (
        mapa_df[mapa_df["unidade_norm"].isin(unidades_criticas_norm)][["unidade", "tipo", "qtd_colaboradores"]]
        .sort_values(["tipo", "unidade"])
        .rename(columns={"unidade": "Unidade", "tipo": "Tipo", "qtd_colaboradores": "Colaboradores"})
    )
    if not unidades_criticas_df.empty:
        section_start("", "", theme="territorial")
        territorial_heading("Unidades verificadas")
        territorial_info_line(
            "Tabela de unidades verificadas",
            "Motivo: detalhar unidades monitoradas como críticas. Para que serve: apoiar validação rápida dos pontos de atenção no efetivo."
        )
        render_elegant_table(
            unidades_criticas_df,
            column_config={
                "Unidade": st.column_config.TextColumn("Unidade"),
                "Tipo": st.column_config.TextColumn("Tipo"),
                "Colaboradores": st.column_config.NumberColumn("Colaboradores", format="%d"),
            },
            emphasis_columns=["Unidade"],
            bar_columns=["Colaboradores"],
            key="mt_tbl_unidades_criticas",
        )
        section_end()

    section_end()

    celk = load_celk_data(_mtime=_celk_mtime())
    if celk is None or celk.empty:
        st.warning("Arquivo CELK não encontrado ou sem dados para gerar o mapa de calor.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    meses_celk = sorted(
        celk["MES_LABEL"].dropna().unique().tolist(),
        key=lambda m: pd.to_datetime(m, format="%b/%y", errors="coerce"),
    )
    opcoes_mes = ["Todos"] + meses_celk
    idx_mes = opcoes_mes.index("Mar/26") if "Mar/26" in opcoes_mes else 0

    mes_heat = st.session_state.get("mt_heat_mes", opcoes_mes[idx_mes] if opcoes_mes else "Todos")
    if mes_heat not in opcoes_mes:
        mes_heat = opcoes_mes[idx_mes] if opcoes_mes else "Todos"
    radius_heat = int(st.session_state.get("mt_heat_radius", 38))
    blur_heat = int(st.session_state.get("mt_heat_suavizacao", 20))
    opacidade_heat = float(st.session_state.get("mt_heat_opacidade", 0.78))
    foco_nucleo = bool(st.session_state.get("mt_heat_focus_core", True))
    mostrar_halo = bool(st.session_state.get("mt_heat_halo", True))

    celk_heat = celk.copy()
    if mes_heat != "Todos":
        celk_heat = celk_heat[celk_heat["MES_LABEL"] == mes_heat]

    if celk_heat.empty:
        st.info("Sem atendimentos para o período selecionado.")
        st.markdown("</div>", unsafe_allow_html=True)
        return

    unidade_col_celk = "UNIDADE" if "UNIDADE" in celk_heat.columns else "UNIDADE_PAINEL"

    def _unit_family(val):
        txt = normalize_text(val) or ""
        if "UPA" in txt:
            return "UPA"
        if any(tag in txt for tag in ["UBSF", "UBS", "PSF"]):
            return "UBS"
        return "OUTROS"

    def _unit_key(val):
        txt = normalize_text(val) or ""
        txt = re.sub(r"[^A-Z0-9 ]", " ", txt)
        tokens = [t for t in txt.split() if t]
        stop = {
            "UBS", "UBSF", "PSF", "UPA", "UNIDADE", "BASICA", "SAUDE", "DE", "DO", "DA", "DAS", "DOS",
            "PRONTO", "ATENDIMENTO", "SERVICO", "CENTRO",
        }
        tokens = [t for t in tokens if t not in stop]
        family = _unit_family(val)
        key_base = " ".join(tokens)
        return f"{family} {key_base}".strip()

    atend_por_unidade = (
        celk_heat[unidade_col_celk]
        .dropna()
        .astype(str)
        .str.strip()
        .to_frame(name="unidade_celk")
    )
    atend_por_unidade["unidade_norm"] = atend_por_unidade["unidade_celk"].map(normalize_text)
    atend_por_unidade["unidade_key"] = atend_por_unidade["unidade_celk"].map(_unit_key)
    atend_por_unidade = (
        atend_por_unidade.groupby(["unidade_norm", "unidade_key"], as_index=False)
        .size()
        .rename(columns={"size": "qtd_atendimentos"})
    )
    atend_por_norm = (
        atend_por_unidade.groupby("unidade_norm", as_index=False)["qtd_atendimentos"].sum()
    )
    atend_por_key = (
        atend_por_unidade.groupby("unidade_key", as_index=False)["qtd_atendimentos"].sum()
    )

    geo_heat = mapa_df[["unidade", "unidade_norm", "latitude", "longitude", "tipo"]].copy()
    geo_heat["unidade_key"] = geo_heat["unidade"].map(_unit_key)
    geo_heat = geo_heat.merge(
        atend_por_norm,
        on="unidade_norm",
        how="left",
    )
    geo_heat = geo_heat.merge(
        atend_por_key.rename(columns={"qtd_atendimentos": "qtd_atendimentos_key"}),
        on="unidade_key",
        how="left",
    )
    geo_heat["qtd_atendimentos"] = geo_heat["qtd_atendimentos"].fillna(geo_heat["qtd_atendimentos_key"])
    geo_heat = geo_heat.drop(columns=["qtd_atendimentos_key"])

    # Fallback de correspondência por aproximação textual quando nomes variam entre GEO e CELK.
    if geo_heat["qtd_atendimentos"].isna().any():
        cnt_map = {
            row["unidade_key"]: int(row["qtd_atendimentos"])
            for _, row in atend_por_key.iterrows()
            if pd.notna(row["unidade_key"])
        }
        cnt_map_norm = {
            row["unidade_norm"]: int(row["qtd_atendimentos"])
            for _, row in atend_por_norm.iterrows()
            if pd.notna(row["unidade_norm"])
        }

        def _approx_count(un_key, un_norm):
            if pd.isna(un_key) and pd.isna(un_norm):
                return 0
            if un_key in cnt_map:
                return cnt_map[un_key]
            if un_norm in cnt_map_norm:
                return cnt_map_norm[un_norm]

            un_key_txt = str(un_key or "")
            un_norm_txt = str(un_norm or "")
            un_tokens = set([t for t in un_key_txt.split() if t])
            best_score = 0.0
            best_value = 0

            for k, v in cnt_map.items():
                if not k:
                    continue
                if k in un_key_txt or un_key_txt in k:
                    return v

                score_seq = difflib.SequenceMatcher(None, un_key_txt, str(k)).ratio()
                score = score_seq
                if un_tokens:
                    k_tokens = set([t for t in str(k).split() if t])
                    union = max(len(un_tokens | k_tokens), 1)
                    score_jaccard = len(un_tokens & k_tokens) / union
                    score = max(score_seq, score_jaccard)
                if score > best_score:
                    best_score = score
                    best_value = v

            if best_score >= 0.62:
                return best_value

            best_norm_score = 0.0
            best_norm_value = 0
            for n, v in cnt_map_norm.items():
                if not n:
                    continue
                score = difflib.SequenceMatcher(None, un_norm_txt, str(n)).ratio()
                if score > best_norm_score:
                    best_norm_score = score
                    best_norm_value = v
            if best_norm_score >= 0.72:
                return best_norm_value
            return 0

        geo_heat["qtd_atendimentos"] = geo_heat.apply(
            lambda r: _approx_count(r["unidade_key"], r["unidade_norm"]),
            axis=1,
        )

    geo_heat["qtd_atendimentos"] = pd.to_numeric(geo_heat["qtd_atendimentos"], errors="coerce").fillna(0).astype(int)
    heat_df = geo_heat.copy()

    if heat_df.empty or int(heat_df["qtd_atendimentos"].sum()) == 0:
        st.warning("Não foi possível casar unidades do GEO com atendimentos do CELK.")
        section_end()
        st.markdown("</div>", unsafe_allow_html=True)
        return

    lat_med = float(heat_df["latitude"].median())
    lon_med = float(heat_df["longitude"].median())
    lat_tol = max(float(heat_df["latitude"].std()) * 0.25, 0.01)
    lon_tol = max(float(heat_df["longitude"].std()) * 0.25, 0.01)

    def _classifica_regiao(lat, lon):
        dlat = float(lat) - lat_med
        dlon = float(lon) - lon_med
        if abs(dlat) <= lat_tol and abs(dlon) <= lon_tol:
            return "Centro"
        if abs(dlat) >= abs(dlon):
            return "Norte" if dlat > 0 else "Sul"
        return "Leste" if dlon > 0 else "Oeste"

    heat_df["regiao"] = heat_df.apply(lambda r: _classifica_regiao(r["latitude"], r["longitude"]), axis=1)

    q_heat_60 = float(heat_df["qtd_atendimentos"].quantile(0.60))
    q_heat_85 = float(heat_df["qtd_atendimentos"].quantile(0.85))
    q_heat_95 = float(heat_df["qtd_atendimentos"].quantile(0.95))
    regiao_lider = str(heat_df.groupby("regiao")["qtd_atendimentos"].sum().sort_values(ascending=False).index[0])
    with summary_slot:
        summary_cols = st.columns(6)
        with summary_cols[0]:
            territorial_kpi_card(
                "Unidades monitoradas",
                format_int(len(mapa_render_df)),
                subtitle="ativos no perímetro filtrado",
                accent="#38BDF8",
                icon="🏥",
            )
        with summary_cols[1]:
            territorial_kpi_card(
                "Média por unidade",
                f"{mapa_render_df['qtd_colaboradores'].mean():.1f}",
                subtitle="densidade média do efetivo",
                accent="#F59E0B",
                icon="📊",
            )
        with summary_cols[2]:
            territorial_kpi_card(
                "Regiões críticas",
                format_int(unidades_criticas_mapa),
                subtitle="unidades acima do P85",
                accent="#F43F5E",
                icon="🚨",
            )
        with summary_cols[3]:
            territorial_kpi_card(
                "Atendimentos no período",
                format_int(int(heat_df["qtd_atendimentos"].sum())),
                subtitle="volume consolidado CELK",
                accent="#38BDF8",
                icon="🧭",
            )
        with summary_cols[4]:
            territorial_kpi_card(
                "Unidades com demanda",
                format_int(int((heat_df["qtd_atendimentos"] > 0).sum())),
                subtitle="com fluxo registrado",
                accent="#34D399",
                icon="🏥",
            )
        with summary_cols[5]:
            territorial_kpi_card(
                "Região líder",
                regiao_lider,
                subtitle="maior concentração operacional",
                accent="#F59E0B",
                icon="📍",
            )

    plot_df = heat_df.copy()
    if foco_nucleo and len(plot_df) >= 8:
        q_low_lon, q_high_lon = plot_df["longitude"].quantile([0.10, 0.90])
        q_low_lat, q_high_lat = plot_df["latitude"].quantile([0.10, 0.90])
        foco_df = plot_df[
            plot_df["longitude"].between(float(q_low_lon), float(q_high_lon))
            & plot_df["latitude"].between(float(q_low_lat), float(q_high_lat))
        ].copy()
        if len(foco_df) >= 5:
            plot_df = foco_df

    map_center_heat = [
        float(plot_df["latitude"].median()),
        float(plot_df["longitude"].median()),
    ]

    mapa_calor = folium.Map(
        location=map_center_heat,
        zoom_start=initial_zoom_level,
        tiles="CartoDB positron",
        control_scale=True,
    )

    max_heat_val = max(float(plot_df["qtd_atendimentos"].max()), 1.0)
    heat_points = []
    for _, r in plot_df.iterrows():
        qtd = float(r["qtd_atendimentos"])
        if qtd <= 0:
            continue
        # Peso mínimo para manter unidades com menor volume perceptíveis no calor.
        peso = 0.25 + 0.75 * ((qtd / max_heat_val) ** 0.55)
        heat_points.append([float(r["latitude"]), float(r["longitude"]), float(peso)])
    HeatMap(
        data=heat_points,
        radius=radius_heat,
        blur=blur_heat,
        min_opacity=max(0.20, float(opacidade_heat) - 0.15),
        max_zoom=14,
        gradient={0.20: "#1d4ed8", 0.38: "#0ea5e9", 0.55: "#22c55e", 0.72: "#facc15", 0.86: "#f97316", 1.0: "#dc2626"},
        name="Calor de atendimentos",
    ).add_to(mapa_calor)

    if mostrar_halo:
        HeatMap(
            data=heat_points,
            radius=min(80, radius_heat + 14),
            blur=min(46, blur_heat + 12),
            min_opacity=max(0.08, float(opacidade_heat) - 0.45),
            max_zoom=14,
            gradient={0.25: "#60a5fa", 0.65: "#f59e0b", 1.0: "#ef4444"},
            name="Halo de intensidade",
            show=True,
        ).add_to(mapa_calor)

    unidades_layer = folium.FeatureGroup(name="Unidades", show=True)
    for _, row in heat_df.iterrows():
        atend = int(row["qtd_atendimentos"])
        if atend >= q_heat_95:
            cor = "#b91c1c"
            nivel = "Critico"
        elif atend >= q_heat_85:
            cor = "#ea580c"
            nivel = "Alto"
        elif atend >= q_heat_60:
            cor = "#0284c7"
            nivel = "Moderado"
        else:
            cor = "#0f766e"
            nivel = "Base"
        raio_u = max(6, min(15, int(round(6 + (atend / max(float(heat_df["qtd_atendimentos"].max()), 1.0)) * 9))))
        tip_heat = (
            f"<b>{html.escape(str(row['unidade']))}</b><br>"
            f"Lat/Lon: <b>{float(row['latitude']):.5f}, {float(row['longitude']):.5f}</b><br>"
            f"Atendimentos: <b>{atend:,}</b><br>"
            f"Regiao: <b>{html.escape(str(row['regiao']))}</b><br>"
            f"Tipo: <b>{html.escape(str(row['tipo']))}</b><br>"
            f"Classe de criticidade: <b>{nivel}</b>"
        )
        folium.CircleMarker(
            location=[float(row["latitude"]), float(row["longitude"])],
            radius=raio_u + 4,
            color=cor,
            weight=0,
            fill=True,
            fill_color=cor,
            fill_opacity=0.15,
            tooltip=folium.Tooltip(tip_heat, sticky=True),
        ).add_to(unidades_layer)
        folium.CircleMarker(
            location=[float(row["latitude"]), float(row["longitude"])],
            radius=raio_u,
            color="#111827",
            weight=1,
            fill=True,
            fill_color=cor,
            fill_opacity=0.92,
            tooltip=folium.Tooltip(tip_heat, sticky=True),
        ).add_to(unidades_layer)
    unidades_layer.add_to(mapa_calor)

    hotspots = heat_df.sort_values("qtd_atendimentos", ascending=False).head(8)
    hotspot_layer = folium.FeatureGroup(name="Top hotspots", show=True)
    for _, row in hotspots.iterrows():
        folium.Marker(
            location=[float(row["latitude"]), float(row["longitude"])],
            icon=folium.DivIcon(
                html=(
                    "<div style='transform: translate(-50%, -50%);'>"
                    f"<span style='background:#0F172A;color:#F8FAFC;padding:3px 7px;border-radius:999px;"
                    "font-size:11px;font-weight:800;border:1px solid rgba(255,255,255,0.35);box-shadow:0 2px 8px rgba(0,0,0,0.40);'>"
                    f"{int(row['qtd_atendimentos'])}</span></div>"
                )
            ),
            tooltip=f"Hotspot: {row['unidade']} | Atendimentos: {int(row['qtd_atendimentos']):,}",
        ).add_to(hotspot_layer)
    hotspot_layer.add_to(mapa_calor)

    folium.LayerControl(collapsed=False).add_to(mapa_calor)
    add_map_lock_control(mapa_calor, initially_locked=True)

    legenda_html = """
    <div style="position: fixed; bottom: 18px; left: 18px; z-index: 9999;
                                background: rgba(255,255,255,0.96); border: 1px solid #dbe3ef; border-radius: 12px;
                                padding: 10px 12px; font-size: 12px; line-height: 1.35; color: #111827; box-shadow:0 8px 20px rgba(15,23,42,0.12);">
            <div style="font-weight: 800; margin-bottom: 7px; color:#0F172A;">Radar de criticidade</div>
            <div><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#0f766e;margin-right:6px;"></span>Base (&lt; P60)</div>
            <div><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#0284c7;margin-right:6px;"></span>Moderado (P60+)</div>
            <div><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#ea580c;margin-right:6px;"></span>Alto (P85+)</div>
            <div><span style="display:inline-block;width:10px;height:10px;border-radius:50%;background:#b91c1c;margin-right:6px;"></span>Critico (P95+)</div>
    </div>
    """
    mapa_calor.get_root().html.add_child(folium.Element(legenda_html))
    mapa_calor_html = mapa_calor.get_root().render()

    with maps_slot:
        section_start(
            "Centro Inteligente de Gestão Territorial e Operacional",
            "",
            theme="territorial",
        )
        st.markdown(
            """
            <style>
            .stTabs [data-baseweb="tab-list"] {
                gap: 0.6rem;
                padding: 0.2rem 0.1rem 0.5rem 0.1rem;
            }

            .stTabs [data-baseweb="tab"] {
                background: linear-gradient(180deg, rgba(15,108,189,0.12) 0%, rgba(15,108,189,0.05) 100%);
                border: 1px solid rgba(15,108,189,0.35);
                border-radius: 12px;
                padding: 0.5rem 0.95rem;
                min-height: 44px;
                box-shadow: 0 2px 8px rgba(15, 108, 189, 0.08);
            }

            .stTabs [data-baseweb="tab"] p {
                font-size: 1rem;
                font-weight: 800;
                letter-spacing: 0.01em;
                color: #0b3b69;
            }

            .stTabs [aria-selected="true"] {
                background: linear-gradient(180deg, rgba(15,108,189,0.24) 0%, rgba(15,108,189,0.12) 100%);
                border: 2px solid #0F6CBD;
                box-shadow: 0 0 0 1px rgba(15,108,189,0.15), 0 8px 18px rgba(15, 108, 189, 0.18);
                transform: translateY(-1px);
            }

            .stTabs [aria-selected="true"] p {
                color: #083055;
            }
            </style>
            """,
            unsafe_allow_html=True,
        )
        tab_mapa, tab_colab, tab_efetivo, tab_resumo_regional, tab_ranking_atend, tab_aderencia = st.tabs([
            "🗺️ Mapas territoriais",
            "👥 Colaboradores por unidade",
            "🏥 Ranking de efetivo por unidade",
            "🌍 Resumo regional de atendimentos",
            "📊 Ranking de atendimentos",
            "🧩 Análise de aderência",
        ])

        with tab_mapa:
            st.caption("Configurações do mapa de calor")
            st.selectbox("Período", opcoes_mes, index=opcoes_mes.index(mes_heat), key="mt_heat_mes")
            with st.expander("⚙ Configurações do mapa", expanded=False):
                conf_col_1, conf_col_2 = st.columns(2)
                with conf_col_1:
                    st.markdown("<div class='mt-control-title' style='font-size:16px;font-weight:700;color:#334155;'>Configurações Heatmap</div>", unsafe_allow_html=True)
                    st.slider("Raio", min_value=18, max_value=65, value=radius_heat, step=1, key="mt_heat_radius")
                    st.slider("Suavização", min_value=8, max_value=36, value=blur_heat, step=1, key="mt_heat_suavizacao")
                    st.slider("Opacidade", min_value=0.25, max_value=0.95, value=opacidade_heat, step=0.05, key="mt_heat_opacidade")
                with conf_col_2:
                    st.markdown("<div class='mt-control-title' style='font-size:16px;font-weight:700;color:#334155;'>Configurações Territoriais</div>", unsafe_allow_html=True)
                    st.checkbox(
                        "Foco geográfico inteligente",
                        value=foco_nucleo,
                        key="mt_heat_focus_core",
                        help="Reduz influência de outliers para ampliar a leitura do núcleo operacional.",
                    )
                    st.checkbox(
                        "Exibir halo de intensidade",
                        value=mostrar_halo,
                        key="mt_heat_halo",
                        help="Liga/desliga a camada adicional de reforço visual da intensidade de atendimento.",
                    )

            col_map_1, col_map_2 = st.columns(2)
            with col_map_1:
                st.markdown(
                    "<div style='font-size:20px;font-weight:900;color:#0B1220;line-height:1.2;'>Mapa de Colaboradores por Unidade " + territorial_help_badge(
                        "Motivo: visualizar a distribuição espacial das unidades e do efetivo. Para que serve: entender cobertura territorial e apoiar decisões de alocação."
                    ) + "</div>",
                    unsafe_allow_html=True,
                )
                components.html(mapa_geo_html, height=690, scrolling=False)
            with col_map_2:
                st.markdown(
                    "<div style='font-size:20px;font-weight:900;color:#0B1220;line-height:1.2;'>Mapa de Calor Assistencial " + territorial_help_badge(
                        "Motivo: evidenciar concentração de demanda assistencial por região. Para que serve: priorizar áreas com maior pressão de atendimento."
                    ) + "</div>",
                    unsafe_allow_html=True,
                )
                components.html(mapa_calor_html, height=690, scrolling=False)

        with tab_colab:
            st.markdown(
                "<div style='font-size:20px;font-weight:900;color:#0B1220;line-height:1.2;'>Colaboradores por Unidade</div>",
                unsafe_allow_html=True,
            )

            if colab_df is None or colab_df.empty:
                st.info("Sem dados de colaboradores na aba COLABORADORES.")
            else:
                unidades_validas = set(mapa_df["unidade_norm"].dropna().unique().tolist())
                colab_view = colab_df[colab_df["unidade_norm"].isin(unidades_validas)].copy()

                if colab_view.empty:
                    st.info("Nenhum colaborador encontrado para o filtro atual de tipo de unidade.")
                else:
                    unidade_options = sorted(colab_view["unidade"].dropna().astype(str).str.strip().unique().tolist())
                    unidade_sel = st.selectbox(
                        "Selecionar unidade",
                        options=unidade_options,
                        key="mt_colab_unidade_sel",
                    )

                    colab_unidade = (
                        colab_view[colab_view["unidade"] == unidade_sel]
                        .copy()
                        .sort_values(["colaborador", "cargo"], ascending=[True, True])
                    )

                    proventos_cols = [c for c in colab_unidade.columns if str(c).startswith("proventos_")]
                    proventos_cols = sorted(proventos_cols)
                    proventos_label_map = {
                        c: f"Proventos {str(c).replace('proventos_', '').replace('_', ' ').title()}"
                        for c in proventos_cols
                    }

                    proventos_col_sel = None
                    if proventos_cols:
                        proventos_col_sel = st.selectbox(
                            "Mês de proventos",
                            options=proventos_cols,
                            index=len(proventos_cols) - 1,
                            format_func=lambda x: proventos_label_map.get(x, x),
                            key="mt_colab_proventos_mes_sel",
                        )

                    c1, c2, c3, c4, c5 = st.columns(5)
                    with c1:
                        territorial_kpi_card(
                            "Total de colaboradores",
                            format_int(int(colab_unidade["colaborador_id"].nunique())),
                            subtitle="na unidade selecionada",
                            accent="#38BDF8",
                            icon="👥",
                        )
                    with c2:
                        territorial_kpi_card(
                            "Com CRM",
                            format_int(int(colab_unidade["crm"].astype(str).str.strip().ne("").sum())),
                            subtitle="profissionais com registro",
                            accent="#22C55E",
                            icon="🩺",
                        )
                    with c3:
                        ativos_mask = colab_unidade["situacao"].map(
                            lambda v: (normalize_text(v) or "") in {"ATIVO", "ATIVA", "EM ATIVIDADE"}
                        )
                        territorial_kpi_card(
                            "Situação ativa",
                            format_int(int(ativos_mask.sum())),
                            subtitle="colaboradores ativos",
                            accent="#14B8A6",
                            icon="✅",
                        )
                    with c4:
                        territorial_kpi_card(
                            "Cargos distintos",
                            format_int(int(colab_unidade["cargo"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().nunique())),
                            subtitle="diversidade de funções",
                            accent="#F59E0B",
                            icon="🧭",
                        )
                    with c5:
                        total_proventos = 0.0
                        if proventos_col_sel:
                            total_proventos = float(pd.to_numeric(colab_unidade[proventos_col_sel], errors="coerce").fillna(0).sum())
                        territorial_kpi_card(
                            "Proventos",
                            format_currency_br(total_proventos),
                            subtitle=proventos_label_map.get(proventos_col_sel, "mês selecionado") if proventos_col_sel else "sem coluna de proventos",
                            accent="#7C3AED",
                            icon="💰",
                        )

                    tabela_colab = colab_unidade.rename(columns={
                        "colaborador": "Colaborador",
                        "cpf": "CPF",
                        "adm": "Adm",
                        "situacao": "Situação",
                        "cargo": "Cargo",
                        "regime_trabalho": "Regime de trabalho",
                        "crm": "CRM",
                    })

                    tabela_cols = [
                        "Colaborador",
                        "CPF",
                        "Adm",
                        "Situação",
                        "Cargo",
                        "Regime de trabalho",
                        "CRM",
                    ]
                    if proventos_col_sel:
                        proventos_col_nome = proventos_label_map.get(proventos_col_sel, "Proventos")
                        tabela_colab[proventos_col_nome] = pd.to_numeric(
                            tabela_colab.get(proventos_col_sel), errors="coerce"
                        )
                        tabela_cols.append(proventos_col_nome)

                    tabela_colab = tabela_colab[tabela_cols]

                    render_elegant_table(
                        tabela_colab,
                        column_config={
                            "Colaborador": st.column_config.TextColumn("Colaborador"),
                            "CPF": st.column_config.TextColumn("CPF"),
                            "Adm": st.column_config.TextColumn("Adm"),
                            "Situação": st.column_config.TextColumn("Situação"),
                            "Cargo": st.column_config.TextColumn("Cargo"),
                            "Regime de trabalho": st.column_config.TextColumn("Regime de trabalho"),
                            "CRM": st.column_config.TextColumn("CRM"),
                            **(
                                {
                                    proventos_label_map.get(proventos_col_sel, "Proventos"): st.column_config.NumberColumn(
                                        proventos_label_map.get(proventos_col_sel, "Proventos"),
                                        format="R$ %.2f",
                                    )
                                }
                                if proventos_col_sel
                                else {}
                            ),
                        },
                        emphasis_columns=["Colaborador"],
                        key="mt_tbl_colaboradores_unidade",
                    )

        with tab_efetivo:
            section_start("", "", theme="territorial")
            territorial_heading("Ranking de efetivo por unidade")
            ranking_efetivo = ranking_df_filtrado.rename(columns={
                "unidade": "Unidade",
                "tipo": "Tipo",
                "qtd_colaboradores": "Colaboradores",
                "qtd_medicos": "Com CRM",
            })

            c_eff1, c_eff2, c_eff3 = st.columns(3)
            with c_eff1:
                territorial_kpi_card(
                    "Maior efetivo",
                    format_int(int(ranking_df_filtrado["qtd_colaboradores"].max())),
                    subtitle="pico de capacidade por unidade",
                    accent="#2563EB",
                    icon="🏥",
                )
            with c_eff2:
                territorial_kpi_card(
                    "Efetivo mediano",
                    f"{float(ranking_df_filtrado['qtd_colaboradores'].median()):.1f}",
                    subtitle="base típica operacional",
                    accent="#22C55E",
                    icon="📈",
                )
            with c_eff3:
                crm_share = 100.0 * float(ranking_df_filtrado["qtd_medicos"].sum()) / max(float(ranking_df_filtrado["qtd_colaboradores"].sum()), 1.0)
                territorial_kpi_card(
                    "Cobertura CRM",
                    f"{crm_share:.1f}%",
                    subtitle="participação médica no efetivo",
                    accent="#F59E0B",
                    icon="🩺",
                )

            territorial_info_line(
                "Tabela de ranking de efetivo",
                "Motivo: ordenar capacidade por unidade. Para que serve: identificar rapidamente concentração e possíveis desequilíbrios de efetivo."
            )
            render_elegant_table(
                ranking_efetivo,
                column_config={
                    "Unidade": st.column_config.TextColumn("Unidade"),
                    "Tipo": st.column_config.TextColumn("Tipo"),
                    "Colaboradores": st.column_config.NumberColumn("Colaboradores", format="%d"),
                    "Com CRM": st.column_config.NumberColumn("Com CRM", format="%d"),
                },
                emphasis_columns=["Unidade"],
                bar_columns=["Colaboradores", "Com CRM"],
                key="mt_tbl_ranking_efetivo",
            )
            section_end()
        section_end()

    if foco_nucleo:
        pass

    with tab_resumo_regional:
        section_start("", "", theme="territorial")
        territorial_heading("Resumo regional de atendimentos")
        resumo_regional = (
            heat_df.groupby("regiao", as_index=False)
            .agg(
                unidades=("unidade", "count"),
                atendimentos=("qtd_atendimentos", "sum"),
                media_unidade=("qtd_atendimentos", "mean"),
            )
            .sort_values("atendimentos", ascending=False)
        )
        total_at = max(float(resumo_regional["atendimentos"].sum()), 1.0)
        resumo_regional["participacao_pct"] = (resumo_regional["atendimentos"] / total_at * 100).round(1)
        resumo_regional_tbl = resumo_regional.rename(columns={
            "regiao": "Região",
            "unidades": "Unidades",
            "atendimentos": "Atendimentos",
            "media_unidade": "Média por unidade",
            "participacao_pct": "Participação %",
        })
        if not resumo_regional.empty:
            top_reg = resumo_regional.iloc[0]
            c_reg1, c_reg2, c_reg3 = st.columns(3)
            with c_reg1:
                territorial_kpi_card(
                    "Região líder",
                    str(top_reg["regiao"]),
                    subtitle="maior concentração de atendimento",
                    accent="#2563EB",
                    icon="🧭",
                )
            with c_reg2:
                territorial_kpi_card(
                    "Participação líder",
                    f"{float(top_reg['participacao_pct']):.1f}%",
                    subtitle="peso sobre o volume total",
                    accent="#22C55E",
                    icon="📊",
                )
            with c_reg3:
                territorial_kpi_card(
                    "Média regional",
                    f"{float(resumo_regional['media_unidade'].mean()):.1f}",
                    subtitle="atendimentos médios por unidade",
                    accent="#F59E0B",
                    icon="📍",
                )
            territorial_heading(f"Dominância da região líder: {float(top_reg['participacao_pct']):.1f}%")
            st.progress(min(1.0, max(0.0, float(top_reg["participacao_pct"]) / 100.0)))

        territorial_info_line(
            "Tabela de resumo regional",
            "Motivo: consolidar o volume por região. Para que serve: comparar participação regional e orientar foco tático."
        )
        render_elegant_table(
            resumo_regional_tbl,
            column_config={
                "Região": st.column_config.TextColumn("Região"),
                "Unidades": st.column_config.NumberColumn("Unidades", format="%d"),
                "Atendimentos": st.column_config.NumberColumn("Atendimentos", format="%d"),
                "Média por unidade": st.column_config.NumberColumn("Média por unidade", format="%.1f"),
                "Participação %": st.column_config.NumberColumn("Participação %", format="%.1f%%"),
            },
            emphasis_columns=["Região"],
            bar_columns=["Atendimentos", "Média por unidade"],
            progress_columns=["Participação %"],
            heatmap_columns=["Participação %"],
            key="mt_tbl_resumo_regional",
        )
        section_end()

    with tab_ranking_atend:
        section_start("", "", theme="territorial")
        territorial_heading("Ranking de atendimentos")
        ranking_atend = heat_df[["unidade", "tipo", "qtd_atendimentos"]]
        ranking_atend = ranking_atend.sort_values("qtd_atendimentos", ascending=False).rename(columns={
            "unidade": "Unidade",
            "tipo": "Tipo",
            "qtd_atendimentos": "Atendimentos",
        })
        territorial_info_line(
            "Tabela de ranking de atendimentos",
            "Motivo: ranquear unidades por volume de atendimento. Para que serve: acelerar priorização operacional e análise de pressão local."
        )
        render_elegant_table(
            ranking_atend,
            column_config={
                "Unidade": st.column_config.TextColumn("Unidade"),
                "Tipo": st.column_config.TextColumn("Tipo"),
                "Atendimentos": st.column_config.NumberColumn("Atendimentos", format="%d"),
            },
            emphasis_columns=["Unidade"],
            bar_columns=["Atendimentos"],
            heatmap_columns=["Atendimentos"],
            key="mt_tbl_ranking_atendimentos",
        )
        section_end()

    # Cálculos base da análise de aderência (renderização ocorre dentro da aba dedicada).

    efetivo_base = (
        mapa_df[["unidade_norm", "unidade", "tipo", "qtd_colaboradores", "qtd_medicos"]]
        .drop_duplicates(subset=["unidade_norm"])
        .copy()
    )
    aderencia_df = heat_df.merge(
        efetivo_base,
        on="unidade_norm",
        how="left",
        suffixes=("_heat", "_efetivo"),
    )
    aderencia_df["unidade_ref"] = aderencia_df["unidade_efetivo"].fillna(aderencia_df["unidade_heat"])
    aderencia_df["tipo_ref"] = aderencia_df["tipo_efetivo"].fillna(aderencia_df["tipo_heat"])
    aderencia_df["qtd_colaboradores"] = pd.to_numeric(aderencia_df["qtd_colaboradores"], errors="coerce").fillna(0)
    aderencia_df["qtd_medicos"] = pd.to_numeric(aderencia_df["qtd_medicos"], errors="coerce").fillna(0)
    aderencia_df = aderencia_df[(aderencia_df["qtd_atendimentos"] > 0) | (aderencia_df["qtd_colaboradores"] > 0)].copy()

    total_atend_ader = max(float(aderencia_df["qtd_atendimentos"].sum()), 1.0)
    total_colab_ader = max(float(aderencia_df["qtd_colaboradores"].sum()), 1.0)

    aderencia_df["share_atend"] = aderencia_df["qtd_atendimentos"] / total_atend_ader
    aderencia_df["share_colab"] = aderencia_df["qtd_colaboradores"] / total_colab_ader

    def _indice_adequacao(row):
        if float(row["share_atend"]) <= 0:
            return 1.0 if float(row["share_colab"]) <= 0 else 9.99
        return float(row["share_colab"]) / float(row["share_atend"])

    aderencia_df["indice_adequacao"] = aderencia_df.apply(_indice_adequacao, axis=1)
    aderencia_df["colaboradores_ideal"] = (aderencia_df["qtd_atendimentos"] / total_atend_ader) * total_colab_ader
    aderencia_df["gap_colaboradores"] = aderencia_df["colaboradores_ideal"] - aderencia_df["qtd_colaboradores"]

    def _status_adequacao(idx):
        if idx < 0.85:
            return "Deficit de efetivo"
        if idx > 1.15:
            return "Excesso relativo de efetivo"
        return "Equilibrado"

    aderencia_df["status_adequacao"] = aderencia_df["indice_adequacao"].map(_status_adequacao)

    total_unid_ader = max(len(aderencia_df), 1)
    qtd_excesso = int((aderencia_df["status_adequacao"] == "Excesso relativo de efetivo").sum())
    qtd_equil = int((aderencia_df["status_adequacao"] == "Equilibrado").sum())
    qtd_critico = int((aderencia_df["status_adequacao"] == "Deficit de efetivo").sum())

    with tab_aderencia:
        section_start("", "", theme="territorial")
        territorial_heading("Análise de aderência")
        territorial_info_line(
            "KPIs de inteligência territorial",
            "Motivo: comparar demanda assistencial versus distribuição de efetivo. Para que serve: orientar recomposição de equipes."
        )
        ex2, ex3, ex4 = st.columns(3)
        with ex2:
            territorial_kpi_card("Equilibradas", format_int(qtd_equil), subtitle="dentro da faixa ideal", accent="#22C55E", icon="✅")
        with ex3:
            territorial_kpi_card("Excesso", format_int(qtd_excesso), subtitle="capacidade acima da demanda", accent="#F59E0B", icon="⚖️")
        with ex4:
            territorial_kpi_card("Déficit", format_int(qtd_critico), subtitle="necessidade de recomposição", accent="#EF4444", icon="🚨")
        section_end()

    

    med_at = float(aderencia_df["qtd_atendimentos"].median()) if not aderencia_df.empty else 0.0
    med_col = float(aderencia_df["qtd_colaboradores"].median()) if not aderencia_df.empty else 0.0

    def _quadrante(row):
        alta_demanda = float(row["qtd_atendimentos"]) >= med_at
        alto_efetivo = float(row["qtd_colaboradores"]) >= med_col
        if alta_demanda and (not alto_efetivo):
            return "Alta demanda / Baixo efetivo"
        if alta_demanda and alto_efetivo:
            return "Alta demanda / Alto efetivo"
        if (not alta_demanda) and alto_efetivo:
            return "Baixa demanda / Alto efetivo"
        return "Baixa demanda / Baixo efetivo"

    aderencia_df["quadrante"] = aderencia_df.apply(_quadrante, axis=1)

    quadrante_ordem = [
        "Alta demanda / Baixo efetivo",
        "Alta demanda / Alto efetivo",
        "Baixa demanda / Alto efetivo",
        "Baixa demanda / Baixo efetivo",
    ]
    resumo_quadrantes = (
        aderencia_df.groupby("quadrante", as_index=False)
        .agg(
            unidades=("unidade_ref", "count"),
            atendimentos=("qtd_atendimentos", "sum"),
            colaboradores=("qtd_colaboradores", "sum"),
        )
    )
    resumo_quadrantes["quadrante"] = pd.Categorical(
        resumo_quadrantes["quadrante"],
        categories=quadrante_ordem,
        ordered=True,
    )
    resumo_quadrantes = resumo_quadrantes.sort_values("quadrante")

    with tab_aderencia:
        section_start("", "", theme="territorial")
        territorial_heading("Resumo por quadrante")
        resumo_quadrantes_tbl = resumo_quadrantes.rename(columns={
            "quadrante": "Quadrante",
            "unidades": "Unidades",
            "atendimentos": "Atendimentos",
            "colaboradores": "Colaboradores",
        })
        if not resumo_quadrantes.empty:
            quad_crit = resumo_quadrantes[resumo_quadrantes["quadrante"] == "Alta demanda / Baixo efetivo"]
            pct_crit = 0.0
            if not quad_crit.empty:
                pct_crit = 100.0 * float(quad_crit["unidades"].sum()) / max(float(resumo_quadrantes["unidades"].sum()), 1.0)
            qd1, qd2, qd3 = st.columns(3)
            with qd1:
                territorial_kpi_card("Quadrante crítico", f"{pct_crit:.1f}%", subtitle="alta demanda e baixo efetivo", accent="#EF4444", icon="🚨")
            with qd2:
                territorial_kpi_card("Unidades em risco", format_int(int(quad_crit["unidades"].sum()) if not quad_crit.empty else 0), subtitle="prioridade de intervenção", accent="#F59E0B", icon="⚠️")
            with qd3:
                territorial_kpi_card("Quadrantes ativos", format_int(int(len(resumo_quadrantes))), subtitle="cenários operacionais mapeados", accent="#2563EB", icon="🧩")
            territorial_heading(f"Pressão no quadrante crítico: {pct_crit:.1f}% das unidades")
            st.progress(min(1.0, max(0.0, pct_crit / 100.0)))

        territorial_info_line(
            "Tabela de quadrantes",
            "Motivo: classificar unidades por combinação de demanda e efetivo. Para que serve: orientar estratégias diferentes para cada cenário operacional."
        )
        render_elegant_table(
            resumo_quadrantes_tbl,
            column_config={
                "Quadrante": st.column_config.TextColumn("Quadrante"),
                "Unidades": st.column_config.NumberColumn("Unidades", format="%d"),
                "Atendimentos": st.column_config.NumberColumn("Atendimentos", format="%d"),
                "Colaboradores": st.column_config.NumberColumn("Colaboradores", format="%d"),
            },
            emphasis_columns=["Quadrante"],
            bar_columns=["Atendimentos", "Colaboradores"],
            heatmap_columns=["Atendimentos"],
            key="mt_tbl_resumo_quadrantes",
        )
        section_end()

    with tab_mapa:
        section_start("", "", theme="territorial")
        territorial_heading("Ranking de priorização")

        gap_pos = aderencia_df.loc[aderencia_df["gap_colaboradores"] > 0, "gap_colaboradores"]
        gap_threshold_crit = float(gap_pos.quantile(0.80)) if not gap_pos.empty else 0.0
        gap_threshold_med = float(gap_pos.quantile(0.45)) if not gap_pos.empty else 0.0

        def _prioridade_operacional(gap):
            g = float(gap)
            if g >= gap_threshold_crit and g > 0:
                return "Critica"
            if g >= gap_threshold_med and g > 0:
                return "Moderada"
            return "Baixa"

        aderencia_df["prioridade_operacional"] = aderencia_df["gap_colaboradores"].map(_prioridade_operacional)

        cprio_1, cprio_2, cprio_3 = st.columns(3)
        with cprio_1:
            territorial_kpi_card(
                "Backlog de GAP",
                f"{float(aderencia_df['gap_colaboradores'].clip(lower=0).sum()):.1f}",
                subtitle="colaboradores a recompor",
                accent="#F97316",
                icon="📉",
            )
        with cprio_2:
            territorial_kpi_card(
                "Prioridade critica",
                format_int(int((aderencia_df["prioridade_operacional"] == "Critica").sum())),
                subtitle="unidades com maior necessidade",
                accent="#EF4444",
                icon="🚨",
            )
        with cprio_3:
            territorial_kpi_card(
                "Indice mediano",
                f"{float(aderencia_df['indice_adequacao'].median()):.2f}",
                subtitle="adequacao demanda x efetivo",
                accent="#38BDF8",
                icon="📊",
            )

        ranking_gap = (
            aderencia_df[[
                "unidade_ref",
                "tipo_ref",
                "regiao",
                "qtd_atendimentos",
                "qtd_colaboradores",
                "colaboradores_ideal",
                "gap_colaboradores",
                "indice_adequacao",
                "status_adequacao",
                "prioridade_operacional",
                "quadrante",
            ]]
            .sort_values("gap_colaboradores", ascending=False)
            .rename(columns={
                "unidade_ref": "Unidade",
                "tipo_ref": "Tipo",
                "regiao": "Regiao",
                "qtd_atendimentos": "Atendimentos",
                "qtd_colaboradores": "Colaboradores",
                "colaboradores_ideal": "Colaboradores ideais",
                "gap_colaboradores": "Gap colaboradores",
                "indice_adequacao": "Indice adequacao",
                "status_adequacao": "Status",
                "prioridade_operacional": "Prioridade",
                "quadrante": "Quadrante",
            })
        )
        territorial_info_line(
            "Tabela de priorização",
            "Motivo: mostrar o gap de colaboradores por unidade. Para que serve: definir sequência de intervenção e redistribuição de equipe."
        )
        render_elegant_table(
            ranking_gap,
            column_config={
                "Unidade": st.column_config.TextColumn("Unidade"),
                "Tipo": st.column_config.TextColumn("Tipo"),
                "Regiao": st.column_config.TextColumn("Região"),
                "Atendimentos": st.column_config.NumberColumn("Atendimentos", format="%d"),
                "Colaboradores": st.column_config.NumberColumn("Colaboradores", format="%d"),
                "Colaboradores ideais": st.column_config.NumberColumn("Colaboradores ideais", format="%.1f"),
                "Gap colaboradores": st.column_config.NumberColumn("Gap colaboradores", format="%.1f"),
                "Indice adequacao": st.column_config.NumberColumn("Índice adequação", format="%.2f"),
                "Status": st.column_config.TextColumn("Status"),
                "Prioridade": st.column_config.TextColumn("Prioridade operacional"),
                "Quadrante": st.column_config.TextColumn("Quadrante"),
            },
            status_columns=["Status", "Prioridade"],
            emphasis_columns=["Unidade", "Gap colaboradores"],
            bar_columns=["Gap colaboradores", "Atendimentos"],
            heatmap_columns=["Indice adequacao", "Gap colaboradores"],
            progress_columns=["Indice adequacao"],
            critical_condition=lambda row: str(row.get("Status", "")).lower().startswith("deficit") or str(row.get("Prioridade", "")).lower().startswith("critica"),
            key="mt_tbl_ranking_priorizacao",
        )

        section_end()

    st.markdown("</div>", unsafe_allow_html=True)

st.sidebar.markdown(
    f"""
    <style>
    section[data-testid="stSidebar"] {{
        min-width: 260px !important;
        max-width: 260px !important;
    }}
    section[data-testid="stSidebar"] > div:first-child {{
        padding-top: 0.65rem;
    }}
    section[data-testid="stSidebar"] div.stButton > button {{
        justify-content: flex-start;
        border-radius: 13px !important;
        font-size: 15px;
        font-weight: 500;
        letter-spacing: 0.1px;
        padding: 8px 11px !important;
        line-height: 1.25;
        border: 1px solid transparent !important;
        margin-bottom: 3px;
        min-height: 42px !important;
    }}
    section[data-testid="stSidebar"] button[id*="menu_unidades_"] {{
        min-height: 47px !important;
        padding-top: 10px !important;
        padding-bottom: 10px !important;
        font-weight: 600;
    }}
    section[data-testid="stSidebar"] button[id*="menu_basicas_"],
    section[data-testid="stSidebar"] button[id*="menu_administrativo_"] {{
        min-height: 40px !important;
        padding-top: 7px !important;
        padding-bottom: 7px !important;
    }}
    section[data-testid="stSidebar"] div.stButton > button[kind="secondary"] {{
        background: transparent !important;
        color: #334155 !important;
    }}
    section[data-testid="stSidebar"] div.stButton > button[kind="secondary"]:hover {{
        background: #F1F5F9 !important;
        color: #0F172A !important;
        border: 1px solid #E2E8F0 !important;
    }}
    section[data-testid="stSidebar"] div.stButton > button[kind="primary"] {{
        background: linear-gradient(90deg, #D2F1E1 0%, #B6E6CC 100%) !important;
        color: #055E45 !important;
        border: 1px solid #63D39B !important;
        box-shadow: none !important;
        font-weight: 700;
    }}
    .sidebar-brand {{
        display: flex;
        align-items: center;
        gap: 10px;
        padding: 8px 2px 11px 2px;
        border-bottom: 1px solid rgba(148, 163, 184, 0.20);
        margin-bottom: 7px;
    }}
    .sidebar-brand-logo {{
        width: 72px;
        height: 72px;
        border-radius: 0;
        object-fit: contain;
        display: block;
    }}
    .sidebar-brand-title {{
        font-size: 18px;
        font-weight: 800;
        line-height: 1.1;
        color: #F8FAFC;
    }}
    .sidebar-brand-sub {{
        font-size: 10px;
        color: rgba(226,232,240,0.75);
        margin-top: 1px;
    }}
    .sidebar-group-label {{
        font-size: 18px;
        letter-spacing: 1.5px;
        text-transform: uppercase;
        font-weight: 900;
        color: #94A3B8;
        margin: 11px 0 5px 0;
    }}
    .sidebar-footer-card {{
        margin-top: 10px;
        background: linear-gradient(135deg, #0E7A5D 0%, #065F46 100%);
        border-radius: 10px;
        border: 1px solid rgba(255,255,255,0.12);
        padding: 10px 12px;
    }}
    .sidebar-footer-card .footer-title {{
        color: #CFFAFE;
        font-size: 11px;
        font-weight: 600;
        margin-bottom: 2px;
    }}
    .sidebar-footer-card .footer-source {{
        color: #FFFFFF;
        font-size: 15px;
        font-weight: 700;
    }}
    </style>
    <div class="sidebar-brand">
        <img class="sidebar-brand-logo" src="data:image/png;base64,{LOGO_SIDEBAR_BASE64}" alt="Patris" />
        <div>
            <div class="sidebar-brand-title">Patris</div>
            <div class="sidebar-brand-sub">Gestão Municipal · Luziânia</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

usuario_logado = st.session_state.get("usuario_logado")

theme_by_user = {
    "admin": "Healthcare Clean (Verde)",
    "vittor": "Healthcare Clean (Verde)",
    "wendel": "Healthcare Clean (Verde)",
    "guilherme": "Healthcare Clean (Verde)",
    "denis": "Healthcare Clean (Verde)",
    "prefeitura": "Healthcare Clean (Verde)",
}

default_theme_for_user = theme_by_user.get(usuario_logado, "Portal Clínico (Azul)")

if "visual_theme" not in st.session_state:
    st.session_state["visual_theme"] = default_theme_for_user

if st.session_state.get("visual_theme_user") != usuario_logado:
    st.session_state["visual_theme"] = default_theme_for_user
    st.session_state["visual_theme_user"] = usuario_logado

visual_theme = st.session_state.get("visual_theme", default_theme_for_user)
apply_visual_theme(visual_theme)

with st.expander("🎨 Temas", expanded=False):
    st.markdown(
        """
        <style>
        .stExpander > div > div {
            padding: 0.2rem 0.4rem !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    theme_col1, theme_col2, theme_col3 = st.columns(3)
    
    if theme_col1.button("Portal Clínico", width="stretch", key="btn_portal"):
        st.session_state["visual_theme"] = "Portal Clínico (Azul)"
    if theme_col2.button("Pro Analytics", width="stretch", key="btn_pro"):
        st.session_state["visual_theme"] = "Pro Analytics (Escuro)"
    if theme_col3.button("Healthcare Clean", width="stretch", key="btn_healthcare"):
        st.session_state["visual_theme"] = "Healthcare Clean (Verde)"
    
    if st.session_state["visual_theme"] != visual_theme:
        apply_visual_theme(st.session_state["visual_theme"])

paginas_unidades = [
    "UPA Luziânia",
    "UPA Jardim Ingá",
    "SAMU",
    "HMJI",
    PAGINA_MAPA_TERRITORIAL,
    PAGINA_HEATMAP,
]

paginas_basicas = [
    "Atenção Primária",
    "Atenção Secundária",
    "Saúde Mental",
]

paginas_administrativo = [
    "Metas do Plano",
    "Gestão de Pessoas",
    "Financeiro",
    PAGINA_ADMIN_ACESSOS,
    PAGINA_PRODUTIVIDADE,
]

todas_paginas = paginas_unidades + paginas_basicas + paginas_administrativo
paginas_disponiveis = todas_paginas

pagina_icons = {
    "UPA Luziânia": "🚑",
    "UPA Jardim Ingá": "🚑",
    "SAMU": "🚨",
    "HMJI": "🏥",
    PAGINA_MAPA_TERRITORIAL: "🗺️",
    PAGINA_HEATMAP: "🔥",
    "Atenção Secundária": "🩺",
    "Saúde Mental": "🧠",
    "Atenção Primária": "💊",
    "Gestão de Pessoas": "👥",
    "Metas do Plano": "📊",
    "Financeiro": "💰",
    PAGINA_ADMIN_ACESSOS: "🔐",
    PAGINA_PRODUTIVIDADE: "📊",
}

if "pagina_selecionada" not in st.session_state or st.session_state["pagina_selecionada"] not in paginas_disponiveis:
    st.session_state["pagina_selecionada"] = paginas_disponiveis[0]

# Compatibilidade: converte rótulo novo para chave interna estável.
if st.session_state.get("pagina_selecionada") == ROTULO_PRODUTIVIDADE:
    st.session_state["pagina_selecionada"] = PAGINA_PRODUTIVIDADE
st.sidebar.markdown('<div class="sidebar-group-label">Unidades</div>', unsafe_allow_html=True)
for page in paginas_unidades:
    if page not in paginas_disponiveis:
        continue
    active = st.session_state["pagina_selecionada"] == page
    page_label = "UPA DE LUZIÂNIA - UPA II" if page == "UPA Luziânia" else page
    if st.sidebar.button(
        f"{pagina_icons.get(page, '📌')}  {page_label}",
        key=f"menu_unidades_{normalize_text(page)}",
        width="stretch",
        type="primary" if active else "secondary"
    ):
        st.session_state["pagina_selecionada"] = page

st.sidebar.markdown('<div class="sidebar-group-label">Unidades basicas</div>', unsafe_allow_html=True)
for page in paginas_basicas:
    if page not in paginas_disponiveis:
        continue
    active = st.session_state["pagina_selecionada"] == page
    if st.sidebar.button(
        f"{pagina_icons.get(page, '📌')}  {page}",
        key=f"menu_basicas_{normalize_text(page)}",
        width="stretch",
        type="primary" if active else "secondary"
    ):
        st.session_state["pagina_selecionada"] = page

st.sidebar.markdown('<div class="sidebar-group-label">Administrativo</div>', unsafe_allow_html=True)
for page in paginas_administrativo:
    if page == PAGINA_PRODUTIVIDADE:
        continue
    if page not in paginas_disponiveis:
        continue
    active = st.session_state["pagina_selecionada"] == page
    page_norm = normalize_text(str(page))
    page_label = ROTULO_PRODUTIVIDADE if "produtividade" in page_norm and "upa" in page_norm else page
    if st.sidebar.button(
        f"{pagina_icons.get(page, '📌')}  {page_label}",
        key=f"menu_administrativo_{normalize_text(page)}",
        width="stretch",
        type="primary" if active else "secondary"
    ):
        st.session_state["pagina_selecionada"] = page

# Botao explicito de produtividade para garantir rotulo final no menu.
active_prod_fixo = st.session_state["pagina_selecionada"] == PAGINA_PRODUTIVIDADE
if st.sidebar.button(
    f"{pagina_icons.get(PAGINA_PRODUTIVIDADE, '📊')}  {ROTULO_PRODUTIVIDADE}",
    key="menu_administrativo_produtividade_medica_upas",
    width="stretch",
    type="primary" if active_prod_fixo else "secondary"
):
    st.session_state["pagina_selecionada"] = PAGINA_PRODUTIVIDADE

pagina = st.session_state["pagina_selecionada"]

if st.session_state.get("last_audit_page") != pagina or st.session_state.get("last_audit_user") != usuario_logado:
    append_audit_event(
        event="page_access",
        user=usuario_logado,
        page=pagina,
        session_id=st.session_state.get("session_id", ""),
        details="Acesso de pagina no painel",
    )
    st.session_state["last_audit_page"] = pagina
    st.session_state["last_audit_user"] = usuario_logado

default_periodo = default_previous_month_selection()
st.sidebar.markdown("## Filtros")
if "meses_selecionados" not in st.session_state:
    st.session_state["meses_selecionados"] = default_periodo

meses_selecionados = st.sidebar.multiselect(
    "Período",
    [MESES_LABEL[m] for m in MESES],
    key="meses_selecionados"
)

st.sidebar.markdown("### Atualizar base")
upload_col1, upload_col2, upload_col3 = st.sidebar.columns([1, 1, 1])

with upload_col1:
    abrir_upload = st.button("📁 Atualizar", width="stretch", key="footer_upload_open")

with upload_col2:
    limpar_upload = st.button("✖", width="stretch", key="footer_upload_clear")

with upload_col3:
    sincronizar = st.button("🔄 Sync", width="stretch", key="footer_sync_local")

if limpar_upload:
    st.session_state.pop("uploaded_file", None)
    st.rerun()

if sincronizar:
    st.session_state.pop("uploaded_file", None)
    load_workbook_data.clear()
    load_metas_data.clear()
    load_financeiro_data.clear()
    load_metas_total_geral_map.clear()
    load_samu_data.clear()
    load_produtividade_data.clear()
    load_mapa_territorial_data.clear()
    load_colaboradores_sheet.clear()
    st.rerun()

uploaded = None
if abrir_upload:
    uploaded = st.sidebar.file_uploader(
        "Selecionar arquivo",
        type=["xlsx"],
        key="upload_hidden"
    )
else:
    uploaded = st.session_state.get("uploaded_file", None)

if uploaded is not None:
    st.session_state["uploaded_file"] = uploaded

if "uploaded_file" in st.session_state:
    st.sidebar.caption("Base atualizada")
else:
    st.sidebar.caption("Usando base local")

file_bytes = uploaded.getvalue() if uploaded else None
_mtime = _local_file_mtime()
data, source_name = load_workbook_data(file_bytes) if uploaded else load_workbook_data(None, _mtime=_mtime)
metas_data = load_metas_data(file_bytes) if uploaded else load_metas_data(None, _mtime=_mtime)
financeiro_data = load_financeiro_data(file_bytes) if uploaded else load_financeiro_data(None, _mtime=_mtime)
metas_total_geral_map = load_metas_total_geral_map(file_bytes) if uploaded else load_metas_total_geral_map(None, _mtime=_mtime)

if data.empty:
    base = Path(__file__).parent
    encontrados = sorted([x.name for x in base.glob("*.xlsx")]) + sorted([x.name for x in base.glob("*.xlsm")])
    st.warning("Não encontrei uma planilha válida automaticamente. Envie um arquivo .xlsx na barra lateral ou deixe o Excel na mesma pasta do app.")
    if encontrados:
        st.info("Arquivos Excel encontrados na pasta do app: " + ", ".join(encontrados))
    else:
        st.info("Nenhum arquivo Excel foi encontrado na mesma pasta do app.")
    st.stop()

st.sidebar.markdown(
    f"""
    <div class="sidebar-footer-card">
        <div class="footer-title">Fonte:</div>
        <div class="footer-source">{source_name}</div>
    </div>
    """,
    unsafe_allow_html=True,
)

if "mes_label" in metas_data.columns:
    metas_data = metas_data[metas_data["mes_label"].isin(meses_selecionados)].copy()
else:
    metas_data = pd.DataFrame(columns=["indicador", "indicador_norm", "mes", "mes_label", "valor"])


def render_admin_access_page():
    if st.session_state.get("usuario_logado") != "admin":
        st.error("Somente o admin pode acessar esta pagina.")
        return

    st.subheader("Administracao de Logins e Senhas")
    store_summary = read_auth_store_summary()
    st.caption(f"Persistencia local ativa em: {store_summary.get('store_path', 'indisponivel')}")

    store_is_deploy_persistent = bool(store_summary.get("store_is_deploy_persistent", False))
    allow_ephemeral = str(os.getenv("AUTH_ALLOW_EPHEMERAL_STORE", "")).strip() == "1"
    is_local_windows = os.name == "nt"
    allow_auth_writes = store_is_deploy_persistent or is_local_windows or allow_ephemeral

    if not allow_auth_writes:
        st.error(
            "Cadastro/edicao de usuarios desativado: armazenamento nao persistente para deploy. "
            "Configure auth.store_dir ou AUTH_STORE_DIR para /mount/data/bi-municipio "
            "(ou habilite volume persistente no Streamlit Cloud)."
        )
        st.caption("Override tecnico temporario: AUTH_ALLOW_EPHEMERAL_STORE=1")

    tab_gestao, tab_auditoria = st.tabs([
        "👤 Gestao de usuarios",
        "🕵️ Auditoria de logins",
    ])

    with tab_gestao:
        usuarios_app = get_usuarios_app()
        permissoes = get_permissoes()
        usuarios_ordenados = sorted(usuarios_app.keys(), key=lambda x: str(x).lower())
        table_rows = []
        for username in usuarios_ordenados:
            perms = permissoes.get(username, [])
            table_rows.append(
                {
                    "usuario": username,
                    "origem": "local" if username in store_summary.get("users", {}) else "secrets",
                    "perfil": "admin" if "*" in perms else "padrao",
                    "permissoes": "*" if "*" in perms else ", ".join(perms),
                }
            )
        if table_rows:
            st.dataframe(pd.DataFrame(table_rows), width="stretch", hide_index=True)

        paginas_opcoes = [
            "UPA Luziânia",
            "UPA Jardim Ingá",
            "SAMU",
            "HMJI",
            PAGINA_MAPA_TERRITORIAL,
            "Atenção Primária",
            "Atenção Secundária",
            "Saúde Mental",
            "Gestão de Pessoas",
            "Financeiro",
            "Metas do Plano",
            PAGINA_PRODUTIVIDADE,
        ]

        st.markdown("### Criar novo usuario")
        c1, c2 = st.columns(2)
        novo_usuario = c1.text_input("Usuario novo", key="adm_new_username")
        senha_nova = c1.text_input("Senha inicial", type="password", key="adm_new_password")
        senha_nova_conf = c2.text_input("Confirmar senha inicial", type="password", key="adm_new_password_confirm")
        novo_admin_total = c2.checkbox("Conceder perfil admin (*)", key="adm_new_full_access")
        novo_permissoes = st.multiselect(
            "Permissoes iniciais",
            paginas_opcoes,
            default=[PAGINA_PRODUTIVIDADE, "SAMU"],
            key="adm_new_permissions",
            disabled=novo_admin_total,
        )
        if st.button("Criar usuario", key="adm_create_user", width="stretch", disabled=(not allow_auth_writes)):
            usuario_norm = novo_usuario.strip()
            if not re.fullmatch(r"[A-Za-z0-9_.-]{3,32}", usuario_norm):
                st.error("Usuario invalido. Use 3-32 caracteres: letras, numeros, _, . ou -")
            elif usuario_norm in USUARIOS_APP:
                st.error("Este usuario ja existe.")
            elif len(senha_nova) < 6:
                st.error("Senha muito curta. Use ao menos 6 caracteres.")
            elif senha_nova != senha_nova_conf:
                st.error("A confirmacao da senha nao confere.")
            else:
                ok_pwd = set_user_password(usuario_norm, senha_nova)
                perms_to_save = ["*"] if novo_admin_total else novo_permissoes
                ok_perm = set_user_permissions(usuario_norm, perms_to_save)
                if ok_pwd and ok_perm:
                    append_audit_event(
                        event="auth_user_create",
                        user=st.session_state.get("usuario_logado", ""),
                        page=PAGINA_ADMIN_ACESSOS,
                        session_id=st.session_state.get("session_id", ""),
                        details=f"Usuario criado: {usuario_norm}",
                    )
                    st.success("Usuario criado com persistencia local.")
                    st.rerun()
                else:
                    st.error("Falha ao gravar usuario. Verifique permissao de escrita em disco.")

        st.markdown("### Alterar senha")
        alvo_senha = st.selectbox("Usuario", usuarios_ordenados, key="adm_password_user") if usuarios_ordenados else None
        n1, n2 = st.columns(2)
        senha_alt = n1.text_input("Nova senha", type="password", key="adm_change_password")
        senha_alt_conf = n2.text_input("Confirmar nova senha", type="password", key="adm_change_password_confirm")
        if st.button("Salvar nova senha", key="adm_change_password_btn", width="stretch", disabled=(not alvo_senha) or (not allow_auth_writes)):
            if len(senha_alt) < 6:
                st.error("Senha muito curta. Use ao menos 6 caracteres.")
            elif senha_alt != senha_alt_conf:
                st.error("A confirmacao da senha nao confere.")
            elif set_user_password(alvo_senha, senha_alt):
                append_audit_event(
                    event="auth_password_change",
                    user=st.session_state.get("usuario_logado", ""),
                    page=PAGINA_ADMIN_ACESSOS,
                    session_id=st.session_state.get("session_id", ""),
                    details=f"Senha alterada para: {alvo_senha}",
                )
                st.success("Senha atualizada e salva em disco.")
                st.rerun()
            else:
                st.error("Falha ao salvar nova senha.")

        st.markdown("### Ajustar permissoes")
        alvo_perm = st.selectbox("Usuario para permissao", usuarios_ordenados, key="adm_perm_user") if usuarios_ordenados else None
        perms_atual = PERMISSOES.get(alvo_perm, []) if alvo_perm else []
        admin_total = st.checkbox("Perfil admin (*)", value=("*" in perms_atual), key="adm_perm_admin")
        selected_perms = st.multiselect(
            "Paginas permitidas",
            paginas_opcoes,
            default=[] if "*" in perms_atual else [p for p in perms_atual if p in paginas_opcoes],
            key="adm_perm_pages",
            disabled=admin_total,
        )
        if st.button("Salvar permissoes", key="adm_perm_save", width="stretch", disabled=(not alvo_perm) or (not allow_auth_writes)):
            perms_to_save = ["*"] if admin_total else selected_perms
            if set_user_permissions(alvo_perm, perms_to_save):
                append_audit_event(
                    event="auth_permissions_change",
                    user=st.session_state.get("usuario_logado", ""),
                    page=PAGINA_ADMIN_ACESSOS,
                    session_id=st.session_state.get("session_id", ""),
                    details=f"Permissoes alteradas para: {alvo_perm}",
                )
                st.success("Permissoes atualizadas e salvas em disco.")
                st.rerun()
            else:
                st.error("Falha ao salvar permissoes.")

        st.markdown("### Desativar usuario")
        candidatos_remocao = [u for u in usuarios_ordenados if u != "admin"]
        remover_usuario = st.selectbox("Usuario para desativar", candidatos_remocao, key="adm_remove_user") if candidatos_remocao else None
        confirma_remocao = st.checkbox("Confirmo a desativacao deste usuario", key="adm_remove_confirm")
        if st.button("Desativar usuario", key="adm_remove_btn", width="stretch", disabled=(not remover_usuario) or (not allow_auth_writes)):
            if not confirma_remocao:
                st.error("Confirme a desativacao para continuar.")
            elif disable_user(remover_usuario):
                append_audit_event(
                    event="auth_user_disable",
                    user=st.session_state.get("usuario_logado", ""),
                    page=PAGINA_ADMIN_ACESSOS,
                    session_id=st.session_state.get("session_id", ""),
                    details=f"Usuario desativado: {remover_usuario}",
                )
                st.success("Usuario desativado com persistencia local.")
                st.rerun()
            else:
                st.error("Falha ao desativar usuario.")

    with tab_auditoria:
        raw_events = read_audit_events(limit=5000)
        login_events = [e for e in raw_events if str(e.get("event", "")).strip().lower() == "login_success"]

        if not login_events:
            st.info("Nenhum registro de login encontrado na auditoria.")
        else:
            audit_df = pd.DataFrame(login_events)
            if "timestamp" in audit_df.columns:
                ts = pd.to_datetime(audit_df["timestamp"], errors="coerce")
                audit_df["timestamp"] = ts.dt.strftime("%d/%m/%Y %H:%M:%S").fillna(audit_df["timestamp"])

            audit_df["user"] = audit_df.get("user", "").astype(str).replace("", "(sem usuario)")

            total_logins = len(audit_df)
            usuarios_distintos = int(audit_df["user"].nunique())
            c1, c2 = st.columns(2)
            c1.metric("Total de logins", f"{total_logins}")
            c2.metric("Usuarios distintos", f"{usuarios_distintos}")

            usuarios_opts = ["Todos"] + sorted(audit_df["user"].unique().tolist())
            usuario_filtro = st.selectbox("Filtrar por usuario", usuarios_opts, key="adm_audit_login_user")

            view_df = audit_df.copy()
            if usuario_filtro != "Todos":
                view_df = view_df[view_df["user"] == usuario_filtro].copy()

            cols = [c for c in ["timestamp", "user", "page", "session_id", "details"] if c in view_df.columns]
            if cols:
                rename_map = {
                    "timestamp": "Data/Hora",
                    "user": "Usuario",
                    "page": "Pagina",
                    "session_id": "Sessao",
                    "details": "Detalhes",
                }
                view_show = view_df[cols].rename(columns=rename_map)
                st.dataframe(view_show, width="stretch", hide_index=True)

                csv_data = view_show.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "Baixar auditoria de logins (CSV)",
                    data=csv_data,
                    file_name="auditoria_logins.csv",
                    mime="text/csv",
                    key="adm_audit_login_download",
                )
            else:
                st.info("Nao foi possivel montar a visualizacao da auditoria.")

hero_header(pagina, source_name, meses_selecionados)

if not usuario_pode_ver_pagina(usuario_logado, pagina):
    st.error("🚫 Você não tem acesso a esta página.")
    st.stop()

if pagina == "UPA Luziânia":
    render_upa_page(data, "UPA DE LUZIÂNIA - UPA II", meses_selecionados)

elif pagina == "UPA Jardim Ingá":
    render_upa_page(data, "UPA JARDIM INGÁ - UPA I", meses_selecionados)

elif pagina == "SAMU":
    render_samu_page()

elif pagina == "HMJI":
    render_hmji(data, meses_selecionados)

elif pagina == "Atenção Secundária":
    render_atencao_secundaria_tabs(data)

elif pagina == "Saúde Mental":
    render_saude_mental_tabs(data)

elif pagina == "Atenção Primária":
    render_atencao_primaria_tabs(data)

elif pagina == "Gestão de Pessoas":
    render_rh_page(data, meses_selecionados, file_bytes=file_bytes, _mtime=_mtime)

elif pagina == "Financeiro":
    render_financeiro_page(financeiro_data, meses_selecionados)

elif pagina == PAGINA_ADMIN_ACESSOS:
    render_admin_access_page()

elif pagina in [PAGINA_PRODUTIVIDADE, ROTULO_PRODUTIVIDADE]:
    render_produtividade_medica_page()

elif pagina == PAGINA_MAPA_TERRITORIAL:
    render_mapa_territorial_page(file_bytes=file_bytes, _mtime=_mtime)

elif pagina == PAGINA_HEATMAP:
    render_heatmap_page()

else:
    render_metas_page(data, metas_data, metas_total_geral_map, meses_selecionados)

with st.expander("Base transformada"):
    if st.checkbox("Mostrar tabela (primeiras 300 linhas)", key="show_base_transformada_table"):
        st.table(data.head(300).reset_index(drop=True))
    else:
        st.caption("Tabela oculta por padrão para reduzir erros de carregamento no navegador.")

st.divider()
footer_col1, footer_col2, footer_col3 = st.columns(3)
footer_col1.caption(f"🔵 VERSAO ATIVA | {BUILD_TAG}")
footer_col2.caption(f"📅 Build local: {globals().get('LOCAL_BUILD_STAMP', 'indisponivel')}")
footer_col3.caption(f"👤 Usuario logado: {usuario_logado}")
