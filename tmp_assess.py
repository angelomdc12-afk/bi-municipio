import json
from openpyxl import load_workbook
import pandas as pd

cand = r"C:\Users\Inovar Soluções\Downloads\PRODUÇÃO CONSOLIDADA DE MARÇO 2026 SISTEMA CELK (1).xlsx"
base = r"C:\Users\Inovar Soluções\Documents\PowerBI\bi-municipio\PowerBI\urgencia_tratado_validado.xlsx"
out_path = r"C:\Users\Inovar Soluções\Documents\PowerBI\bi-municipio\heatmap_file_assessment.json"

report = {"candidate": cand, "base": base, "candidate_exists": False, "base_exists": False}

import os
report["candidate_exists"] = os.path.exists(cand)
report["base_exists"] = os.path.exists(base)

if report["candidate_exists"]:
    wb = load_workbook(cand, read_only=True, data_only=True)
    sheets = []
    for sh in wb.sheetnames:
        ws = wb[sh]
        headers = []
        try:
            first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = list(first)
        except Exception:
            headers = []
        sheets.append({"sheet": sh, "rows": int(ws.max_row or 0), "cols": [str(x) if x is not None else None for x in headers[:20]]})
    report["candidate_sheets"] = sheets

if report["base_exists"]:
    xls = pd.ExcelFile(base)
    base_sheets = []
    for sh in xls.sheet_names:
        d = xls.parse(sh, nrows=2)
        base_sheets.append({"sheet": sh, "rows": int(len(xls.parse(sh))), "cols": [str(c) for c in d.columns[:20]]})
    report["base_sheets"] = base_sheets

# Tentativa de cruzamento diário por unidade
cross = {"status": "not_possible", "reason": "colunas-chave não identificadas"}
if report.get("candidate_sheets"):
    # pega primeira aba com mais linhas
    top = sorted(report["candidate_sheets"], key=lambda x: x["rows"], reverse=True)[0]
    sh = top["sheet"]
    df = pd.read_excel(cand, sheet_name=sh)
    cols_norm = {c: str(c).strip().lower() for c in df.columns}

    date_col = None
    hour_col = None
    unit_col = None
    qty_col = None

    for c,n in cols_norm.items():
        if date_col is None and any(k in n for k in ["data", "date"]):
            date_col = c
        if hour_col is None and any(k in n for k in ["hora", "hour", "atendimento em"]):
            hour_col = c
        if unit_col is None and any(k in n for k in ["unidade", "setor", "origem", "estabelecimento"]):
            unit_col = c
        if qty_col is None and any(k in n for k in ["quant", "qtd", "atendimento", "total", "volume"]):
            qty_col = c

    cross.update({
        "candidate_top_sheet": sh,
        "candidate_columns": [str(c) for c in df.columns[:30]],
        "detected": {
            "date_col": str(date_col) if date_col else None,
            "hour_col": str(hour_col) if hour_col else None,
            "unit_col": str(unit_col) if unit_col else None,
            "qty_col": str(qty_col) if qty_col else None,
        }
    })

    if date_col and unit_col and qty_col:
        dfx = df[[date_col, unit_col, qty_col]].copy()
        dfx[date_col] = pd.to_datetime(dfx[date_col], errors="coerce", dayfirst=True)
        dfx[qty_col] = pd.to_numeric(dfx[qty_col], errors="coerce").fillna(0)
        dfx = dfx.dropna(subset=[date_col])
        dfx["dia"] = dfx[date_col].dt.date
        agg_new = dfx.groupby(["dia", unit_col], dropna=False)[qty_col].sum().reset_index()

        base_daily = pd.read_excel(base, sheet_name="KPI_DIARIO_GERAL")
        base_daily["Data"] = pd.to_datetime(base_daily["Data"], errors="coerce", dayfirst=True)
        rows = []
        for unit_col_base in ["UPA II DE LUZIÂNIA", "UPA I JARDIM INGÁ", "SAMU"]:
            if unit_col_base in base_daily.columns:
                tmp = base_daily[["Data", unit_col_base]].copy()
                tmp.columns = ["dia_dt", "valor"]
                tmp["dia"] = tmp["dia_dt"].dt.date
                tmp["unidade_base"] = unit_col_base
                rows.append(tmp[["dia", "unidade_base", "valor"]])
        if rows:
            base_long = pd.concat(rows, ignore_index=True)
            base_long["valor"] = pd.to_numeric(base_long["valor"], errors="coerce").fillna(0)
            # comparação apenas total diário (sem depender do nome de unidade)
            new_total = agg_new.groupby("dia", as_index=False)[qty_col].sum().rename(columns={qty_col:"novo_total"})
            base_total = base_long.groupby("dia", as_index=False)["valor"].sum().rename(columns={"valor":"base_total"})
            cmp = new_total.merge(base_total, on="dia", how="inner")
            cmp["dif"] = cmp["novo_total"] - cmp["base_total"]
            cross["status"] = "partial"
            cross["matched_days"] = int(len(cmp))
            cross["avg_abs_diff"] = float(cmp["dif"].abs().mean()) if len(cmp) else None
            cross["max_abs_diff"] = float(cmp["dif"].abs().max()) if len(cmp) else None

report["cross"] = cross

with open(out_path, "w", encoding="utf-8") as f:
    json.dump(report, f, ensure_ascii=False, indent=2)
print(out_path)
