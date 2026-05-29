from openpyxl import load_workbook
import os

file_path = "PowerBI/data_raw/producao_consolidada_marco_2026_celk.xlsx"
wb = load_workbook(file_path, data_only=True)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Total rows: {ws.max_row}")
print(f"Total cols: {ws.max_column}")

print("\n" + "="*80)
print("First 20 rows of data:")
print("="*80)
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i <= 20:
        print(f"Row {i}: {row}")
    else:
        break

print("\n" + "="*80)
print("Sampling random rows to check DATE patterns:")
print("="*80)
# Get every 100th row
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i % 100 == 0 or i in [50, 150, 250, 500]:
        print(f"Row {i}: DATA={row[0]}")
